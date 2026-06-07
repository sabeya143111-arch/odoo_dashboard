"""
SWAG Product Comparison Dashboard
Version 3.0 — Ultra Premium Dark Design
"""

import io
import re
import hashlib
import time
import xmlrpc.client
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import streamlit as st

st.set_page_config(
    page_title="SWAG Dashboard",
    page_icon="◆",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:ital,wght@0,300;0,600;1,300&family=Tajawal:wght@400;700;900&family=Outfit:wght@400;500;600;700&display=swap');

/* ── RESET ─────────────────────────────────────────────── */
*,html,body,[class*="css"]{
  font-family:'Outfit','Tajawal',sans-serif;
  box-sizing:border-box;
}

/* ── APP BACKGROUND — WHITE ─────────────────────────────── */
.stApp{background:#F5F7FA !important;}
.stApp > header{background:transparent !important;}
.block-container{padding-top:0 !important;padding-bottom:0 !important;max-width:100% !important;}
.main .block-container{padding:0 !important;}

/* ── SIDEBAR — light grey ───────────────────────────────── */
section[data-testid="stSidebar"]{
  background:#FFFFFF !important;
  border-right:2px solid #E2E8F0 !important;
}
section[data-testid="stSidebar"] *{
  color:#111827 !important;
  font-weight:600 !important;
}
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3{
  color:#1A7A82 !important;
  font-family:'Outfit',sans-serif !important;
  font-size:9px !important;
  font-weight:700 !important;
  letter-spacing:4px !important;
  text-transform:uppercase !important;
}
section[data-testid="stSidebar"] input{color:#111827 !important;}

/* ── METRICS ─────────────────────────────────────────────── */
[data-testid="stMetric"]{
  background:#FFFFFF !important;
  border:1.5px solid #E2E8F0 !important;
  border-radius:10px !important;
  padding:20px 24px !important;
  box-shadow:0 1px 4px rgba(0,0,0,0.06) !important;
  transition:border-color 0.2s,box-shadow 0.2s !important;
}
[data-testid="stMetric"]:hover{
  border-color:#1A7A82 !important;
  box-shadow:0 4px 12px rgba(26,122,130,0.1) !important;
}
[data-testid="stMetricLabel"]{
  font-family:'Outfit',sans-serif !important;
  font-size:10px !important;
  font-weight:800 !important;
  letter-spacing:3px !important;
  text-transform:uppercase !important;
  color:#374151 !important;
}
[data-testid="stMetricValue"]{
  font-family:'Cormorant Garamond',serif !important;
  font-size:46px !important;
  font-weight:600 !important;
  color:#0A0A0A !important;
  line-height:1.1 !important;
}

/* ── TABS ────────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"]{
  background:#FFFFFF !important;
  border-bottom:2px solid #E2E8F0 !important;
  gap:0 !important;padding:0 !important;
}
.stTabs [data-baseweb="tab"]{
  font-family:'Outfit',sans-serif !important;
  font-size:10px !important;
  font-weight:700 !important;
  letter-spacing:2px !important;
  text-transform:uppercase !important;
  color:#6B7280 !important;
  font-size:11px !important;
  padding:14px 22px !important;
  border-radius:0 !important;
  border-bottom:2px solid transparent !important;
  background:transparent !important;
  transition:all 0.2s !important;
}
.stTabs [aria-selected="true"]{
  color:#0A0A0A !important;
  border-bottom:3px solid #1A7A82 !important;
  background:transparent !important;
}

/* ── INPUTS ──────────────────────────────────────────────── */
.stTextInput input,
.stNumberInput input,
.stTextArea textarea{
  background:#FFFFFF !important;
  border:1.5px solid #D1D5DB !important;
  border-radius:8px !important;
  color:#0A0A0A !important;
  font-family:'Outfit',sans-serif !important;
  font-size:15px !important;
  font-weight:600 !important;
  caret-color:#1A7A82 !important;
  transition:all 0.2s !important;
}
.stTextInput input:focus,
.stTextArea textarea:focus{
  border-color:#1A7A82 !important;
  box-shadow:0 0 0 3px rgba(26,122,130,0.12) !important;
}
.stTextInput input::placeholder,
.stTextArea textarea::placeholder{color:#9CA3AF !important;}
.stTextInput label,
.stNumberInput label,
.stTextArea label,
.stSelectbox label,
.stMultiSelect label{
  font-family:'Outfit',sans-serif !important;
  font-size:9px !important;
  font-weight:700 !important;
  letter-spacing:3px !important;
  text-transform:uppercase !important;
  color:#1A7A82 !important;
}

/* ── SELECT ──────────────────────────────────────────────── */
[data-baseweb="select"] div,
[data-baseweb="select"] span,
[data-baseweb="select"] input{
  background:#FFFFFF !important;
  color:#111827 !important;
  border-color:#D1D5DB !important;
  border-radius:8px !important;
  font-family:'Outfit',sans-serif !important;
  font-size:13px !important;
  font-weight:600 !important;
}
/* Dropdown popover / listbox */
[data-baseweb="popover"],
[data-baseweb="popover"] *,
[role="listbox"],
[role="listbox"] *,
[role="option"],
[data-baseweb="menu"],
[data-baseweb="menu"] *,
ul[role="listbox"],
ul[role="listbox"] li{
  background:#FFFFFF !important;
  color:#111827 !important;
  font-family:'Outfit',sans-serif !important;
  font-size:13px !important;
  font-weight:600 !important;
}
/* Hover state on option */
[role="option"]:hover,
[data-baseweb="menu"] li:hover{
  background:#EEF9FA !important;
  color:#1A7A82 !important;
}
/* Selected option */
[aria-selected="true"]{
  background:#EEF9FA !important;
  color:#1A7A82 !important;
  font-weight:700 !important;
}
/* Multiselect tag */
[data-baseweb="tag"]{
  background:#EEF9FA !important;
  color:#1A7A82 !important;
  border-radius:100px !important;
  border:1.5px solid #1A7A82 !important;
  font-weight:700 !important;
}

/* ── BUTTONS ─────────────────────────────────────────────── */
.stButton button{
  font-family:'Outfit',sans-serif !important;
  font-size:10px !important;
  font-weight:700 !important;
  letter-spacing:2px !important;
  text-transform:uppercase !important;
  border-radius:100px !important;
  transition:all 0.2s !important;
}
.stButton button[kind="primary"],
.stFormSubmitButton button{
  background:#1A7A82 !important;
  color:#FFFFFF !important;
  border:none !important;
  padding:10px 28px !important;
  border-radius:100px !important;
  box-shadow:0 2px 8px rgba(26,122,130,0.25) !important;
}
.stButton button[kind="primary"]:hover,
.stFormSubmitButton button:hover{
  background:#145F66 !important;
  transform:translateY(-1px) !important;
  box-shadow:0 4px 16px rgba(26,122,130,0.35) !important;
}
.stButton button[kind="secondary"]{
  background:#FFFFFF !important;
  color:#1A7A82 !important;
  border:1.5px solid #1A7A82 !important;
  border-radius:100px !important;
}
.stButton button[kind="secondary"]:hover{
  background:#F0FAFA !important;
}

/* ── DOWNLOAD BUTTONS ────────────────────────────────────── */
.stDownloadButton button{
  background:#FFFFFF !important;
  color:#1A7A82 !important;
  border:1.5px solid #D1D5DB !important;
  border-radius:100px !important;
  font-family:'Outfit',sans-serif !important;
  font-size:9px !important;
  font-weight:700 !important;
  letter-spacing:2px !important;
  text-transform:uppercase !important;
  padding:6px 16px !important;
  transition:all 0.2s !important;
}
.stDownloadButton button:hover{
  border-color:#1A7A82 !important;
  background:#F0FAFA !important;
}

/* ── TOGGLE / RADIO / CHECKBOX ───────────────────────────── */
.stToggle label,
.stCheckbox label,
.stRadio label,
div[data-testid="stRadio"] p{
  color:#111827 !important;
  font-family:'Outfit',sans-serif !important;
  font-size:11px !important;
  font-weight:700 !important;
  letter-spacing:0.5px !important;
}
[data-testid="stToggle"] span[data-checked="true"]{background:#1A7A82 !important;}

/* ── EXPANDER ────────────────────────────────────────────── */
[data-testid="stExpander"]{
  background:#FFFFFF !important;
  border:1.5px solid #E2E8F0 !important;
  border-radius:10px !important;
  box-shadow:0 1px 3px rgba(0,0,0,0.05) !important;
}
[data-testid="stExpander"] summary,
[data-testid="stExpander"] summary p{
  color:#1A7A82 !important;
  font-family:'Outfit',sans-serif !important;
  font-size:10px !important;
  font-weight:700 !important;
  letter-spacing:2px !important;
  text-transform:uppercase !important;
}

/* ── FILE UPLOADER ───────────────────────────────────────── */
[data-testid="stFileUploader"]{
  background:#FAFAFA !important;
  border:2px dashed #D1D5DB !important;
  border-radius:10px !important;
}
[data-testid="stFileUploader"] p,
[data-testid="stFileUploader"] span{
  color:#6B7280 !important;
  font-family:'Outfit',sans-serif !important;
  font-size:12px !important;
  font-weight:500 !important;
}

/* ── PROGRESS BAR ────────────────────────────────────────── */
[data-testid="stProgressBar"]>div{
  background:linear-gradient(90deg,#1A7A82,#D4A84B) !important;
  border-radius:100px !important;
}
[data-testid="stProgressBar"]{
  background:#E2E8F0 !important;
  border-radius:100px !important;
  height:4px !important;
}

/* ── SCROLLBAR ───────────────────────────────────────────── */
::-webkit-scrollbar{width:4px;height:4px;}
::-webkit-scrollbar-track{background:#F5F7FA;}
::-webkit-scrollbar-thumb{background:#1A7A82;border-radius:100px;}

/* ── DIVIDER ─────────────────────────────────────────────── */
hr{
  border:none !important;
  height:2px !important;
  background:#E2E8F0 !important;
  margin:20px 0 !important;
}

/* ── CAPTION ─────────────────────────────────────────────── */
.stCaption,[data-testid="stCaptionContainer"] p{
  color:#6B7280 !important;
  font-family:'Outfit',sans-serif !important;
  font-size:11px !important;
  font-weight:600 !important;
  letter-spacing:0.5px !important;
}

/* ── TEXT ────────────────────────────────────────────────── */
h1,h2,h3,h4,h5,h6{
  color:#0A0A0A !important;
  font-family:'Outfit',sans-serif !important;
  font-weight:800 !important;
}
.stMarkdown p,.stMarkdown li{color:#111827 !important;font-weight:600 !important;font-size:14px !important;}
p,div,span,label{color:#111827;font-weight:500;}

/* ── NUMBER INPUT ────────────────────────────────────────── */
.stNumberInput button{
  color:#1A7A82 !important;
  background:#F0FAFA !important;
  border-color:#D1D5DB !important;
}

/* ── DATA TABLE ──────────────────────────────────────────── */
[data-testid="stDataFrame"] th{
  background:#1A7A82 !important;
  color:#FFFFFF !important;
  font-family:'Outfit',sans-serif !important;
  font-size:11px !important;
  font-weight:700 !important;
  letter-spacing:1px !important;
  text-transform:uppercase !important;
}
[data-testid="stDataFrame"] td{
  font-family:'Outfit',sans-serif !important;
  font-size:13px !important;
  font-weight:600 !important;
  color:#0A0A0A !important;
}
[data-testid="stDataFrame"]{
  border:1.5px solid #E2E8F0 !important;
  border-radius:10px !important;
  overflow:hidden !important;
}

/* ── CUSTOM BANNERS ──────────────────────────────────────── */
.info-banner{
  background:#EEF9FA;
  border-left:3px solid #1A7A82;
  padding:10px 16px;
  font-family:'Outfit',sans-serif;
  font-size:10px;
  font-weight:600;
  letter-spacing:1.5px;
  text-transform:uppercase;
  color:#1A7A82;
  margin-bottom:10px;
  border-radius:0 6px 6px 0;
}
.warn-banner{
  background:#FFFBEB;
  border-left:3px solid #D97706;
  padding:10px 16px;
  font-family:'Outfit',sans-serif;
  font-size:10px;
  font-weight:600;
  letter-spacing:1.5px;
  text-transform:uppercase;
  color:#92400E;
  margin-bottom:10px;
  border-radius:0 6px 6px 0;
}
.alert-missing{
  background:#FEF2F2;
  border-left:3px solid #EF4444;
  padding:10px 16px;
  font-family:'Outfit',sans-serif;
  font-size:10px;
  font-weight:600;
  letter-spacing:1.5px;
  text-transform:uppercase;
  color:#991B1B;
  margin-bottom:8px;
  border-radius:0 6px 6px 0;
}
.alert-price{
  background:#FFFBEB;
  border-left:3px solid #F59E0B;
  padding:10px 16px;
  font-family:'Outfit',sans-serif;
  font-size:10px;
  font-weight:600;
  letter-spacing:1.5px;
  text-transform:uppercase;
  color:#92400E;
  margin-bottom:8px;
  border-radius:0 6px 6px 0;
}

/* ── SECTION TAG ─────────────────────────────────────────── */
.section-tag{
  font-family:'Outfit',sans-serif;
  font-size:9px;
  font-weight:700;
  letter-spacing:4px;
  text-transform:uppercase;
  color:#1A7A82;
  margin:20px 0 12px 0;
  display:flex;
  align-items:center;
  gap:10px;
}
.section-tag::before{
  content:'';width:20px;height:2px;background:#1A7A82;
}

/* ── HERO SECTION ────────────────────────────────────────── */
.hero-section{
  padding:40px 0 28px;
  border-bottom:2px solid #E2E8F0;
  position:relative;
  overflow:hidden;
}
.hero-title{
  font-family:'Cormorant Garamond',serif !important;
  font-size:52px !important;
  font-weight:600 !important;
  color:#111827 !important;
  letter-spacing:-1px;
  margin-bottom:0;
}
.hero-title em{color:#1A7A82;font-style:normal;}

/* ── STATUS TAGS ─────────────────────────────────────────── */
.status-online{
  background:#D1FAE5;color:#065F46;
  padding:3px 10px;border-radius:100px;
  font-size:9px;font-weight:700;letter-spacing:1px;
}
.status-offline{
  background:#FEE2E2;color:#991B1B;
  padding:3px 10px;border-radius:100px;
  font-size:9px;font-weight:700;letter-spacing:1px;
}

/* ── SWAG TABLE ──────────────────────────────────────────── */
.swag-tbl{
  width:100%;
  border-collapse:collapse;
  font-family:'Outfit',sans-serif;
  font-size:13px;
}
.swag-tbl th{
  background:#1A7A82;
  color:#FFFFFF;
  font-size:9px;font-weight:700;
  letter-spacing:2px;text-transform:uppercase;
  padding:12px 16px;text-align:left;
}
.swag-tbl td{
  padding:10px 16px;
  color:#111827;
  font-weight:500;
  border-bottom:1px solid #F3F4F6;
}
.swag-tbl tbody tr:nth-child(even) td{background:#F9FAFB;}
.swag-tbl tbody tr:hover td{background:#EEF9FA;}
.swag-tbl tbody tr.rl td{background:#FFFBEB;}

/* ── SIZE TABLE ──────────────────────────────────────────── */
.sz-tbl{width:100%;border-collapse:collapse;font-family:'Outfit',sans-serif;font-size:12px;}
.sz-tbl th{background:#1A7A82;color:#fff;font-size:9px;font-weight:700;
  letter-spacing:2px;text-transform:uppercase;padding:10px 12px;text-align:center;}
.sz-tbl td{padding:8px 12px;color:#111827;font-weight:500;
  border-bottom:1px solid #F3F4F6;text-align:center;}
.sz-tbl tbody tr:nth-child(even) td{background:#F9FAFB;}
.sz-tbl tbody tr:hover td{background:#EEF9FA;}

/* ── SNAPSHOT (Today widget) ─────────────────────────────── */
.snap-card{
  background:#FFFFFF;
  border:2px solid #E2E8F0;
  border-radius:12px;
  padding:18px;
  transition:border-color 0.2s,box-shadow 0.2s;
}
.snap-card:hover{
  border-color:#1A7A82;
  box-shadow:0 4px 12px rgba(26,122,130,0.12);
}
.sc-label{
  font-family:'Outfit',sans-serif;font-size:9px;font-weight:800;
  letter-spacing:3px;text-transform:uppercase;color:#374151;margin-bottom:8px;
}
.sc-val{
  font-family:'Cormorant Garamond',serif;font-size:36px;
  font-weight:700;color:#0A0A0A;line-height:1;margin-bottom:3px;
}
.sc-val.teal{color:#1A7A82;}
.sc-val.gold{color:#B45309;}
.sc-val.red-v{color:#DC2626;}
.sc-sub{font-family:'Outfit',sans-serif;font-size:11px;font-weight:600;color:#374151;}
.snap-divider{height:2px;margin:24px 0 20px;background:#E2E8F0;}

/* ── SYSTEM PILLS (snapshot) ─────────────────────────────── */
.sp{display:flex;align-items:center;gap:7px;border-radius:100px;padding:6px 14px;}
.sp-online{background:#D1FAE5;border:1.5px solid #6EE7B7;}
.sp-offline{background:#F3F4F6;border:1.5px solid #E5E7EB;}
.sp-error{background:#FEE2E2;border:1.5px solid #FCA5A5;}
.sp-nodata{background:#FFFBEB;border:1.5px solid #FDE68A;}
.sd{width:8px;height:8px;border-radius:50%;flex-shrink:0;}
.sd-online{background:#059669;}
.sd-offline{background:#9CA3AF;}
.sd-error{background:#EF4444;}
.sd-nodata{background:#D97706;}
.sn{font-family:'Outfit',sans-serif;font-size:11px;font-weight:700;letter-spacing:0.5px;}
.sn-online{color:#065F46;}.sn-offline{color:#6B7280;}
.sn-error{color:#991B1B;}.sn-nodata{color:#92400E;}
.sb{font-family:'Outfit',sans-serif;font-size:8px;font-weight:700;
  letter-spacing:1.5px;text-transform:uppercase;padding:2px 8px;border-radius:100px;}
.sb-online{background:#A7F3D0;color:#065F46;}
.sb-offline{background:#E5E7EB;color:#6B7280;}
.sb-error{background:#FECACA;color:#991B1B;}
.sb-nodata{background:#FDE68A;color:#92400E;}

/* ── SEASON COMPARISON PILLS ─────────────────────────────── */
.sc-pill-ok{background:#FFFFFF;border:1.5px solid #1A7A82;border-radius:10px;padding:12px;text-align:center;box-shadow:0 1px 4px rgba(26,122,130,0.1);}
.sc-pill-no{background:#F9FAFB;border:1.5px solid #E2E8F0;border-radius:10px;padding:12px;text-align:center;}
.sc-pill-label{font-size:8px;font-weight:700;letter-spacing:2px;text-transform:uppercase;margin-bottom:4px;}
.sc-pill-field{font-size:9px;color:#6B7280;font-weight:500;}
.sc-pill-count{font-family:'Cormorant Garamond',serif;font-size:24px;font-weight:600;color:#111827;}
.sc-pill-unit{font-size:8px;color:#9CA3AF;font-weight:500;}

/* ── LAST RUN CARD ───────────────────────────────────────── */
.snap-last{
  display:flex;align-items:center;gap:14px;flex-wrap:wrap;
  background:#EEF9FA;
  border:1.5px solid #1A7A82;
  border-radius:8px;
  padding:12px 18px;margin-top:12px;
}
.sl-label{font-family:'Outfit',sans-serif;font-size:8px;font-weight:700;
  letter-spacing:3px;text-transform:uppercase;color:#1A7A82;flex-shrink:0;}
.sl-val{font-family:'Cormorant Garamond',serif;font-size:18px;
  font-weight:600;color:#111827;letter-spacing:1px;}
.sl-meta{font-family:'Outfit',sans-serif;font-size:10px;font-weight:500;color:#6B7280;}
.sl-ago{font-family:'Outfit',sans-serif;font-size:10px;font-weight:500;color:#9CA3AF;}
.sl-rows{font-family:'Outfit',sans-serif;font-size:11px;color:#1A7A82;font-weight:700;}
</style>
""", unsafe_allow_html=True)

# TABLE CSS injected separately so it's reusable
_TABLE_CSS = """<style>
.swag-wrap{width:100%;overflow-x:auto;border:2px solid #E2E8F0;border-radius:10px;overflow:hidden;margin-bottom:8px;}
.swag-tbl{width:100%;border-collapse:collapse;font-family:'Outfit','Tajawal',sans-serif;}
.swag-tbl thead tr{background:#1A7A82;}
.swag-tbl thead th{
  color:#FFFFFF;font-family:'Outfit',sans-serif;
  font-size:10px;letter-spacing:2px;text-transform:uppercase;font-weight:700;
  padding:14px 16px;text-align:center;white-space:nowrap;
}
.swag-tbl tbody tr{border-bottom:1px solid #F3F4F6;transition:background 0.15s;}
.swag-tbl tbody tr:nth-child(even) td{background:#F9FAFB;}
.swag-tbl tbody tr:last-child{border-bottom:none;}
.swag-tbl tbody tr:hover td{background:#EEF9FA !important;}
.swag-tbl tbody td{
  padding:12px 16px;text-align:center;
  font-size:13px;font-weight:600;
  color:#111827;
}
.swag-tbl tbody td.cf{
  font-family:'Outfit',monospace;font-size:12px;letter-spacing:0.5px;
  color:#1A7A82;font-weight:700;
  border-right:2px solid #E2E8F0;
  text-align:left;
}
.swag-tbl tbody tr.rl td{background:#FFFBEB !important;color:#92400E;}
.swag-tbl tbody td.na-cell{
  color:#DC2626 !important;
  font-weight:700 !important;
  font-size:13px !important;
}
.swag-tbl tbody td.zero-cell{
  color:#9CA3AF !important;
  font-weight:600 !important;
}
</style>"""

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
SYSTEM_KEYS = ["SWAG", "STOCK", "LAROUCHE", "DIFFC", "FASHIONLIMITS"]

# ─────────────────────────────────────────────────────────────────────────────
# LANGUAGE
# ─────────────────────────────────────────────────────────────────────────────
def get_lang():
    return st.session_state.get("lang", "EN")

def t(en, ar):
    return ar if get_lang() == "AR" else en

def get_system_name(key):
    cfg = get_system_config(key) or {}
    return cfg.get("name_ar", cfg.get("name", key)) if get_lang() == "AR" else cfg.get("name", key)

def translate_system_names(df):
    if df is None or df.empty:
        return df
    sys_col = t("System", "النظام")
    if sys_col not in df.columns:
        return df
    key_to_name = {k: get_system_name(k) for k in SYSTEM_KEYS}
    out = df.copy()
    out[sys_col] = out[sys_col].map(lambda v: key_to_name.get(v, v))
    return out

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
_DEF = {
    "authenticated": False, "user_email": "", "lang": "EN",
    "last_run": None, "total_df": None, "branch_df": None,
    "transfers_df": None, "reorder_df": None, "sys_stats": {},
    "search_exact": False, "low_stock_thresh": 5,
    "show_transfers": False, "show_reorder": False,
    "reorder_target_days": 30,
    "reorder_point": 10,
    "pdf_codes": None, "pdf_mode": "total",
    "so_analytics_df": None, "so_last_model": "", "so_last_system": "",
}
for k, v in _DEF.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ─────────────────────────────────────────────────────────────────────────────
# SESSION LOGIN RESTORE
# ─────────────────────────────────────────────────────────────────────────────
_COOKIE_SECRET = "swag_2025_secure"

def _make_token(email):
    return hashlib.sha256(f"{_COOKIE_SECRET}_{email}".encode()).hexdigest()[:32]

def _verify_token(email, token):
    return bool(email and token and token == _make_token(email))

def restore_session():
    if st.session_state.get("authenticated"):
        return
    try:
        params = st.query_params
        email = params.get("u", "")
        token = params.get("t", "")
        if email and token and _verify_token(email, token):
            st.session_state.authenticated = True
            st.session_state.user_email = email
    except Exception:
        pass

# ─────────────────────────────────────────────────────────────────────────────
# XML-RPC
# ─────────────────────────────────────────────────────────────────────────────
# CONFIG HELPER — normalises key aliases + strips trailing /odoo from URL
# ─────────────────────────────────────────────────────────────────────────────
_KEY_ALIASES: dict = {
    # Both spellings map to the canonical key stored in secrets
    "FASHION_LIMITS" : "FASHIONLIMITS",
    "FASHIONLIMITS"  : "FASHIONLIMITS",
}

def _canonical_key(key: str) -> str:
    """Return the canonical secrets key for any alias."""
    return _KEY_ALIASES.get(key, key)

def get_system_config(key: str) -> dict | None:
    """
    Fetch config from st.secrets, trying canonical key then known aliases.
    Strips a trailing '/odoo' from the URL so XML-RPC proxy works correctly:
      proxy builds  url + /xmlrpc/2/common
      so url must be  https://host.swag.com.sa  (no /odoo suffix)
    Returns dict or None.
    """
    canonical = _canonical_key(key)
    # Try canonical first, then the raw key as fallback
    cfg = st.secrets.get(canonical) or st.secrets.get(key)
    if not cfg:
        return None
    cfg = dict(cfg)                         # make mutable copy
    url = str(cfg.get("url", "")).rstrip("/")
    if url.endswith("/odoo"):
        url = url[: -len("/odoo")]
    cfg["url"] = url
    return cfg

# ─────────────────────────────────────────────────────────────────────────────
@st.cache_resource
def _proxy(url, ep):
    return xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/{ep}", allow_none=True)

@st.cache_data(ttl=28800, show_spinner=False)
def _auth(url, db, user, api_key):
    """
    Returns structured dict:
      {"ok": True,  "uid": <int>}
      {"ok": False, "error": "<TAG>: <detail>"}
    Tags: NO_RESPONSE | BAD_CREDENTIALS | AUTH_EXCEPTION
    """
    try:
        uid = _proxy(url, "common").authenticate(db, user, api_key, {})
        if uid:
            return {"ok": True, "uid": uid}
        return {"ok": False, "error": "BAD_CREDENTIALS: uid=False — check user/api_key"}
    except ConnectionRefusedError as e:
        return {"ok": False, "error": f"NO_RESPONSE: {e}"}
    except Exception as e:
        return {"ok": False, "error": f"AUTH_EXCEPTION: {e}"}

def _auth_uid(url, db, user, api_key):
    """Backward-compat wrapper — returns uid int or None."""
    r = _auth(url, db, user, api_key)
    return r["uid"] if r["ok"] else None

def _x(url, db, uid, key, model, method, domain, kw):
    return _proxy(url, "object").execute_kw(db, uid, key, model, method, domain, kw)

# ─────────────────────────────────────────────────────────────────────────────
# DOMAIN BUILDER
# ─────────────────────────────────────────────────────────────────────────────
def _domain(codes, exact):
    if exact:
        return [["default_code", "in", codes]]
    if len(codes) == 1:
        return [["default_code", "=like", f"{codes[0]}%"]]
    parts = [["default_code", "=like", f"{c}%"] for c in codes]
    return ["|"] * (len(parts) - 1) + parts

# ─────────────────────────────────────────────────────────────────────────────
# PDF PARSING
# ─────────────────────────────────────────────────────────────────────────────
_RE_BRACKET = re.compile(r'\[([A-Za-z0-9\-_()]{3,30})\]')
_RE_SR_LINE = re.compile(
    r'(?:^|\s)([A-Z]{2,6}\d+(?:-\d+)?(?:-[A-Z0-9()]{1,10})?)\s+.{0,80}?\d+\.?\d*\s+SR',
    re.MULTILINE)
_RE_GENERAL = re.compile(
    r'\b([A-Z]{2,6}\d+(?:-\d+)?(?:-[A-Z0-9]{1,4})?(?:\([^)]{1,15}\))?)\b')
_EXCLUDE = frozenset([
    'SR','VAT','TAX','PCS','QTY','NO','REF','INV','PO','SO',
    'DO','ID','EN','AR','PDF','AED','SAR','USD','KWD','OMR',
    'BHD','JOD','EGP','TRY'
])

def _valid(code):
    c = code.strip().upper()
    return (bool(re.search(r'[A-Z]', c)) and bool(re.search(r'\d', c))
            and 4 <= len(c) <= 25 and c not in _EXCLUDE)

def extract_base_model(code):
    code = re.sub(r'\([^)]*\)', '', code)
    for s in ['-2XL','-3XL','-4XL','-XXL','-XL','-L','-M','-S','-XS','-2X','-3X']:
        if code.upper().endswith(s.upper()):
            code = code[:-len(s)]; break
    return re.sub(r'-\d{2,3}$', '', code).strip()

def get_unique_base_models(raw):
    seen, out = set(), []
    for item in raw:
        b = extract_base_model(item["code"])
        if b and b not in seen:
            seen.add(b)
            out.append({"sequence": item["sequence"], "code": b})
    return out

@st.cache_data(show_spinner=False)
def parse_invoice_pdf_cached(file_bytes):
    try:
        from pypdf import PdfReader
    except ImportError:
        return []
    text = ""
    for page in PdfReader(io.BytesIO(file_bytes)).pages:
        text += (page.extract_text() or "") + "\n"
    if not text.strip():
        return []
    raw = (_RE_BRACKET.findall(text)
           + [m.group(1) for m in _RE_SR_LINE.finditer(text)]
           + _RE_GENERAL.findall(text))
    seen, out = set(), []
    seq = 1
    for c in raw:
        u = c.strip().upper()
        if _valid(u) and u not in seen:
            seen.add(u)
            out.append({"sequence": seq, "code": u})
            seq += 1
    return out

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _style_worksheet(ws, df_clean, lang="EN"):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule
    if lang == "AR":
        ws.sheet_view.rightToLeft = True
    # White professional theme
    hdr_fill  = PatternFill("solid", fgColor="1A7A82")   # teal header
    hdr_font  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin      = Side(border_style="thin", color="D0D0D0")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill  = PatternFill("solid", fgColor="F2FAFA")   # very light teal
    zero_fill = PatternFill("solid", fgColor="FFF3CD")   # light amber for zero
    zero_font = Font(color="856404", bold=True, name="Calibri")
    norm_font = Font(name="Calibri", size=10, color="333333")
    num_align = Alignment(horizontal="right",  vertical="center")
    ctr_align = Alignment(horizontal="center", vertical="center")
    tot_fill  = PatternFill("solid", fgColor="E8F5F6")   # light teal total row
    tot_font  = Font(bold=True, name="Calibri", color="1A7A82")
    max_row   = ws.max_row
    max_col   = ws.max_column
    ws.row_dimensions[1].height = 28
    for col_num in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = border
    col_names = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    on_hand_col = None
    for i, name in enumerate(col_names, 1):
        if name in ("On Hand", "متوفر"): on_hand_col = i
    for row in ws.iter_rows(min_row=2, max_row=max_row):
        is_zero = False
        if on_hand_col:
            val = ws.cell(row=row[0].row, column=on_hand_col).value
            is_zero = (val is None or
                       str(val).strip() in ['0','Not Available','غير متوفر','—','-','']
                       or val == 0)
        for cell in row:
            cell.border = border
            cell.font   = zero_font if is_zero else norm_font
            if is_zero:              cell.fill = zero_fill
            elif cell.row % 2 == 0: cell.fill = alt_fill
            cell.alignment = num_align if isinstance(cell.value, (int, float)) else ctr_align
        ws.row_dimensions[row[0].row].height = 18
    for col_num in range(1, max_col + 1):
        col_letter = get_column_letter(col_num)
        max_len = 0
        for r in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in r:
                if cell.value: max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    if on_hand_col and max_row > 1:
        col_letter = get_column_letter(on_hand_col)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{max_row}",
            DataBarRule(start_type="min", end_type="max", color="1A7A82"))
    total_row = max_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font      = tot_font
    ws.cell(row=total_row, column=1).fill      = tot_fill
    ws.cell(row=total_row, column=1).alignment = ctr_align
    if on_hand_col:
        col = get_column_letter(on_hand_col)
        ws.cell(row=total_row, column=on_hand_col,
                value=f"=SUM({col}2:{col}{max_row})")
        ws.cell(row=total_row, column=on_hand_col).font      = tot_font
        ws.cell(row=total_row, column=on_hand_col).fill      = tot_fill
        ws.cell(row=total_row, column=on_hand_col).alignment = ctr_align
    ws.row_dimensions[total_row].height = 20
    ws.sheet_properties.tabColor = "4AACB4"
    footer_row = total_row + 2
    ws.cell(row=footer_row, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  SWAG Dashboard")
    ws.cell(row=footer_row, column=1).font = Font(italic=True, color="1A7A82", size=9, name="Calibri")
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows       = "1:1"
    ws.sheet_view.zoomScale   = 85

def to_csv(df):
    return df.drop(columns=["_status"], errors="ignore").to_csv(index=False).encode("utf-8-sig")

def to_excel(df):
    lang  = st.session_state.get('lang', 'EN')
    buf   = io.BytesIO()
    clean = df.drop(columns=['_status','_avail'], errors='ignore').copy()

    # Clean product names — remove [CODE] prefix from display_name
    if "Product" in clean.columns:
        def _strip_bracket(v):
            s=str(v or "").strip()
            if s.startswith("[") and "]" in s:
                end=s.index("]"); part=s[end+1:].strip()
                if part: return part
            return s
        clean["Product"]=clean["Product"].apply(_strip_bracket)

    oh    = 'On Hand' if 'On Hand' in clean.columns else ('متوفر' if 'متوفر' in clean.columns else None)
    if oh:
        na = 'غير متوفر' if lang == 'AR' else 'Not Available'
        clean[oh] = clean[oh].apply(
            lambda x: na if (pd.isna(x) or str(x).strip() in ['0','']) or x == 0 else x)
    desired = [t("Model Code","رمز الموديل"), t("System","النظام"),
               t("Branch","الفرع"), t("Location","الموقع"),
               t("Sale Price","سعر البيع"), t("On Hand","متوفر")]
    ordered  = [c for c in desired if c in clean.columns]
    remaining = [c for c in clean.columns if c not in ordered]
    clean    = clean[ordered + remaining]
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        clean.to_excel(w, index=False, sheet_name='Data')
        _style_worksheet(w.sheets['Data'], clean, lang=lang)
    return buf.getvalue()

def to_excel_bulk(df):
    lang    = st.session_state.get("lang", "EN")
    buf     = io.BytesIO()
    sys_col = t("System", "النظام")
    desired = [t("Model Code","رمز الموديل"), t("System","النظام"),
               t("Branch","الفرع"), t("Location","الموقع"),
               t("Sale Price","سعر البيع"), t("On Hand","متوفر")]
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        def _ws(data, name):
            c  = data.drop(columns=["_status"], errors="ignore").copy()
            oh = t("On Hand", "متوفر")
            if oh in c.columns:
                na = 'غير متوفر' if lang == 'AR' else 'Not Available'
                c[oh] = c[oh].apply(
                    lambda x: na if (pd.isna(x) or str(x).strip() in ['0','']) or x == 0 else x)
            ordered   = [col for col in desired if col in c.columns]
            remaining = [col for col in c.columns if col not in ordered]
            c = c[ordered + remaining]
            c.to_excel(w, index=False, sheet_name=name[:31])
            _style_worksheet(w.sheets[name[:31]], c, lang=lang)
        _ws(df, t("All Systems", "كل الأنظمة"))
        if sys_col in df.columns:
            for key in SYSTEM_KEYS:
                nm  = get_system_name(key)
                sub = df[df[sys_col] == nm]
                if not sub.empty:
                    _ws(sub, nm)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# PURCHASE SUMMARY
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=3600, show_spinner=False)
def get_purchase_summary_by_model(model_codes_tuple, date_from, date_to, system_key="SWAG"):
    empty = pd.DataFrame(columns=["Model Code", "Purchase Qty"])
    cfg = get_system_config(system_key)
    if not cfg: return empty
    ar = _auth(cfg["url"], cfg["db"], cfg["user"], cfg["api_key"])
    if not ar["ok"]: return empty
    uid = ar["uid"]; u = cfg["url"]; db = cfg["db"]; ak = cfg["api_key"]
    try:
        dom = [
            ["order_id.state", "in", ["purchase", "done"]],
            ["order_id.date_order", ">=", f"{date_from} 00:00:00"],
            ["order_id.date_order", "<=", f"{date_to} 23:59:59"],
        ]
        if model_codes_tuple:
            dom.append(["product_id.default_code", "in", list(model_codes_tuple)])
        lines = _x(u, db, uid, ak, "purchase.order.line", "search_read", [dom],
                   {"fields": ["product_id", "product_qty"], "limit": 10000, "order": "id desc"})
        if not lines: return empty
        pids = list({l["product_id"][0] for l in lines if isinstance(l.get("product_id"), list)})
        prods = _x(u, db, uid, ak, "product.product", "search_read",
                   [[["id", "in", pids]]],
                   {"fields": ["id", "default_code"], "limit": len(pids) + 10})
        pmap = {p["id"]: p for p in prods}
        agg = {}
        for line in lines:
            pid  = line["product_id"][0] if isinstance(line.get("product_id"), list) else None
            mc   = pmap.get(pid, {}).get("default_code", "").strip()
            if not mc: continue
            agg[mc] = agg.get(mc, 0) + float(line.get("product_qty") or 0)
        if not agg: return empty
        df = pd.DataFrame([{"Model Code": mc, "Purchase Qty": qty} for mc, qty in agg.items()])
        return df.groupby("Model Code", as_index=False)["Purchase Qty"].sum()
    except Exception:
        return empty

# ─────────────────────────────────────────────────────────────────────────────
# PURCHASE HISTORY
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=1800, show_spinner=False)
def fetch_swag_purchase_history(model_code, date_from, date_to, system_key="SWAG"):
    cols = ["Date","PO","Vendor","Brand Category","Category",
            "Model Code","Product","Qty","Unit Price","Subtotal"]
    empty = pd.DataFrame(columns=cols)
    cfg = get_system_config(system_key)
    if not cfg: return empty
    ar = _auth(cfg["url"], cfg["db"], cfg["user"], cfg["api_key"])
    if not ar["ok"]: return empty
    uid = ar["uid"]; u = cfg["url"]; db = cfg["db"]; ak = cfg["api_key"]
    try:
        dom = [
            ["order_id.state", "in", ["purchase", "done"]],
            ["order_id.date_order", ">=", f"{date_from} 00:00:00"],
            ["order_id.date_order", "<=", f"{date_to} 23:59:59"],
        ]
        if model_code and model_code.strip():
            dom.append(["product_id.default_code", "=", model_code.strip()])
        lines = _x(u, db, uid, ak, "purchase.order.line", "search_read", [dom],
                   {"fields": ["order_id","product_id","product_qty","price_unit"],
                    "limit": 5000, "order": "order_id desc"})
        if not lines: return empty
        oids = list({l["order_id"][0] for l in lines if isinstance(l.get("order_id"), list)})
        pids = list({l["product_id"][0] for l in lines if isinstance(l.get("product_id"), list)})
        orders  = _x(u, db, uid, ak, "purchase.order", "search_read",
                     [[["id","in",oids]]],
                     {"fields":["id","name","partner_id","date_order"],"limit":len(oids)+10})
        omap    = {o["id"]: o for o in orders}
        prods   = _x(u, db, uid, ak, "product.product", "search_read",
                     [[["id","in",pids]]],
                     {"fields":["id","default_code","display_name","categ_id","product_tmpl_id"],
                      "limit":len(pids)+10})
        pmap    = {p["id"]: p for p in prods}
        tids    = list({p["product_tmpl_id"][0] for p in prods
                        if isinstance(p.get("product_tmpl_id"), list)})
        tmap    = {}
        if tids:
            try:
                tmpls = _x(u, db, uid, ak, "product.template", "search_read",
                           [[["id","in",tids]]],
                           {"fields":["id","x_brand_category_id"],"limit":len(tids)+10})
                tmap = {t_["id"]: t_ for t_ in tmpls}
            except Exception:
                pass
        rows = []
        for line in lines:
            oid   = line["order_id"][0] if isinstance(line.get("order_id"), list) else None
            pid   = line["product_id"][0] if isinstance(line.get("product_id"), list) else None
            order = omap.get(oid, {})
            prod  = pmap.get(pid, {})
            raw_d = order.get("date_order") or ""
            try:    ds = datetime.strptime(raw_d, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
            except: ds = raw_d[:10] if raw_d else "—"
            partner = order.get("partner_id")
            vendor  = partner[1] if isinstance(partner, list) else (str(partner) if partner else "—")
            categ   = prod.get("categ_id")
            category = categ[1] if isinstance(categ, list) else (str(categ) if categ else "")
            brand_category = ""
            tr = prod.get("product_tmpl_id")
            if isinstance(tr, list) and tr:
                tmpl = tmap.get(tr[0], {})
                bc   = tmpl.get("x_brand_category_id")
                if isinstance(bc, list): brand_category = bc[1] if len(bc) > 1 else ""
                elif bc: brand_category = str(bc)
            qty   = float(line.get("product_qty") or 0)
            price = float(line.get("price_unit") or 0)
            rows.append({
                "Date": ds, "PO": order.get("name") or "—",
                "Vendor": vendor, "Brand Category": brand_category,
                "Category": category,
                "Model Code": prod.get("default_code") or "",
                "Product": prod.get("display_name") or "",
                "Qty": qty, "Unit Price": price,
                "Subtotal": round(qty * price, 2),
            })
        if not rows: return empty
        return pd.DataFrame(rows).sort_values(by="Date", ascending=False).reset_index(drop=True)
    except Exception:
        return empty

# ─────────────────────────────────────────────────────────────────────────────
# SALES HISTORY
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=3600, show_spinner=False)
def fetch_swag_sales_history(model_code=None, date_from=None, date_to=None, system_key="SWAG"):
    empty = pd.DataFrame(columns=[
        "Date","SO","Customer","Branch","Brand Category","Category",
        "Model Code","Product","Qty","Unit Price","Subtotal"])
    cfg = get_system_config(system_key)
    if not cfg: return empty
    ar = _auth(cfg["url"], cfg["db"], cfg["user"], cfg["api_key"])
    if not ar["ok"]: return empty
    uid = ar["uid"]; u, db, ak = cfg["url"], cfg["db"], cfg["api_key"]
    try:
        dom = [
            ["order_id.state","in",["sale","done"]],
            ["order_id.date_order",">=",f"{date_from} 00:00:00"],
            ["order_id.date_order","<=",f"{date_to} 23:59:59"],
        ]
        if model_code:
            dom.append(["product_id.default_code","=like",f"{model_code}%"])
        lines = _x(u, db, uid, ak, "sale.order.line", "search_read", [dom],
                   {"fields":["order_id","product_id","product_uom_qty","price_unit","price_subtotal"],
                    "limit":15000,"order":"order_id desc"})
        if not lines: return empty
        oids  = list({l["order_id"][0] for l in lines if isinstance(l.get("order_id"), list)})
        orders = _x(u, db, uid, ak, "sale.order", "search_read",
                    [[["id","in",oids]]],
                    {"fields":["id","name","partner_id","date_order","branch_id"],
                     "limit":len(oids)+10})
        omap  = {o["id"]: o for o in orders}
        pids  = list({l["product_id"][0] for l in lines if isinstance(l.get("product_id"), list)})
        prods = _x(u, db, uid, ak, "product.product", "search_read",
                   [[["id","in",pids]]],
                   {"fields":["id","default_code","name","categ_id","product_tmpl_id"],
                    "limit":len(pids)+10})
        pmap  = {p["id"]: p for p in prods}
        tids  = list({p["product_tmpl_id"][0] for p in prods
                      if isinstance(p.get("product_tmpl_id"), list)})
        tmap  = {}
        if tids:
            try:
                tmpls = _x(u, db, uid, ak, "product.template", "search_read",
                           [[["id","in",tids]]],
                           {"fields":["id","x_studio_brand_category"],"limit":len(tids)+10})
                tmap = {tt["id"]: tt for tt in tmpls}
            except Exception:
                pass
        rows = []
        for line in lines:
            oid = line["order_id"][0] if isinstance(line.get("order_id"), list) else None
            pid = line["product_id"][0] if isinstance(line.get("product_id"), list) else None
            o   = omap.get(oid, {})
            p   = pmap.get(pid, {})
            tr  = p.get("product_tmpl_id")
            tid = tr[0] if isinstance(tr, list) else tr
            tmpl = tmap.get(tid, {})
            br   = o.get("branch_id")
            branch = br[1] if isinstance(br, list) and len(br)>1 else (str(br) if br else "Unknown")
            cat  = p.get("categ_id")
            categ = cat[1] if isinstance(cat, list) and len(cat)>1 else (str(cat) if cat else "")
            bcr  = tmpl.get("x_studio_brand_category", "")
            brand_cat = bcr[1] if isinstance(bcr, list) and len(bcr)>1 else (str(bcr) if bcr else "")
            partner = o.get("partner_id")
            customer = partner[1] if isinstance(partner, list) and len(partner)>1 else (str(partner) if partner else "")
            pname = line.get("product_id")
            pdisplay = pname[1] if isinstance(pname, list) and len(pname)>1 else p.get("name","")
            raw_d = str(o.get("date_order",""))
            rows.append({
                "Date": raw_d[:10] if raw_d else "",
                "SO": o.get("name",""), "Customer": customer,
                "Branch": branch, "Brand Category": brand_cat or "(No Brand)",
                "Category": categ or "(No Category)",
                "Model Code": str(p.get("default_code","")).strip(),
                "Product": pdisplay,
                "Qty": float(line.get("product_uom_qty") or 0),
                "Unit Price": float(line.get("price_unit") or 0),
                "Subtotal": float(line.get("price_subtotal") or 0),
            })
        df = pd.DataFrame(rows)
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        return df.sort_values("Date", ascending=False).reset_index(drop=True)
    except Exception:
        return empty


# ─────────────────────────────────────────────────────────────────────────────
# DEAD STOCK FINDER — SWAG only (FAST VERSION)
# Strategy:
#   - Use product.product qty_available directly (no quant scan)
#   - Get last sale via a SINGLE search_read with date filter + group_by trick
#   - Chunk size 1000 to minimize API round trips
#   - Cache 10 min so repeat runs are instant
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=600, show_spinner=False)
def fetch_dead_stock(threshold_days=60, system_key="SWAG",
                     _progress=None, _status_text=None):
    """
    _progress     : st.progress() object — updated 0.0→1.0
    _status_text  : st.empty()     object — live step messages
    Returns (df, is_partial) tuple.
      is_partial=True  means timeout/error interrupted midway — partial results returned
      is_partial=False means full run completed
    """
    empty_cols = [
        "Model Code","Product","Category","On Hand",
        "Unit Price","Frozen Value (SAR)",
        "Last Sale Date","Days Since Sale","Status"
    ]
    empty = pd.DataFrame(columns=empty_cols)

    def _prog(pct, msg=""):
        if _progress:
            _progress.progress(min(pct, 1.0))
        if _status_text and msg:
            _status_text.markdown(
                f"<div class='info-banner' style='margin:4px 0;'>{msg}</div>",
                unsafe_allow_html=True)

    cfg = get_system_config(system_key)
    if not cfg:
        _prog(1.0, "No config found for system.")
        return empty, False
    ar = _auth(cfg["url"], cfg["db"], cfg["user"], cfg["api_key"])
    if not ar["ok"]:
        _prog(1.0, f"Auth failed: {ar['error']}")
        return empty, False
    uid = ar["uid"]; u, db, ak = cfg["url"], cfg["db"], cfg["api_key"]

    today  = datetime.now().date()
    cutoff = (today - timedelta(days=threshold_days)).strftime("%Y-%m-%d")
    rows         = []
    is_partial   = False
    last_sale_map= {}

    try:
        # ── Step 1: In-stock products ────────────────────────────────────
        _prog(0.05, t(
            "Step 1/4 — Loading in-stock products...",
            "الخطوة 1/4 — تحميل المنتجات المتوفرة..."))

        all_prods = _x(u, db, uid, ak, "product.product", "search_read",
                       [[["qty_available",">",0],["sale_ok","=",True]]],
                       {"fields":["id","default_code","display_name",
                                  "categ_id","list_price","qty_available"],
                        "limit":10000,"order":"default_code asc"})

        if not all_prods:
            _prog(1.0, t("No in-stock products found.","لا توجد منتجات في المخزون."))
            return empty, False

        all_pids = [p["id"] for p in all_prods]
        prod_map = {p["id"]: p for p in all_prods}
        _prog(0.15, t(
            f"Step 1/4 — Found {len(all_pids):,} in-stock products.",
            f"الخطوة 1/4 — تم العثور على {len(all_pids):,} منتج في المخزون."))

        # ── Step 2: Recently sold pids (single fast call) ────────────────
        _prog(0.20, t(
            f"Step 2/4 — Checking recent sales (last {threshold_days} days)...",
            f"الخطوة 2/4 — فحص المبيعات الأخيرة (آخر {threshold_days} يوم)..."))

        recent_sol = _x(u, db, uid, ak, "sale.order.line", "search_read",
                        [[["product_id","in",all_pids],
                          ["order_id.state","in",["sale","done"]],
                          ["order_id.date_order",">=",f"{cutoff} 00:00:00"]]],
                        {"fields":["product_id"],"limit":50000})

        recently_sold = set()
        for ln in (recent_sol or []):
            pid = ln["product_id"][0] if isinstance(ln.get("product_id"),list) else None
            if pid: recently_sold.add(pid)

        dead_pids = [p for p in all_pids if p not in recently_sold]
        _prog(0.35, t(
            f"Step 2/4 — {len(dead_pids):,} items have no recent sale (dead candidates).",
            f"الخطوة 2/4 — {len(dead_pids):,} صنف بلا مبيعات حديثة (مرشحون للركود)."))

        if not dead_pids:
            _prog(1.0, t(
                "All products sold recently — no dead stock!",
                "جميع المنتجات بيعت مؤخراً — لا مخزون راكد!"))
            return empty, False

        # ── Step 3: Last sale date per dead candidate ────────────────────
        # Smaller chunk = more progress updates + less timeout risk
        CHUNK      = 200
        n_chunks   = max(1, (len(dead_pids) + CHUNK - 1) // CHUNK)
        prog_start = 0.35
        prog_end   = 0.85

        _prog(prog_start, t(
            f"Step 3/4 — Fetching sale history for {len(dead_pids):,} items "
            f"({n_chunks} batches)...",
            f"الخطوة 3/4 — جلب تاريخ المبيعات لـ {len(dead_pids):,} صنف "
            f"({n_chunks} دفعة)..."))

        for batch_idx, i in enumerate(range(0, len(dead_pids), CHUNK)):
            chunk = dead_pids[i:i+CHUNK]
            batch_num = batch_idx + 1

            # live progress per batch
            pct = prog_start + (prog_end - prog_start) * (batch_idx / n_chunks)
            _prog(pct, t(
                f"Step 3/4 — Batch {batch_num}/{n_chunks} "
                f"({min(i+CHUNK, len(dead_pids)):,}/{len(dead_pids):,} items)...",
                f"الخطوة 3/4 — الدفعة {batch_num}/{n_chunks} "
                f"({min(i+CHUNK, len(dead_pids)):,}/{len(dead_pids):,} صنف)..."))

            try:
                sol_chunk = _x(u, db, uid, ak, "sale.order.line", "search_read",
                               [[["product_id","in",chunk],
                                 ["order_id.state","in",["sale","done"]]]],
                               {"fields":["product_id","order_id"],
                                "limit":50000,"order":"id desc"})
            except Exception as chunk_err:
                # One batch failed — mark partial and continue with what we have
                is_partial = True
                _prog(pct, t(
                    f"⚠️ Batch {batch_num} failed ({chunk_err}) — continuing with partial results.",
                    f"⚠️ فشلت الدفعة {batch_num} — متابعة بنتائج جزئية."))
                continue

            if not sol_chunk: continue

            oids = list({ln["order_id"][0] for ln in sol_chunk
                         if isinstance(ln.get("order_id"),list)})
            if not oids: continue

            try:
                orders_ch = _x(u, db, uid, ak, "sale.order", "search_read",
                               [[["id","in",oids]]],
                               {"fields":["id","date_order"],
                                "limit":len(oids)+5})
            except Exception:
                is_partial = True
                continue

            odate = {}
            for o in orders_ch:
                raw = o.get("date_order","")
                if raw:
                    try:
                        odate[o["id"]] = datetime.strptime(
                            raw, "%Y-%m-%d %H:%M:%S").date()
                    except Exception:
                        pass

            for ln in sol_chunk:
                pid = ln["product_id"][0] if isinstance(ln.get("product_id"),list) else None
                oid = ln["order_id"][0]    if isinstance(ln.get("order_id"),list) else None
                if pid is None or oid is None: continue
                d = odate.get(oid)
                if d is None: continue
                if pid not in last_sale_map or d > last_sale_map[pid]:
                    last_sale_map[pid] = d

        # ── Step 4: Build rows ───────────────────────────────────────────
        _prog(0.90, t(
            "Step 4/4 — Building results...",
            "الخطوة 4/4 — بناء النتائج..."))

        for pid in dead_pids:
            prod      = prod_map.get(pid, {})
            code      = str(prod.get("default_code") or "").strip()
            name      = prod.get("display_name") or ""
            cat       = prod.get("categ_id")
            categ     = cat[1] if isinstance(cat,list) and len(cat)>1 else ""
            price     = float(prod.get("list_price") or 0)
            qty       = float(prod.get("qty_available") or 0)
            if qty <= 0: continue
            frozen_val = round(qty * price, 2)

            last_sale = last_sale_map.get(pid)
            if last_sale is None:
                days_since = 99999
                status     = "Never Sold"
            else:
                days_since = (today - last_sale).days
                status     = "Dead Stock"

            rows.append({
                "Model Code"        : code if code else "—",
                "Product"           : name,
                "Category"          : categ,
                "On Hand"           : int(qty),
                "Unit Price"        : price,
                "Frozen Value (SAR)": frozen_val,
                "Last Sale Date"    : last_sale.strftime("%Y-%m-%d") if last_sale else "Never",
                "Days Since Sale"   : days_since,
                "Status"            : status,
            })

        if not rows:
            _prog(1.0, t("No dead stock found.","لا يوجد مخزون راكد."))
            return empty, False

        df = pd.DataFrame(rows)
        df = df.sort_values("Frozen Value (SAR)", ascending=False).reset_index(drop=True)
        _prog(1.0, t(
            f"Done — {len(df):,} dead stock items found.",
            f"اكتمل — تم العثور على {len(df):,} صنف راكد."))
        return df, is_partial

    except Exception as e:
        # Top-level failure — return whatever rows we built so far
        is_partial = True
        _prog(1.0, f"⚠️ Error: {e}")
        if rows:
            df = pd.DataFrame(rows)
            df = df.sort_values("Frozen Value (SAR)", ascending=False).reset_index(drop=True)
            return df, is_partial
        return empty, is_partial

# ─────────────────────────────────────────────────────────────────────────────
# COLUMN MAPS
# ─────────────────────────────────────────────────────────────────────────────
_COL_EN = {
    "System":"System","Model Code":"Model Code","Product":"Product",
    "Sale Price":"Sale Price","On Hand":"On Hand","Branch":"Branch",
    "Location":"Location","Reference":"Reference","Type":"Type",
    "State":"State","From":"From","To":"To","Qty":"Qty",
    "Scheduled":"Scheduled","Sold(30d)":"Sold(30d)","Daily Vel":"Daily Vel",
    "Days Left":"Days Left","Suggest":"Suggest","Priority":"Priority",
    "Purchase Qty":"Purchase Qty",
}
_COL_AR = {
    "System":"النظام","Model Code":"رمز الموديل","Product":"المنتج",
    "Sale Price":"سعر البيع","On Hand":"متوفر","Branch":"الفرع",
    "Location":"الموقع","Reference":"المرجع","Type":"النوع",
    "State":"الحالة","From":"من","To":"إلى","Qty":"الكمية",
    "Scheduled":"المجدول","Sold(30d)":"مباع(30ي)","Daily Vel":"معدل/يوم",
    "Days Left":"أيام متبقية","Suggest":"المقترح","Priority":"الأولوية",
    "Purchase Qty":"كمية المشتريات",
}

def localize_columns(df):
    if df is None or df.empty: return df
    return df.rename(columns=_COL_AR if get_lang() == "AR" else _COL_EN)

def prepare_df(df):
    return translate_system_names(localize_columns(df))

# ─────────────────────────────────────────────────────────────────────────────
# FETCH ALL DATA
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=30, show_spinner=False)
def fetch_all_data(codes_tuple, exact=False, need_branch=False,
                   need_transfers=False, need_reorder=False,
                   target_days=30, reorder_point=10):
    DAYS  = 30
    dfrom = (datetime.now() - timedelta(days=DAYS)).strftime("%Y-%m-%d 00:00:00")
    codes = list(codes_tuple)
    dom   = _domain(codes, exact)

    CS="System";CM="Model Code";CPR="Product";CP="Sale Price"
    CQ="On Hand";CB="Branch";CR="Reference";CT="Type"
    CST="State";CF="From";CTO="To";CQT="Qty"
    CD="Scheduled";CSOLD="Sold(30d)";CVEL="Daily Vel"
    CDAY="Days Left";CSUGG="Suggest";CPRI="Priority"
    SM={"draft":"Draft","waiting":"Waiting","confirmed":"Confirmed","assigned":"Ready"}

    def _one(key):
        cfg = get_system_config(key)
        # Store RAW KEY in System column so ns-matching works correctly.
        # prepare_df() will translate it to display name afterwards.
        sn  = key
        R   = {"key":key,"total":[],"branch":[],"transfers":[],"reorder":[]}
        if not cfg:
            R["total"].append({CS:sn,CM:"—",
                CPR:f"No config — add [{_canonical_key(key)}] to secrets.toml",
                CP:0.0,CQ:0,"_status":"ERROR"})
            return R
        auth_r = _auth(cfg["url"],cfg["db"],cfg["user"],cfg["api_key"])
        if not auth_r["ok"]:
            err_short = auth_r["error"].split(":")[0]   # e.g. BAD_CREDENTIALS
            R["total"].append({CS:sn,CM:"—",
                CPR:f"{err_short} — {auth_r['error']}",
                CP:0.0,CQ:0,"_status":"ERROR"})
            return R
        uid=auth_r["uid"];u=cfg["url"];db=cfg["db"];ak=cfg["api_key"]
        try:
            prods = _x(u,db,uid,ak,"product.product","search_read",[dom],
                       {"fields":["id","display_name","default_code","list_price"],
                        "limit":2000,"order":"default_code asc"})
            if not prods:
                R["total"].append({CS:sn,CM:"—",CPR:"Not found",CP:0.0,CQ:0,"_status":"NOT_FOUND"})
                return R
            pids = [p["id"] for p in prods]
            pmap = {p["id"]:p for p in prods}

            # Use stock.quant for accurate qty across ALL internal locations
            _all_locs = _x(u,db,uid,ak,"stock.location","search_read",
                          [[["usage","=","internal"],["active","=",True]]],
                          {"fields":["id"],"limit":10000})
            _loc_ids_all = [l["id"] for l in (_all_locs or [])]
            _qty_map = {}  # pid → total qty
            if _loc_ids_all:
                _qs_all = _x(u,db,uid,ak,"stock.quant","search_read",
                             [[["product_id","in",pids],
                               ["location_id","in",_loc_ids_all],
                               ["quantity",">",0]]],
                             {"fields":["product_id","quantity"],"limit":50000})
                for _q in (_qs_all or []):
                    _pid = _q["product_id"][0] if isinstance(_q.get("product_id"),list) else _q.get("product_id")
                    _qty_map[_pid] = _qty_map.get(_pid,0) + float(_q.get("quantity") or 0)

            for p in prods:
                R["total"].append({
                    CS:sn,CM:p.get("default_code") or "—",
                    CPR:p.get("display_name") or "",
                    CP:float(p.get("list_price") or 0),
                    CQ:int(_qty_map.get(p["id"],0)),
                    "_status":"OK"})
            if need_branch:
                locs = _x(u,db,uid,ak,"stock.location","search_read",
                          [[["usage","=","internal"],["active","=",True]]],
                          {"fields":["id"],"limit":10000})
                loc_ids = {l["id"] for l in locs}
                qs = _x(u,db,uid,ak,"stock.quant","search_read",
                        [[["product_id","in",pids],
                          ["location_id","in",list(loc_ids)],
                          ["quantity",">",0]]],
                        {"fields":["product_id","location_id","quantity"],"limit":5000})
                for q in qs:
                    _pr = q.get("product_id")
                    pid = (_pr[0] if isinstance(_pr,list) and _pr else _pr)
                    loc = q.get("location_id") or [None,"—"]
                    ln  = loc[1] if isinstance(loc,list) and len(loc)>1 else str(loc)
                    pm  = pmap.get(pid,{})
                    if not pm: continue  # skip if product not in pmap
                    _code = pm.get("default_code") or "—"
                    _name = pm.get("display_name") or ""
                    # Clean [CODE] prefix from display_name
                    if _name.startswith("[") and "]" in _name:
                        _name = _name[_name.index("]")+1:].strip()
                    R["branch"].append({
                        CS:sn,CB:ln,CM:_code,
                        CPR:_name,
                        CP:float(pm.get("list_price") or 0),
                        CQ:int(q.get("quantity") or 0),"_status":"OK"})
            if need_transfers:
                mvs = _x(u,db,uid,ak,"stock.move","search_read",
                         [[["product_id","in",pids],
                           ["state","in",["draft","waiting","confirmed","assigned"]]]],
                         {"fields":["picking_id","product_id","product_uom_qty"],"limit":2000})
                if mvs:
                    pkids = list({m["picking_id"][0] for m in mvs if isinstance(m.get("picking_id"),list)})
                    if pkids:
                        pks   = _x(u,db,uid,ak,"stock.picking","search_read",
                                   [[["id","in",pkids]]],
                                   {"fields":["id","name","picking_type_id","state",
                                              "location_id","location_dest_id","scheduled_date"]})
                        pkmap = {p["id"]:p for p in pks}
                        for mv in mvs:
                            pr = mv.get("picking_id")
                            if not isinstance(pr,list): continue
                            pk = pkmap.get(pr[0],{})
                            def _n(f,_p=pk):
                                v=_p.get(f); return v[1] if isinstance(v,list) else (v or "—")
                            sd = pk.get("scheduled_date") or "—"
                            if sd != "—":
                                try: sd=datetime.strptime(sd,"%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
                                except: pass
                            pid2 = mv["product_id"][0] if isinstance(mv.get("product_id"),list) else None
                            pm2  = pmap.get(pid2,{})
                            R["transfers"].append({
                                CS:sn,CR:pk.get("name") or "—",
                                CT:_n("picking_type_id"),
                                CST:SM.get(pk.get("state",""),pk.get("state","")),
                                CF:_n("location_id"),CTO:_n("location_dest_id"),
                                CM:pm2.get("default_code") or "—",
                                CQT:int(mv.get("product_uom_qty") or 0),
                                CD:sd,"_status":"OK"})
            if need_reorder:
                sl = _x(u,db,uid,ak,"sale.order.line","search_read",
                        [[["product_id","in",pids],
                          ["order_id.state","in",["sale","done"]],
                          ["order_id.date_order",">=",dfrom]]],
                        {"fields":["product_id","product_uom_qty"],"limit":10000})
                sm2 = {}
                for l in sl:
                    pid = l["product_id"][0] if isinstance(l.get("product_id"),list) else None
                    if pid: sm2[pid] = sm2.get(pid,0)+float(l.get("product_uom_qty") or 0)
                for p in prods:
                    pid  = p["id"]; cq=int(p.get("qty_available") or 0)
                    sold = sm2.get(pid,0); vel=round(sold/DAYS,2)
                    dl   = str(round(cq/vel,1)) if vel>0 else "∞"
                    sg   = max(0,round(target_days*vel-cq))
                    pr2  = ("Critical" if cq<=0 else "Low" if cq<=reorder_point else "OK")
                    R["reorder"].append({
                        CS:sn,CM:p.get("default_code") or "—",
                        CPR:p.get("display_name") or "",
                        CQ:cq,CSOLD:int(sold),CVEL:vel,
                        CDAY:dl,CSUGG:sg,CPRI:pr2,"_status":"OK"})
        except Exception as e:
            R["total"].append({CS:sn,CM:"—",CPR:f"Error: {e}",CP:0.0,CQ:0,"_status":"ERROR"})
        return R

    at=[];ab=[];atr=[];ar=[]
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = {ex.submit(_one,k):k for k in SYSTEM_KEYS}
        for f in as_completed(futs):
            r = f.result()
            at.extend(r["total"]);ab.extend(r["branch"])
            atr.extend(r["transfers"]);ar.extend(r["reorder"])

    def _df(rows,cols):
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=cols)
    return {
        "total"    : _df(at,  ["System","Model Code","Product","Sale Price","On Hand","_status"]),
        "branch"   : _df(ab,  ["System","Branch","Model Code","Sale Price","On Hand","_status"]),
        "transfers": _df(atr, ["System","Reference","Type","State","From","To","Model Code","Qty","Scheduled","_status"]),
        "reorder"  : _df(ar,  ["System","Model Code","Product","On Hand","Sold(30d)","Daily Vel","Days Left","Suggest","Priority","_status"]),
    }

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL PURCHASE / SALES EXPORT
# ─────────────────────────────────────────────────────────────────────────────
def _excel_generic(df, sheet_name, hdr_color="1A7A82", hdr_txt="FFFFFF"):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    buf = io.BytesIO()
    clean = df.copy()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        clean.to_excel(w, index=False, sheet_name=sheet_name)
        ws  = w.sheets[sheet_name]
        hfill = PatternFill("solid", fgColor=hdr_color)
        hfont = Font(bold=True, color=hdr_txt, size=11, name="Calibri")
        halign = Alignment(horizontal="center", vertical="center")
        thin   = Side(border_style="thin", color="1A2A2C")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        afill  = PatternFill("solid", fgColor="F2FAFA")
        nfont  = Font(name="Calibri", size=10, color="333333")
        num_a  = Alignment(horizontal="right",  vertical="center")
        ctr_a  = Alignment(horizontal="center", vertical="center")
        tfill  = PatternFill("solid", fgColor="1A7A82")
        tfont  = Font(bold=True, name="Calibri", color="1A7A82")
        mr, mc = ws.max_row, ws.max_column
        ws.row_dimensions[1].height = 28
        for c in range(1, mc+1):
            cell = ws.cell(row=1, column=c)
            cell.fill=hfill; cell.font=hfont; cell.alignment=halign; cell.border=border
        for row in ws.iter_rows(min_row=2, max_row=mr):
            for cell in row:
                cell.border=border; cell.font=nfont
                if cell.row%2==0: cell.fill=afill
                cell.alignment = num_a if isinstance(cell.value,(int,float)) else ctr_a
            ws.row_dimensions[row[0].row].height=18
        for c in range(1, mc+1):
            cl = get_column_letter(c)
            ml = max((len(str(ws.cell(row=r,column=c).value or "")) for r in range(1,mr+1)), default=8)
            ws.column_dimensions[cl].width = min(max(ml+3,12),50)
        ws.freeze_panes=f"A2"; ws.auto_filter.ref=f"A1:{get_column_letter(mc)}{mr}"
        tr = mr+1
        ws.cell(row=tr,column=1,value="TOTAL").font=tfont
        ws.cell(row=tr,column=1).fill=tfill
        ws.cell(row=tr,column=1).alignment=ctr_a
        cnames = [ws.cell(row=1,column=c).value for c in range(1,mc+1)]
        for cn in ("Qty","Subtotal"):
            if cn in cnames:
                ci=cnames.index(cn)+1; cl=get_column_letter(ci)
                ws.cell(row=tr,column=ci,value=f"=SUM({cl}2:{cl}{mr})")
                ws.cell(row=tr,column=ci).font=tfont
                ws.cell(row=tr,column=ci).fill=tfill
                ws.cell(row=tr,column=ci).alignment=ctr_a
        ws.sheet_properties.tabColor="4AACB4"
    return buf.getvalue()

def to_excel_purchase(df): return _excel_generic(df, "SWAG Purchase")
def to_excel_sales(df):
    d = df.copy()
    if "Date" in d.columns: d["Date"] = d["Date"].astype(str).str[:10]
    return _excel_generic(d, "SWAG Sales")

# ─────────────────────────────────────────────────────────────────────────────
# BRANCH MATRIX EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def to_excel_branch_matrix(df_branch_filtered, lang="EN"):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    if df_branch_filtered is None or df_branch_filtered.empty: return b""
    cm = t("Model Code","رمز الموديل"); cb = t("Branch","الفرع")
    cl = t("Location","الموقع");        cp = t("Sale Price","سعر البيع")
    cq = t("On Hand","متوفر");          cpr = t("Product","المنتج")
    lp = t("Purchase Qty","كمية المشتريات")
    df = df_branch_filtered.copy()
    pc = cl if cl in df.columns else (cb if cb in df.columns else None)
    if not pc or cm not in df.columns:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            df.to_excel(w, index=False, sheet_name="BranchMatrix")
        return buf.getvalue()
    if cq in df.columns: df[cq] = pd.to_numeric(df[cq], errors="coerce").fillna(0)
    else: df[cq] = 0
    pivot = (df.pivot_table(index=cm, columns=pc, values=cq, aggfunc="sum", fill_value=0)
             .reset_index())
    pivot.columns.name = None
    if cp in df.columns:
        pm = df.groupby(cm)[cp].first().reset_index()
        pivot = pivot.merge(pm, on=cm, how="left")
        pivot[cp] = pd.to_numeric(pivot[cp], errors="coerce").fillna(0).round(2)
    else: pivot[cp] = 0.0
    tdf = st.session_state.get("total_df")
    pmap = {}
    if tdf is not None and not tdf.empty and cm in tdf.columns and cpr in tdf.columns:
        pmap = tdf.groupby(cm)[cpr].first().dropna().to_dict()
    pivot[cpr] = pivot[cm].map(pmap).fillna("")
    purmap = {}
    if tdf is not None and not tdf.empty:
        for k in ["Purchase Qty","كمية المشتريات",lp]:
            if k in tdf.columns and cm in tdf.columns:
                tmp = tdf.groupby(cm)[k].sum().to_dict()
                if tmp: purmap = tmp; break
    if not purmap:
        um = pivot[cm].dropna().unique().tolist()
        if um:
            try:
                ed = datetime.now().date(); sd = ed - timedelta(days=365)
                pdf = get_purchase_summary_by_model(tuple(um),
                      sd.strftime("%Y-%m-%d"), ed.strftime("%Y-%m-%d"))
                if not pdf.empty: purmap = dict(zip(pdf["Model Code"], pdf["Purchase Qty"]))
            except Exception: pass
    pivot[lp] = pivot[cm].map(purmap).fillna(0).astype(int)
    fixed = [cm,cpr,cp,lp]
    locs  = sorted(c for c in pivot.columns if c not in fixed)
    pivot = pivot[[c for c in fixed if c in pivot.columns] + locs]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pivot.to_excel(writer, index=False, sheet_name="BranchMatrix")
        ws = writer.sheets["BranchMatrix"]
        if lang == "AR": ws.sheet_view.rightToLeft = True
        hfill = PatternFill("solid", fgColor="1A7A82")
        hfont = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        halign = Alignment(horizontal="center", vertical="center")
        thin   = Side(border_style="thin", color="1A2A2C")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        afill  = PatternFill("solid", fgColor="F2FAFA")
        nfont  = Font(name="Calibri", size=10, color="333333")
        num_a  = Alignment(horizontal="right",  vertical="center")
        ctr_a  = Alignment(horizontal="center", vertical="center")
        tfill  = PatternFill("solid", fgColor="1A7A82")
        tfont  = Font(bold=True, color="1A7A82", name="Calibri")
        zfill  = PatternFill("solid", fgColor="FFF3CD")
        zfont  = Font(color="1A7A82", bold=True, name="Calibri")
        mr, mc = ws.max_row, ws.max_column
        cnames = [ws.cell(row=1, column=c).value for c in range(1, mc+1)]
        ws.row_dimensions[1].height = 28
        for c in range(1, mc+1):
            cell = ws.cell(row=1, column=c)
            cell.fill=hfill; cell.font=hfont; cell.alignment=halign; cell.border=border
        for ri in range(2, mr+1):
            for ci in range(1, mc+1):
                cell = ws.cell(row=ri, column=ci)
                cn   = cnames[ci-1]
                is_loc = cn not in (cm,cpr,cp,lp,None)
                cell.border=border; cell.font=nfont
                if ri%2==0: cell.fill=afill
                if is_loc and isinstance(cell.value,(int,float)) and cell.value==0:
                    cell.fill=zfill; cell.font=zfont
                cell.alignment = num_a if isinstance(cell.value,(int,float)) else ctr_a
            ws.row_dimensions[ri].height=18
        for c in range(1, mc+1):
            cl2 = get_column_letter(c)
            ml  = max((len(str(ws.cell(row=r,column=c).value or "")) for r in range(1,mr+1)), default=8)
            ws.column_dimensions[cl2].width = min(max(ml+3,12),50)
        ws.freeze_panes=f"A2"; ws.auto_filter.ref=f"A1:{get_column_letter(mc)}{mr}"
        tr = mr+1
        tc = ws.cell(row=tr,column=1,value=t("TOTAL","الإجمالي"))
        tc.font=tfont; tc.fill=tfill; tc.alignment=ctr_a
        ws.row_dimensions[tr].height=22
        for ci,cn in enumerate(cnames,start=1):
            if cn in (None,cm,cpr,cp): continue
            cl2 = get_column_letter(ci)
            tot = ws.cell(row=tr,column=ci)
            tot.value=f"=SUM({cl2}2:{cl2}{mr})"
            tot.font=tfont; tot.fill=tfill; tot.alignment=num_a
        ws.sheet_properties.tabColor="4AACB4"
        ws.page_setup.orientation="landscape"; ws.page_setup.fitToPage=True
        ws.page_setup.fitToWidth=1; ws.print_title_rows="1:1"
        ws.sheet_view.zoomScale=85
    return buf.getvalue()

def dl_name(tag, ext):
    return f"swag_{tag}_{datetime.now().strftime('%Y%m%d_%H%M')}.{ext}"

# ─────────────────────────────────────────────────────────────────────────────
# PRICE HISTORY
# ─────────────────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────────────────
# QTY DISPLAY
# ─────────────────────────────────────────────────────────────────────────────
def get_qty_display(qty, lang="EN"):
    """Show actual qty including 0. Only NaN → Not Available."""
    try:
        v = float(qty)
        if pd.isna(v):
            return "—"
        return int(v)   # show 0 as 0
    except Exception:
        return "—"

# ─────────────────────────────────────────────────────────────────────────────
# DISPLAY DF
# ─────────────────────────────────────────────────────────────────────────────
def display_df(df, thresh=0, table_key="tbl"):
    if df is None or df.empty:
        st.info(t("No data.", "لا بيانات."))
        return pd.DataFrame()
    work    = df.copy()
    sys_col = t("System","النظام"); mc_col = t("Model Code","رمز الموديل")
    pr_col  = t("Product","المنتج"); br_col = t("Branch","الفرع")
    loc_col = t("Location","الموقع"); qc = t("On Hand","متوفر")
    pc      = t("Sale Price","سعر البيع")
    has_sys = sys_col in work.columns; has_br = br_col in work.columns
    fc = st.columns([2,2,2,1.5])
    if has_sys:
        all_sys = sorted(work[sys_col].dropna().unique().tolist())
        with fc[0]:
            sel_sys = st.multiselect(t("Company","الشركة"), options=all_sys,
                                     default=all_sys, key=f"{table_key}_sys")
        if sel_sys: work = work[work[sys_col].isin(sel_sys)]
    if has_br:
        all_br = sorted(work[br_col].dropna().unique().tolist())
        with fc[1]:
            sel_br = st.multiselect(t("Branch","الفرع"), options=all_br,
                                    default=all_br, key=f"{table_key}_br")
        if sel_br: work = work[work[br_col].isin(sel_br)]
    with fc[2]:
        q = st.text_input(t("Search model / product","بحث موديل / منتج"),
                          value="", placeholder=t("e.g. XP6013","مثال: XP6013"),
                          key=f"{table_key}_q").strip()
    if q:
        ql   = q.lower()
        mask = pd.Series([False]*len(work), index=work.index)
        for col in [mc_col,pr_col,loc_col]:
            if col in work.columns:
                mask = mask | work[col].fillna("").str.lower().str.contains(ql, regex=False)
        work = work[mask]
    with fc[3]:
        sortable = [c for c in work.columns if c != "_status"]
        sort_by  = st.selectbox(t("Sort by","ترتيب"), options=["—"]+sortable,
                                index=0, key=f"{table_key}_sort")
    if sort_by and sort_by != "—" and sort_by in work.columns:
        try:
            work = work.sort_values(
                by=sort_by,
                key=lambda s: pd.to_numeric(s,errors="coerce").fillna(0)
                              if pd.api.types.is_numeric_dtype(pd.to_numeric(s,errors="coerce"))
                              else s, ascending=True)
        except Exception:
            work = work.sort_values(by=sort_by)
    if work.empty:
        st.warning(t("No rows match your filters.","لا توجد نتائج بعد الفلتر."))
        return pd.DataFrame()
    if qc in work.columns:
        raw_q = pd.to_numeric(work[qc],errors="coerce")
        mn,mx = int(raw_q.min() or 0), int(raw_q.max() or 0)
        if mx > mn:
            qr    = st.slider(t("Qty range","نطاق الكمية"),
                              min_value=0,max_value=mx,value=(0,mx),
                              key=f"{table_key}_qrange")
            raw_q2 = pd.to_numeric(work[qc],errors="coerce")
            work   = work[(raw_q2>=qr[0])&(raw_q2<=qr[1])]
    ok_work = work[work["_status"]=="OK"] if "_status" in work.columns else work
    sm1,sm2,sm3,sm4 = st.columns(4)
    sm1.metric(t("Rows","الصفوف"), len(work))
    if qc in ok_work.columns:
        sm2.metric(t("Total Qty","إجمالي الكمية"),
                   int(pd.to_numeric(ok_work[qc],errors="coerce").fillna(0).sum()))
    if pc in ok_work.columns:
        vp = pd.to_numeric(ok_work[pc],errors="coerce")
        sm3.metric(t("Avg Price (SAR)","متوسط السعر ر.س"),
                   f"{vp[vp>0].mean():,.0f}" if not vp[vp>0].empty else "—")
    if has_sys and sys_col in ok_work.columns:
        sm4.metric(t("Companies","الشركات"), ok_work[sys_col].nunique())
    show     = work.drop(columns=["_status"],errors="ignore").copy()
    _raw_qty = (pd.to_numeric(work[qc],errors="coerce").fillna(0)
                if qc in work.columns else pd.Series(dtype=float,index=work.index))
    if pc in show.columns:
        show[pc] = pd.to_numeric(show[pc],errors="coerce").map(
            lambda v: f"{v:.2f} SAR" if pd.notna(v) else "—")
    if qc in show.columns:
        _lang = get_lang()
        _na_text = "Not Available" if _lang=="EN" else "غير متوفر"
        _status_col = work["_status"] if "_status" in work.columns else pd.Series("OK",index=work.index)
        def _qty_fmt(pair):
            idx, v = pair
            if _status_col.get(idx,"OK") == "not_available":
                return _na_text
            return get_qty_display(v, _lang)
        show[qc] = pd.Series(
            [_qty_fmt((i,v)) for i,v in zip(show.index, pd.to_numeric(show[qc],errors="coerce"))],
            index=show.index)
    low_idx  = set()
    if thresh > 0 and qc in work.columns:
        raw_q3  = pd.to_numeric(work[qc],errors="coerce")
        low_idx = set(work.index[(raw_q3>0)&(raw_q3<=thresh)])
    _zero_set = set(_raw_qty.index[_raw_qty==0]) if not _raw_qty.empty else set()
    _na_en = "Not Available"; _na_ar = "غير متوفر"
    cols = show.columns.tolist()
    th_  = "".join(f"<th>{c}</th>" for c in cols)
    def _row(idx_row):
        i,row = idx_row
        is_zero = i in _zero_set
        cls = " na-row" if is_zero else (" rl" if i in low_idx else "")
        cells = "".join(
            f'<td class="cf">{v}</td>' if ci==0
            else (f'<td class="na-cell">{v}</td>'
                  if is_zero and isinstance(v,str) and v in (_na_en,_na_ar)
                  else f"<td>{v}</td>")
            for ci,v in enumerate(row))
        return f'<tr class="{cls}">{cells}</tr>'
    tbody = "".join(_row(x) for x in show.iterrows())
    st.markdown(
        f'{_TABLE_CSS}<div class="swag-wrap">'
        f'<table class="swag-tbl"><thead><tr>{th_}</tr></thead>'
        f'<tbody>{tbody}</tbody></table></div>',
        unsafe_allow_html=True)
    st.caption(f"{len(show)} {t('rows shown','صفوف معروضة')} / {len(df)} {t('total','إجمالي')}")
    return work.drop(columns=["_status"],errors="ignore").copy()

def _render_html_table(df_display):
    if df_display is None or df_display.empty:
        st.info(t("No data.","لا بيانات.")); return
    cols = df_display.columns.tolist()
    th_  = "".join(f"<th>{c}</th>" for c in cols)
    def _row(idx_row):
        _,row = idx_row
        cells = "".join(
            f'<td class="cf">{v}</td>' if ci==0 else f"<td>{v}</td>"
            for ci,v in enumerate(row))
        return f"<tr>{cells}</tr>"
    tbody = "".join(_row(x) for x in df_display.iterrows())
    st.markdown(
        f'{_TABLE_CSS}<div class="swag-wrap">'
        f'<table class="swag-tbl"><thead><tr>{th_}</tr></thead>'
        f'<tbody>{tbody}</tbody></table></div>',
        unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# SIZE BREAKDOWN HELPER
# ─────────────────────────────────────────────────────────────────────────────

# Canonical size order for columns
_SIZE_ORDER = ["2XS","XS","S","M","L","XL","XXL","2XL","3XL","4XL","5XL","OSFA"]

# Regex: extract size suffix from model code
import re as _re
_SIZE_RE = _re.compile(
    r'-?(2XS|XS|S|M|L|XL|XXL|2XL|3XL|4XL|5XL|OSFA|OS)$',
    _re.IGNORECASE
)

def _extract_size(code: str):
    """Return (base_model, size) from a model code like XP6013-M."""
    code = str(code).strip()
    m = _SIZE_RE.search(code)
    if m:
        size = m.group(0).lstrip("-").upper()
        base = code[:m.start()].rstrip("-").strip()
        return base, size
    return code, ""   # no recognisable size suffix

def build_size_pivot(df, mc_col, qc_col, sc_col, pc_col, thr=0):
    """
    Build size-breakdown pivot from the flat stock dataframe.
    Returns (pivot_df, size_cols_found) or (None, []) if not possible.

    Columns: System | Base Model | Price | S | M | L | XL | XXL | ... | Total
    Rows   : one per (system × base_model)
    Color  : applied via HTML in caller
    """
    if df is None or df.empty: return None, []
    if mc_col not in df.columns or qc_col not in df.columns: return None, []

    work = df.copy()
    work["_qty_num"] = pd.to_numeric(work[qc_col], errors="coerce").fillna(0)
    work[["_base","_size"]] = work[mc_col].apply(
        lambda c: pd.Series(_extract_size(str(c))))

    # Only rows that have a recognised size
    sized = work[work["_size"] != ""].copy()
    if sized.empty: return None, []

    # Pivot: index = (system, base_model), columns = size, values = sum qty
    idx_cols = [c for c in [sc_col, "_base"] if c in sized.columns]
    pivot = (sized
             .pivot_table(index=idx_cols, columns="_size",
                          values="_qty_num", aggfunc="sum", fill_value=0)
             .reset_index())
    pivot.columns.name = None

    # Add price (first price for that base model)
    if pc_col in sized.columns:
        price_map = (sized.groupby("_base")[pc_col]
                     .first().reset_index()
                     .rename(columns={"_base":"_base", pc_col:"_price"}))
        pivot = pivot.merge(price_map, on="_base", how="left")
    else:
        pivot["_price"] = 0.0

    # Add Total column
    size_cols_found = [s for s in _SIZE_ORDER if s in pivot.columns]
    # Also pick up any sizes NOT in our canonical list
    extra_sizes = [c for c in pivot.columns
                   if c not in idx_cols + ["_price","_base"]
                   and c not in _SIZE_ORDER]
    size_cols_found = size_cols_found + sorted(extra_sizes)

    pivot["Total"] = pivot[size_cols_found].sum(axis=1)

    # Final column order
    base_col_name  = t("Base Model","الموديل الأساسي")
    price_col_name = t("Unit Price","سعر الوحدة")
    sys_col_label  = sc_col if sc_col in pivot.columns else None

    rename_map = {"_base": base_col_name, "_price": price_col_name}
    pivot = pivot.rename(columns=rename_map)

    ordered = []
    if sys_col_label and sys_col_label in pivot.columns:
        ordered.append(sys_col_label)
    ordered += [base_col_name, price_col_name]
    ordered += size_cols_found + ["Total"]
    ordered  = [c for c in ordered if c in pivot.columns]
    pivot    = pivot[ordered].sort_values(
        [sys_col_label, base_col_name] if sys_col_label else [base_col_name]
    ).reset_index(drop=True)

    return pivot, size_cols_found


def render_size_pivot(pivot_df, size_cols, thr=0):
    """Render the pivot as a colour-coded HTML table."""
    if pivot_df is None or pivot_df.empty:
        return

    cols = pivot_df.columns.tolist()
    th   = "".join(f"<th>{c}</th>" for c in cols)

    def _cell(col, val):
        # Size columns — colour by qty
        if col in size_cols or col == "Total":
            try:
                v = float(val)
            except Exception:
                return f"<td>{val}</td>"
            if v == 0:
                return (f'<td style="color:#DC2626;'
                        f'font-size:11px;">0</td>')
            elif thr > 0 and v <= thr:
                return (f'<td style="color:#D4A84B;font-weight:600;">'
                        f'{int(v)}</td>')
            else:
                return (f'<td style="color:#7FCDD3;font-weight:500;">'
                        f'{int(v)}</td>')
        # Total column bold
        if col == "Total":
            return f'<td style="color:#111827;font-weight:600;">{int(float(val)) if val else 0}</td>'
        # Price column
        if "Price" in str(col) or "سعر" in str(col):
            try:
                return f'<td style="color:#D4A84B;font-family:Outfit,monospace;font-size:11px;">{float(val):.2f}</td>'
            except Exception:
                return f"<td>{val}</td>"
        # Base model — monospace
        return f'<td class="cf">{val}</td>'

    def _row(ir):
        _, row = ir
        cells = "".join(_cell(col, val) for col, val in row.items())
        return f"<tr>{cells}</tr>"

    tbody = "".join(_row(x) for x in pivot_df.iterrows())

    _SZ_CSS = """<style>
.sz-wrap{width:100%;overflow-x:auto;border:1px solid rgba(74,172,180,0.08);
  border-radius:4px;overflow:hidden;margin-bottom:4px;}
.sz-tbl{width:100%;border-collapse:collapse;
  font-family:'Outfit','Tajawal',sans-serif;}
.sz-tbl thead tr{background:#E0F4F5;
  border-bottom:1px solid rgba(74,172,180,0.15);}
.sz-tbl thead th{color:#1A7A82;font-family:'Outfit',sans-serif;
  font-size:8px;letter-spacing:3px;text-transform:uppercase;
  font-weight:600;padding:12px 14px;text-align:center;white-space:nowrap;}
.sz-tbl tbody tr{border-bottom:1px solid #F3F4F6;
  transition:background 0.15s;}
.sz-tbl tbody tr:hover td{background:#EEF9FA;}
.sz-tbl tbody td{padding:10px 14px;text-align:center;
  font-size:12px;color:#6B7280;}
.sz-tbl tbody td.cf{font-family:'Outfit',monospace;font-size:11px;
  letter-spacing:0.5px;color:#111827;font-weight:500;
  border-right:1px solid rgba(74,172,180,0.08);}
</style>"""

    st.markdown(
        f'{_SZ_CSS}<div class="sz-wrap">'
        f'<table class="sz-tbl"><thead><tr>{th}</tr></thead>'
        f'<tbody>{tbody}</tbody></table></div>',
        unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────────────────────────────────────
def show_login():
    # ── Language toggle — fixed top right, above animated bg ─────────────
    # Inject CSS to push the radio widget to fixed top-right corner
    st.markdown("""
    <style>
    /* Force language toggle to fixed top-right above everything */
    div[data-testid="stRadio"]:has(label[style*="display: none"]) {
      position: fixed !important;
      top: 16px !important;
      right: 20px !important;
      z-index: 9999 !important;
      background: rgba(6,13,14,0.85) !important;
      border: 1px solid rgba(74,172,180,0.3) !important;
      border-radius: 100px !important;
      padding: 4px 12px !important;
      backdrop-filter: blur(10px) !important;
    }
    div[data-testid="stRadio"]:has(label[style*="display: none"]) label {
      color: rgba(255,255,255,0.7) !important;
      font-family: Outfit, sans-serif !important;
      font-size: 10px !important;
      letter-spacing: 2px !important;
    }
    div[data-testid="stRadio"]:has(label[style*="display: none"]) label[data-checked="true"],
    div[data-testid="stRadio"]:has(label[style*="display: none"]) [aria-checked="true"] + div {
      color: #4AACB4 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    lg = st.radio("",["EN","AR"],horizontal=True,
                  index=0 if get_lang()=="EN" else 1,
                  label_visibility="collapsed", key="llr")
    if lg != get_lang():
        st.session_state.lang = lg
        st.rerun()

    # ── Animated background + form ────────────────────────────────────────
    st.markdown("""
<style>
@keyframes float1{0%,100%{transform:translateY(0) rotate(0deg);}50%{transform:translateY(-18px) rotate(5deg);}}
@keyframes float2{0%,100%{transform:translateY(0) rotate(45deg);}50%{transform:translateY(-24px) rotate(50deg);}}
@keyframes float3{0%,100%{transform:translateY(0) rotate(20deg);}50%{transform:translateY(-12px) rotate(15deg);}}
@keyframes float4{0%,100%{transform:translateY(0) rotate(70deg);}60%{transform:translateY(-20px) rotate(65deg);}}
@keyframes float5{0%,100%{transform:translateY(0) rotate(30deg);}40%{transform:translateY(-16px) rotate(35deg);}}
@keyframes glowPulse{0%,100%{box-shadow:0 0 40px rgba(74,172,180,0.15),0 0 80px rgba(74,172,180,0.06);}
  50%{box-shadow:0 0 60px rgba(74,172,180,0.3),0 0 120px rgba(74,172,180,0.12),0 0 180px rgba(212,168,75,0.06);}}
@keyframes logoSpin{0%{transform:rotate(0deg);}100%{transform:rotate(360deg);}}
@keyframes titleReveal{from{opacity:0;transform:translateY(20px);}to{opacity:1;transform:translateY(0);}}
@keyframes shimmerBtn{0%{background-position:-200% center;}100%{background-position:200% center;}}
@keyframes beamRotate{0%{transform:rotate(0deg);}100%{transform:rotate(360deg);}}
@keyframes fadeInUp{from{opacity:0;transform:translateY(30px);}to{opacity:1;transform:translateY(0);}}
@keyframes borderGlow{0%,100%{border-color:rgba(74,172,180,0.2);}50%{border-color:#1A7A82;}}
@keyframes dotPulse{0%,100%{transform:scale(1);opacity:1;}50%{transform:scale(1.4);opacity:0.7;}}

.login-bg{
  display:none !important;  /* hidden by default — login page enables via style attr */
  position:fixed;inset:0;background:#F5F7FA;overflow:hidden;z-index:0;
  pointer-events:none;
}
.login-particle{
  position:absolute;opacity:0.06;
}
.login-particle svg path,.login-particle svg rect{stroke:#4AACB4;}

/* Radial glow blobs */
.login-glow-teal{
  position:absolute;width:600px;height:600px;border-radius:50%;
  background:radial-gradient(circle,rgba(74,172,180,0.12) 0%,transparent 70%);
  left:-150px;top:-150px;pointer-events:none;
}
.login-glow-gold{
  position:absolute;width:400px;height:400px;border-radius:50%;
  background:radial-gradient(circle,rgba(212,168,75,0.07) 0%,transparent 70%);
  right:-100px;bottom:-100px;pointer-events:none;
}

/* Grid lines */
.login-grid{
  position:absolute;inset:0;
  background-image:
    linear-gradient(rgba(74,172,180,0.03) 1px,transparent 1px),
    linear-gradient(90deg,rgba(74,172,180,0.03) 1px,transparent 1px);
  background-size:60px 60px;
}

.login-wrap{
  position:relative;z-index:1;
  display:flex;flex-direction:column;align-items:center;
  padding:40px 20px 24px;
  animation:fadeInUp 0.8s ease forwards;
}

/* Logo ring */
.login-logo-ring{
  position:relative;width:100px;height:100px;margin:0 auto 28px;
}
.ring-outer{
  position:absolute;inset:0;border-radius:50%;
  border:1px solid rgba(74,172,180,0.25);
  animation:glowPulse 3s ease-in-out infinite,borderGlow 3s ease-in-out infinite;
}
.ring-inner{
  position:absolute;inset:10px;border-radius:50%;
  border:1px dashed rgba(74,172,180,0.12);
  animation:logoSpin 20s linear infinite;
}
.ring-center{
  position:absolute;inset:20px;
  display:flex;align-items:center;justify-content:center;
}
.ring-dot{
  position:absolute;width:6px;height:6px;border-radius:50%;background:#D4A84B;
  animation:dotPulse 2s ease-in-out infinite;
}
.ring-dot.t{top:3px;left:50%;transform:translateX(-50%);}
.ring-dot.r{right:3px;top:50%;transform:translateY(-50%);animation-delay:0.5s;}
.ring-dot.b{bottom:3px;left:50%;transform:translateX(-50%);animation-delay:1s;}
.ring-dot.l{left:3px;top:50%;transform:translateY(-50%);animation-delay:1.5s;}

.login-title-big{
  font-family:'Cormorant Garamond',serif;
  font-size:52px;font-weight:300;color:#111827;
  text-align:center;letter-spacing:8px;
  margin-bottom:6px;
  text-shadow:0 0 40px rgba(74,172,180,0.3);
}
.login-eyebrow{
  font-family:'Outfit',sans-serif;font-size:9px;letter-spacing:5px;
  text-transform:uppercase;color:#1A7A82;text-align:center;
  margin-bottom:32px;
}

/* Glass card */
.login-glass{
  width:100%;max-width:380px;
  background:#F9FAFB;
  backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);
  border:1.5px solid #B3D9DB;
  border-radius:16px;padding:32px;
  box-shadow:0 24px 64px rgba(0,0,0,0.4),inset 0 1px 0 rgba(255,255,255,0.05);
  animation:fadeInUp 0.9s 0.2s ease both;
}

.login-footer{
  font-family:'Outfit',sans-serif;font-size:8px;
  letter-spacing:3px;color:#D1D5DB;
  text-align:center;margin-top:24px;text-transform:uppercase;
}
</style>

<div class="login-bg" style="display:block !important;">
  <div class="login-glow-teal"></div>
  <div class="login-glow-gold"></div>
  <div class="login-grid"></div>

  <!-- Floating particles -->
  <div class="login-particle" style="top:8%;left:6%;animation:float1 7s ease-in-out infinite;">
    <svg width="48" height="48" viewBox="0 0 48 48" fill="none">
      <path d="M24 4 L44 24 L24 44 L4 24 Z" stroke="#4AACB4" stroke-width="0.8" fill="none"/>
      <path d="M24 12 L36 24 L24 36 L12 24 Z" stroke="#4AACB4" stroke-width="0.5" fill="none" opacity="0.5"/>
    </svg>
  </div>
  <div class="login-particle" style="top:15%;right:8%;animation:float2 9s ease-in-out infinite;">
    <svg width="32" height="32" viewBox="0 0 32 32" fill="none">
      <rect x="4" y="4" width="24" height="24" stroke="#D4A84B" stroke-width="0.6" fill="none" opacity="0.6"/>
    </svg>
  </div>
  <div class="login-particle" style="top:60%;left:4%;animation:float3 8s ease-in-out infinite;">
    <svg width="24" height="24" viewBox="0 0 24 24" fill="none">
      <path d="M12 2 L22 12 L12 22 L2 12 Z" stroke="#4AACB4" stroke-width="0.8" fill="rgba(74,172,180,0.06)"/>
    </svg>
  </div>
  <div class="login-particle" style="bottom:20%;right:5%;animation:float4 10s ease-in-out infinite;">
    <svg width="40" height="40" viewBox="0 0 40 40" fill="none">
      <path d="M20 3 L37 20 L20 37 L3 20 Z" stroke="#4AACB4" stroke-width="0.6" fill="none"/>
      <circle cx="20" cy="3" r="1.5" fill="#D4A84B"/>
      <circle cx="37" cy="20" r="1.5" fill="#D4A84B"/>
      <circle cx="20" cy="37" r="1.5" fill="#D4A84B"/>
      <circle cx="3" cy="20" r="1.5" fill="#D4A84B"/>
    </svg>
  </div>
  <div class="login-particle" style="top:40%;right:3%;animation:float5 6s ease-in-out infinite;">
    <svg width="20" height="20" viewBox="0 0 20 20" fill="none">
      <rect x="2" y="2" width="16" height="16" stroke="#D4A84B" stroke-width="0.6" fill="none" opacity="0.5" transform="rotate(45 10 10)"/>
    </svg>
  </div>
  <div class="login-particle" style="bottom:35%;left:8%;animation:float2 11s ease-in-out infinite;">
    <svg width="28" height="28" viewBox="0 0 28 28" fill="none">
      <path d="M14 2 L26 14 L14 26 L2 14 Z" stroke="#4AACB4" stroke-width="0.5" fill="none" opacity="0.4"/>
    </svg>
  </div>
</div>
""", unsafe_allow_html=True)

    _,col,_ = st.columns([1,1.2,1])
    with col:
        st.markdown("""
        <div class="login-wrap">
          <div class="login-logo-ring">
            <div class="ring-outer"></div>
            <div class="ring-inner"></div>
            <div class="ring-dot t"></div>
            <div class="ring-dot r"></div>
            <div class="ring-dot b"></div>
            <div class="ring-dot l"></div>
            <div class="ring-center">
              <svg width="44" height="44" viewBox="0 0 44 44" fill="none">
                <path d="M22 4 L38 22 L22 40 L6 22 Z" stroke="#4AACB4" stroke-width="1.2" fill="none"/>
                <path d="M22 10 L32 22 L22 34 L12 22 Z" stroke="#4AACB4" stroke-width="0.7" fill="none" opacity="0.5"/>
                <path d="M22 15 L28 22 L22 29 L16 22 Z" fill="#4AACB4" opacity="0.5"/>
              </svg>
            </div>
          </div>
          <div class="login-title-big">SWAG</div>
          <div class="login-eyebrow">Product Intelligence · 5 Systems</div>
        </div>
        """, unsafe_allow_html=True)

        # Style the Streamlit form container to look like glass card
        st.markdown("""
        <style>
        /* Target the form container directly */
        [data-testid="stForm"]{
          background:#F9FAFB !important;
          backdrop-filter:blur(20px) !important;
          -webkit-backdrop-filter:blur(20px) !important;
          border:1.5px solid #1A7A82 !important;
          border-radius:16px !important;
          padding:24px !important;
          box-shadow:0 24px 64px rgba(0,0,0,0.4),
                     inset 0 1px 0 rgba(255,255,255,0.05) !important;
          animation:fadeInUp 0.9s 0.2s ease both !important;
        }
        </style>
        """, unsafe_allow_html=True)

        with st.form("lf", clear_on_submit=False):
            em = st.text_input(
                t("Email","البريد الإلكتروني"),
                placeholder="you@swag.com.sa")
            pw = st.text_input(
                t("Password","كلمة المرور"),
                type="password", placeholder="••••••••")
            st.markdown("<br>", unsafe_allow_html=True)
            sub = st.form_submit_button(
                t("Sign In →","تسجيل الدخول →"),
                use_container_width=True, type="primary")

        # Fix 2: Footer — brighter so it's readable
        st.markdown("""
        <div style='text-align:center;margin-top:20px;
                    font-family:Outfit,sans-serif;font-size:9px;
                    letter-spacing:3px;text-transform:uppercase;
                    color:#6B7280;'>
          SWAG DASHBOARD · 2025 · POWERED BY ODOO
        </div>
        """, unsafe_allow_html=True)

        if sub:
            if not em or not pw:
                st.error(t("Fill in both fields.","يرجى ملء جميع الحقول.")); return
            if "LOGIN" not in st.secrets:
                st.error("LOGIN section missing in secrets.toml"); return
            cfg = st.secrets["LOGIN"]
            with st.spinner(t("Authenticating...","جارٍ التحقق...")):
                try:
                    _login_url = str(cfg.get("url","")).rstrip("/")
                    if _login_url.endswith("/odoo"):
                        _login_url = _login_url[:-len("/odoo")]
                    proxy = xmlrpc.client.ServerProxy(
                        f"{_login_url}/xmlrpc/2/common", allow_none=True)
                    uid = proxy.authenticate(cfg["db"], em, pw, {})
                    if uid:
                        token = _make_token(em)
                        st.query_params["u"] = em
                        st.query_params["t"] = token
                        st.session_state.authenticated = True
                        st.session_state.user_email    = em
                        time.sleep(0.3); st.rerun()
                    else:
                        st.error(t("Wrong email or password.",
                                   "بريد إلكتروني أو كلمة مرور خاطئة."))
                except Exception as e:
                    st.error(f"Connection error: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# LOGOUT
# ─────────────────────────────────────────────────────────────────────────────
def do_logout():
    try: st.query_params.clear()
    except Exception: pass
    st.session_state.authenticated = False
    st.session_state.user_email    = ""
    st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
def show_dashboard():

    # ── SIDEBAR ──────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown(f"""
        <div style='padding:24px 0 20px;border-bottom:1px solid rgba(74,172,180,0.08);margin-bottom:20px;'>
          <div style='display:flex;align-items:center;gap:10px;'>
            <svg width="28" height="28" viewBox="0 0 32 32" fill="none">
              <path d="M16 2 L28 16 L16 30 L4 16 Z" stroke="#4AACB4" stroke-width="1" fill="rgba(74,172,180,0.04)"/>
              <path d="M16 9 L23 16 L16 23 L9 16 Z" fill="#4AACB4" opacity="0.3"/>
              <circle cx="16" cy="2"  r="1.5" fill="#D4A84B"/>
              <circle cx="28" cy="16" r="1.5" fill="#D4A84B"/>
              <circle cx="16" cy="30" r="1.5" fill="#D4A84B"/>
              <circle cx="4"  cy="16" r="1.5" fill="#D4A84B"/>
            </svg>
            <div>
              <div style='font-family:Outfit,sans-serif;font-size:13px;font-weight:600;
                          color:#111827;letter-spacing:2px;text-transform:uppercase;'>SWAG</div>
              <div style='font-family:Outfit,sans-serif;font-size:7px;
                          letter-spacing:3px;color:#1A7A82;text-transform:uppercase;'>Dashboard</div>
            </div>
          </div>
        </div>""", unsafe_allow_html=True)

        lc2 = st.radio(t("Language","اللغة"),["EN","AR"],
                       index=0 if get_lang()=="EN" else 1, horizontal=True)
        if lc2!=get_lang(): st.session_state.lang=lc2; st.rerun()

        st.markdown(f"""
        <div style='margin:16px 0 8px;font-family:Outfit,sans-serif;font-size:7px;
                    letter-spacing:3px;text-transform:uppercase;color:#9CA3AF;'>
          {st.session_state.user_email}
        </div>""", unsafe_allow_html=True)
        if st.button(t("Logout →","خروج →"), use_container_width=True, type="secondary"):
            do_logout()
        if st.button(t("🔄 Reload Data","🔄 تحديث البيانات"), use_container_width=True, type="secondary"):
            try:
                fetch_all_data.clear()
                st.cache_data.clear()
            except Exception:
                pass
            for _k in ["total_df","branch_df","transfers_df","reorder_df",
                       "sc_info","sc_long","sc_sname"]:
                st.session_state.pop(_k, None)
            st.rerun()

        st.divider()

        st.markdown(
            f"<div class='section-tag'>{t('Page','الصفحة')}</div>",
            unsafe_allow_html=True)
        _pg_opts=[
            t("🔍 Product Comparison","🔍 مقارنة المنتجات"),
            t("🌾 Season Comparison","🌾 مقارنة الموسم")]
        _cur_pg=st.session_state.get("_cur_page",_pg_opts[0])
        for _pgo in _pg_opts:
            _is_sel=_cur_pg==_pgo
            _bg="#1A7A82" if _is_sel else "transparent"
            _cl="#fff" if _is_sel else "#374151"
            _brd="#1A7A82"
            if st.button(
                _pgo,
                key=f"page_btn_{_pgo}",
                use_container_width=True,
                type="primary" if _is_sel else "secondary"):
                st.session_state["_cur_page"]=_pgo
                # Clear season state when switching
                if "Season" in _pgo:
                    pass
                st.rerun()

        st.divider()
        st.markdown(f"<div class='section-tag'>{t('Search Mode','وضع البحث')}</div>",
                    unsafe_allow_html=True)
        et = st.toggle(t("Exact match","تطابق تام"), value=st.session_state.search_exact)
        if et!=st.session_state.search_exact:
            st.session_state.search_exact=et
            st.session_state.total_df=None
            st.session_state.branch_df=None
            st.session_state.transfers_df=None
            st.rerun()

        st.divider()
        st.markdown(f"<div class='section-tag'>{t('Low Stock Alert','تنبيه المخزون')}</div>",
                    unsafe_allow_html=True)
        thr = st.number_input(t("Threshold (qty ≤)","الحد (كمية ≤)"),
                              min_value=0, max_value=1000,
                              value=st.session_state.low_stock_thresh, step=1)
        if thr!=st.session_state.low_stock_thresh:
            st.session_state.low_stock_thresh = int(thr)

        st.divider()
        if st.session_state.last_run:
            st.markdown(f"<div class='section-tag'>{t('Last Run','آخر تشغيل')}</div>",
                        unsafe_allow_html=True)
            st.caption(st.session_state.last_run.get("time",""))

    # Hide login page animated background on dashboard
    st.markdown("""
    <style>
    .login-bg{display:none !important;}
    .login-particle{display:none !important;}
    </style>
    """, unsafe_allow_html=True)

    # ── PAGE ROUTING — check early so season page skips all product content ──
    _cur_page=st.session_state.get("_cur_page",t("🔍 Product Comparison","🔍 مقارنة المنتجات"))
    _on_season="Season" in _cur_page or "موسم" in _cur_page

    if _on_season:
        # Render season comparison page directly, skip all product content
        import re as _re2
        from concurrent.futures import ThreadPoolExecutor as _TPE2, as_completed as _asc2
    if _on_season:
        ti = 0
        import re as _re2
        from concurrent.futures import ThreadPoolExecutor as _TPE, as_completed as _asc

        # ── Season type detection ─────────────────────────────────────────────
        _STC = [
            (("صيفي","صيف","summer","ss","su"),"SUMMER"),
            (("شتوي","شتاء","winter","aw","fw","wi"),"WINTER"),
            (("ربيعي","ربيع","spring","sp"),"SPRING"),
            (("خريفي","خريف","fall","autumn","fa"),"FALL"),
        ]
        _STL = {"SUMMER":"Summer / صيفي","WINTER":"Winter / شتوي",
                "SPRING":"Spring / ربيعي","FALL":"Fall / خريفي"}
        _SKF = [  # known season fields in priority order
            ("season_id","many2one"),("x_season_id","many2one"),
            ("x_studio_season_id","many2one"),("x_studio_season","many2one"),
            ("x_season","char"),("season","char"),
        ]
        _SBL = {"res.users","res.partner","res.company","stock.location",
                "stock.warehouse","stock.quant","ir.model","ir.ui.view"}
        _SHI = ["season","saison","موسم","فصل","x_season","x_collection","mawsim"]
        _LC  = ["System","Branch","Model Code","Product","Season","Qty","Price","_na"]

        def _stype(lbl):
            s=str(lbl or "").strip().lower()
            for words,canon in _STC:
                for w in words:
                    if (len(w)<=2 and _re2.search(rf"\b{_re2.escape(w)}\b",s)) or (len(w)>2 and w in s):
                        return canon
            return None

        def _syear(lbl):
            ns=_re2.findall(r"\d+",str(lbl or ""))
            if not ns: return ""
            n=ns[-1]
            if len(n)>=4: return n[:4]
            return "20"+n if len(n)==2 else n

        def _snorm(v):
            return _re2.sub(r"[\s\-_/]","",str(v or "").strip().lower())

        def _cname(dn,dc=""):
            s=str(dn or "").strip()
            if s.startswith("[") and "]" in s:
                p=s[s.index("]")+1:].strip()
                if p: return p
            return s

        def _ccode(dc,dn=""):
            c=str(dc or "").strip()
            if not c or c=="False": return ""
            if c.startswith("[") and "]" in c:
                c=c[1:c.index("]")].strip()
            return c

        def _chunks(seq,n):
            for i in range(0,len(seq),n): yield seq[i:i+n]

        # Location name fragments that indicate warehouse zones, not branches
        _LOC_SKIP_WORDS = {
            "input","output","pack","packing","quality","control","qc",
            "virtual","scrap","adjustment","inventory","transit",
            "production","manufacturing","repair","rma","return",
            "المناطق","منطقة","المستودع الافتراضي",
        }

        @st.cache_data(ttl=3600,show_spinner=False)
        def _slocs(sys_key):
            cfg=get_system_config(sys_key)
            if not cfg: return {}
            ar=_auth(cfg["url"],cfg["db"],cfg["user"],cfg["api_key"])
            if not ar["ok"]: return {}
            uid=ar["uid"]; u,db,ak=cfg["url"],cfg["db"],cfg["api_key"]
            try:
                rs=_proxy(u,"object").execute_kw(
                    db,uid,ak,"stock.location","search_read",
                    [[["usage","=","internal"],["active","=",True]]],
                    {"fields":["id","complete_name","name","location_id"],"limit":10000})
                out={}
                for l in rs or []:
                    nm=str(l.get("complete_name") or l.get("name") or str(l["id"])).strip()
                    out[l["id"]]=nm
                return out
            except: return {}

        def _get_seasons(x,db,uid,ak,fn,ft):
            seasons={}
            try:
                gs=x(db,uid,ak,"product.template","read_group",
                     [[[fn,"!=",False]]],[fn],[fn],{"lazy":False})
                for g in gs or []:
                    v=g.get(fn)
                    if not v and v is not False: continue
                    if ft=="many2one":
                        if isinstance(v,list) and len(v)>=2: seasons[v[0]]=str(v[1]).strip()
                    else:
                        sv=str(v).strip()
                        if sv: seasons[v]=sv
            except:
                try:
                    rs=x(db,uid,ak,"product.template","search_read",
                         [[[fn,"!=",False]]],{"fields":[fn],"limit":5000})
                    for r in rs:
                        v=r.get(fn)
                        if not v: continue
                        if ft=="many2one":
                            if isinstance(v,list) and len(v)>=2: seasons[v[0]]=str(v[1]).strip()
                        else:
                            sv=str(v).strip()
                            if sv: seasons[v]=sv
                except: pass
            return [(k,lb) for k,lb in seasons.items() if lb]

        @st.cache_data(ttl=3600,show_spinner=False)
        def _sdiscover(sys_key):
            cfg=get_system_config(sys_key)
            if not cfg: return None
            ar=_auth(cfg["url"],cfg["db"],cfg["user"],cfg["api_key"])
            if not ar["ok"]: return None
            uid=ar["uid"]; u,db,ak=cfg["url"],cfg["db"],cfg["api_key"]
            x=_proxy(u,"object").execute_kw
            try:
                fm=x(db,uid,ak,"product.template","fields_get",[],
                     {"attributes":["string","type","relation"]})
            except: return None
            existing=set(fm.keys())
            # Try known fields first
            for fn,ft in _SKF:
                if fn not in existing: continue
                fi=fm[fn]; ft2=fi.get("type",ft); rel=fi.get("relation","") or ""
                sl=_get_seasons(x,db,uid,ak,fn,ft2)
                if sl and len(sl)<=80:
                    return {"field":fn,"ftype":ft2,"relation":rel,
                            "seasons":sorted(sl,key=lambda z:str(z[1]))}
            # Scored fallback
            scored=[]
            for fn,fi in fm.items():
                ft2=fi.get("type",""); rel=fi.get("relation","") or ""; fl=fi.get("string",fn)
                if ft2 not in ("many2one","char","selection"): continue
                if rel in _SBL: continue
                sc=sum(30*(h in fn.lower())+25*(h in fl.lower()) for h in _SHI)
                sc+=40 if fn in {"season_id","x_season_id","x_studio_season_id"} else 0
                sc+=20 if any(h in rel.lower() for h in _SHI) else 0
                if sc>0: scored.append((sc,fn,ft2,rel))
            if not scored: return None
            scored.sort(reverse=True)
            fn,ft2,rel=scored[0][1],scored[0][2],scored[0][3]
            sl=_get_seasons(x,db,uid,ak,fn,ft2)
            if not sl or len(sl)>80: return None
            return {"field":fn,"ftype":ft2,"relation":rel,
                    "seasons":sorted(sl,key=lambda z:str(z[1]))}

        def _sresolve(query,info,mode):
            seasons=info.get("seasons",[])
            qt=_stype(query); out_v,out_l=[],[]
            for val,lbl in seasons:
                if mode=="type":
                    if qt and _stype(lbl)==qt: out_v.append(val); out_l.append(lbl)
                else:
                    if _snorm(lbl)==_snorm(query): out_v.append(val); out_l.append(lbl)
            return out_v,out_l

        # SWAG = master system. Its season defines what we compare.
        _MASTER_SYS = "SWAG"

        def _fetch_swag_season(info, query, mode, inc_archived, sys_key=None):
            """
            Fetch products for selected season from any system.
            Returns {model_code: {name, season, price}} — master dict.
            sys_key: which system to fetch from (default: first system with season match)
            """
            # Find which system to use
            _use_sys=sys_key or _MASTER_SYS
            # Find the system that has this season
            for _ts in [_use_sys]+[s for s in SYSTEM_KEYS if s!=_use_sys]:
                _ti=_sc_info.get(_ts) if '_sc_info' in dir() else info
                if _ti:
                    _tv,_=_sresolve(query,_ti,mode)
                    if _tv:
                        _use_sys=_ts; info=_ti; break
            cfg=get_system_config(_use_sys)
            if not cfg: return {}, pd.DataFrame(columns=_LC)
            ar=_auth(cfg["url"],cfg["db"],cfg["user"],cfg["api_key"])
            if not ar["ok"]: return {}, pd.DataFrame(columns=_LC)
            uid=ar["uid"]; u,db,ak=cfg["url"],cfg["db"],cfg["api_key"]
            x=_proxy(u,"object").execute_kw
            field=info["field"]; ftype=info["ftype"]
            ctx={"active_test":False} if inc_archived else {}

            vals,lbls=_sresolve(query,info,mode)
            val_to_lbl=dict(zip(vals,lbls))
            if not vals: return {}, pd.DataFrame(columns=_LC)

            try:
                # Get templates matching season
                dom=[[[field,"in",vals]]] if len(vals)>1 else [[[field,"=",vals[0]]]]
                trecs=x(db,uid,ak,"product.template","search_read",
                        dom,{"fields":["id",field],"limit":50000,"context":ctx}) or []
                tmpl_season={}
                for tr in trecs:
                    v=tr.get(field)
                    if isinstance(v,list) and v: v=v[0]
                    tmpl_season[tr["id"]]=val_to_lbl.get(v,", ".join(lbls))

                if not tmpl_season: return {}, pd.DataFrame(columns=_LC)

                # Get all product variants
                pmap={}
                for chunk in _chunks(list(tmpl_season.keys()),200):
                    recs=x(db,uid,ak,"product.product","search_read",
                           [[["product_tmpl_id","in",chunk],
                             ["default_code","!=",False]]],
                           {"fields":["id","default_code","display_name",
                                      "list_price","lst_price","product_tmpl_id"],
                            "limit":50000,"context":ctx})
                    if recs:
                        for p in recs:
                            if str(p.get("default_code") or "").strip():
                                pmap[p["id"]]=p

                if not pmap: return {}, pd.DataFrame(columns=_LC)
                pids=list(pmap.keys())

                # Quants
                locs=_slocs(_use_sys); loc_ids=list(locs.keys())
                quants=[]
                if loc_ids:
                    for chunk in _chunks(pids,1000):
                        qs=x(db,uid,ak,"stock.quant","search_read",
                             [[["product_id","in",chunk],
                               ["location_id","in",loc_ids],
                               ["quantity",">",0]]],
                             {"fields":["product_id","location_id","quantity"],
                              "limit":500000,"context":ctx})
                        if qs: quants.extend(qs)

                def _m(pid):
                    p=pmap.get(pid,{})
                    rc=str(p.get("default_code") or "").strip()
                    rn=str(p.get("display_name") or "").strip()
                    code=_ccode(rc,rn); name=_cname(rn,rc)
                    price=float(p.get("lst_price") or p.get("list_price") or 0)
                    tid=p.get("product_tmpl_id")
                    tid=tid[0] if isinstance(tid,list) and tid else tid
                    return code,name,price,tmpl_season.get(tid,"")

                rows=[]; seen=set(); master={}
                for q in quants:
                    pr=q.get("product_id"); pid=pr[0] if isinstance(pr,list) else pr
                    if pid not in pmap: continue
                    loc=q.get("location_id")
                    bn=(str(loc[1] if len(loc)>1 else locs.get(loc[0],"—")).strip()
                        if isinstance(loc,list) else str(locs.get(loc,"—")).strip())
                    seen.add(pid)
                    code,name,price,slbl=_m(pid)
                    if not code: continue
                    master[code]={"name":name,"season":slbl,"price":price}
                    rows.append({"System":_use_sys,"Branch":bn,
                                 "Model Code":code,"Product":name,"Season":slbl,
                                 "Qty":float(q.get("quantity") or 0),
                                 "Price":price,"_na":""})

                for pid in pids:
                    if pid not in seen:
                        code,name,price,slbl=_m(pid)
                        if not code: continue
                        master[code]={"name":name,"season":slbl,"price":price}
                        rows.append({"System":_use_sys,"Branch":"—",
                                     "Model Code":code,"Product":name,"Season":slbl,
                                     "Qty":0.0,"Price":price,"_na":""})

                df=pd.DataFrame(rows,columns=_LC)
                if not df.empty:
                    df=(df.groupby(["System","Branch","Model Code","_na"],as_index=False)
                          .agg({"Product":"first","Season":"first","Qty":"sum","Price":"max"}))
                return master, df

            except Exception:
                return {}, pd.DataFrame(columns=_LC)

        def _sfetch(sys_key, info, query, mode, inc_archived, master=None):
            """
            Fetch for NON-SWAG systems using SWAG master model codes.
            Does NOT use season field — matches by default_code only.
            Models in master but not here → NOT AVAILABLE.
            """
            if sys_key==_MASTER_SYS:
                # Should not be called for SWAG directly
                return pd.DataFrame(columns=_LC)

            cfg=get_system_config(sys_key)
            if not cfg:
                rows=[{"System":sys_key,"Branch":"—","Model Code":mc,
                       "Product":d["name"],"Season":d["season"],
                       "Qty":0.0,"Price":0.0,"_na":"NOT AVAILABLE"}
                      for mc,d in (master or {}).items()]
                return pd.DataFrame(rows,columns=_LC) if rows else pd.DataFrame(columns=_LC)

            ar=_auth(cfg["url"],cfg["db"],cfg["user"],cfg["api_key"])
            if not ar["ok"]:
                rows=[{"System":sys_key,"Branch":"—","Model Code":mc,
                       "Product":d["name"],"Season":d["season"],
                       "Qty":0.0,"Price":0.0,"_na":"NOT AVAILABLE"}
                      for mc,d in (master or {}).items()]
                return pd.DataFrame(rows,columns=_LC) if rows else pd.DataFrame(columns=_LC)

            uid=ar["uid"]; u,db,ak=cfg["url"],cfg["db"],cfg["api_key"]
            x=_proxy(u,"object").execute_kw
            ctx={"active_test":False} if inc_archived else {}
            master_codes=list(master.keys()) if master else []

            try:
                # Fetch only products whose default_code is in SWAG master
                pmap={}
                for chunk in _chunks(master_codes,500):
                    recs=x(db,uid,ak,"product.product","search_read",
                           [[["default_code","in",chunk]]],
                           {"fields":["id","default_code","display_name",
                                      "list_price","lst_price"],
                            "limit":50000,"context":ctx})
                    if recs:
                        for p in recs:
                            code=_ccode(str(p.get("default_code") or ""))
                            if code: pmap[p["id"]]=p

                found_codes={_ccode(str(p.get("default_code") or "")) for p in pmap.values()}

                pids=list(pmap.keys())
                locs=_slocs(sys_key); loc_ids=list(locs.keys())
                quants=[]
                if loc_ids and pids:
                    for chunk in _chunks(pids,1000):
                        qs=x(db,uid,ak,"stock.quant","search_read",
                             [[["product_id","in",chunk],
                               ["location_id","in",loc_ids],
                               ["quantity",">",0]]],
                             {"fields":["product_id","location_id","quantity"],
                              "limit":500000,"context":ctx})
                        if qs: quants.extend(qs)

                rows=[]; seen=set()
                for q in quants:
                    pr=q.get("product_id"); pid=pr[0] if isinstance(pr,list) else pr
                    if pid not in pmap: continue
                    loc=q.get("location_id")
                    bn=(str(loc[1] if len(loc)>1 else locs.get(loc[0],"—")).strip()
                        if isinstance(loc,list) else str(locs.get(loc,"—")).strip())
                    seen.add(pid)
                    p=pmap[pid]
                    code=_ccode(str(p.get("default_code") or ""))
                    name=_cname(str(p.get("display_name") or ""),str(p.get("default_code") or ""))
                    price=float(p.get("lst_price") or p.get("list_price") or 0)
                    # Use SWAG season for this model
                    mdata=master.get(code,{})
                    slbl=mdata.get("season","")
                    rows.append({"System":sys_key,"Branch":bn,
                                 "Model Code":code,"Product":name,"Season":slbl,
                                 "Qty":float(q.get("quantity") or 0),
                                 "Price":price,"_na":""})

                # Zero-stock: found in this system but no quant
                for pid in pids:
                    if pid not in seen:
                        p=pmap[pid]
                        code=_ccode(str(p.get("default_code") or ""))
                        name=_cname(str(p.get("display_name") or ""),str(p.get("default_code") or ""))
                        price=float(p.get("lst_price") or p.get("list_price") or 0)
                        mdata=master.get(code,{})
                        slbl=mdata.get("season","")
                        rows.append({"System":sys_key,"Branch":"—",
                                     "Model Code":code,"Product":name,"Season":slbl,
                                     "Qty":0.0,"Price":price,"_na":""})

                # NOT AVAILABLE: in SWAG master but not found in this system
                for mc,mdata in (master or {}).items():
                    if mc not in found_codes:
                        rows.append({"System":sys_key,"Branch":"—",
                                     "Model Code":mc,"Product":mdata.get("name",""),
                                     "Season":mdata.get("season",""),
                                     "Qty":0.0,"Price":0.0,"_na":"NOT AVAILABLE"})

                df=pd.DataFrame(rows,columns=_LC)
                if df.empty: return df
                return (df.groupby(["System","Branch","Model Code","_na"],as_index=False)
                          .agg({"Product":"first","Season":"first","Qty":"sum","Price":"max"}))

            except Exception as _e:
                # Return NOT AVAILABLE for all master codes on error
                rows=[{"System":sys_key,"Branch":"—","Model Code":mc,
                       "Product":(d.get("name","") if isinstance(d,dict) else str(d)),
                       "Season":(d.get("season","") if isinstance(d,dict) else ""),
                       "Qty":0.0,"Price":0.0,"_na":"NOT AVAILABLE"}
                      for mc,d in (master or {}).items()]
                return pd.DataFrame(rows,columns=_LC) if rows else pd.DataFrame(columns=_LC)

        def _sfetch_noseas(sys_key,inc_archived,master):
            """Fetch all products from no-season system. Mark NOT AVAILABLE if missing from master."""
            # master can be {code: {name,season,price}} or {code: name}
            def _mn(mc):
                d=(master or {}).get(mc,{})
                return d.get("name","") if isinstance(d,dict) else str(d)
            def _ms(mc):
                d=(master or {}).get(mc,{})
                return d.get("season","") if isinstance(d,dict) else ""
            zeros=[{"System":sys_key,"Branch":"—","Model Code":mc,"Product":_mn(mc),
                    "Season":_ms(mc),"Qty":0.0,"Price":0.0,"_na":"NOT AVAILABLE"}
                   for mc in (master or {})]
            cfg=get_system_config(sys_key)
            if not cfg: return pd.DataFrame(zeros,columns=_LC) if zeros else pd.DataFrame(columns=_LC)
            ar=_auth(cfg["url"],cfg["db"],cfg["user"],cfg["api_key"])
            if not ar["ok"]: return pd.DataFrame(zeros,columns=_LC) if zeros else pd.DataFrame(columns=_LC)
            uid=ar["uid"]; u,db,ak=cfg["url"],cfg["db"],cfg["api_key"]
            x=_proxy(u,"object").execute_kw
            ctx={"active_test":False} if inc_archived else {}
            try:
                prods=x(db,uid,ak,"product.product","search_read",
                        [[["default_code","!=",False],["default_code","!=",""]]],
                        {"fields":["id","default_code","display_name","list_price","lst_price"],
                         "limit":100000,"context":ctx}) or []
                pmap={p["id"]:p for p in prods}; pids=list(pmap.keys())
                this_codes={_ccode(str(p.get("default_code") or "")) for p in prods}-{""}
                locs=_slocs(sys_key); loc_ids=list(locs.keys())
                quants=[]
                if loc_ids and pids:
                    for chunk in _chunks(pids,1000):
                        qs=x(db,uid,ak,"stock.quant","search_read",
                             [[["product_id","in",chunk],["location_id","in",loc_ids],
                               ["quantity",">",0]]],
                             {"fields":["product_id","location_id","quantity"],"limit":500000,"context":ctx})
                        if qs: quants.extend(qs)
                rows=[]; seen=set()
                for q in quants:
                    pr=q.get("product_id"); pid=pr[0] if isinstance(pr,list) else pr
                    if pid not in pmap: continue
                    loc=q.get("location_id")
                    bn=(str(loc[1] if len(loc)>1 else locs.get(loc[0],"—")).strip()
                        if isinstance(loc,list) else str(locs.get(loc,"—")).strip())
                    seen.add(pid)
                    p=pmap[pid]
                    rc=str(p.get("default_code") or "").strip(); rn=str(p.get("display_name") or "").strip()
                    code=_ccode(rc,rn); name=_cname(rn,rc)
                    if not code: continue
                    price=float(p.get("lst_price") or p.get("list_price") or 0)
                    rows.append({"System":sys_key,"Branch":bn,"Model Code":code,"Product":name,
                                 "Season":"","Qty":float(q.get("quantity") or 0),"Price":price,"_na":""})
                for pid in pids:
                    if pid not in seen:
                        p=pmap[pid]; rc=str(p.get("default_code") or "").strip(); rn=str(p.get("display_name") or "").strip()
                        code=_ccode(rc,rn); name=_cname(rn,rc)
                        if not code: continue
                        price=float(p.get("lst_price") or p.get("list_price") or 0)
                        rows.append({"System":sys_key,"Branch":"—","Model Code":code,"Product":name,
                                     "Season":"","Qty":0.0,"Price":price,"_na":""})
                if master:
                    for mc,pn in master.items():
                        if mc and mc not in this_codes:
                            rows.append({"System":sys_key,"Branch":"—","Model Code":mc,"Product":pn,
                                         "Season":"","Qty":0.0,"Price":0.0,"_na":"NOT AVAILABLE"})
                df=pd.DataFrame(rows,columns=_LC)
                if df.empty: return pd.DataFrame(zeros,columns=_LC) if zeros else df
                return (df.groupby(["System","Branch","Model Code","_na"],as_index=False)
                          .agg({"Product":"first","Season":"first","Qty":"sum","Price":"max"}))
            except:
                return pd.DataFrame(zeros,columns=_LC) if zeros else pd.DataFrame(columns=_LC)

        # ═══════════════════════════════════════════════════════
        # UI — Header
        # ═══════════════════════════════════════════════════════
        st.markdown("""
        <style>
        .sc-hdr{font-family:'Cormorant Garamond',serif;font-size:36px;font-weight:300;
          color:#111827;margin-bottom:4px;}
        .sc-hdr em{font-style:normal;color:#1A7A82;}
        .sc-sub{font-family:'Outfit',sans-serif;font-size:9px;letter-spacing:4px;
          text-transform:uppercase;color:#9CA3AF;margin-bottom:24px;}
        .sc-pill-ok{background:#EEF9FA;border:1.5px solid #1A7A82;
          border-radius:8px;padding:10px;text-align:center;}
        .sc-pill-no{background:#F9FAFB;border:1px solid #E2E8F0;
          border-radius:8px;padding:10px;text-align:center;}
        .sc-pill-label{font-size:8px;letter-spacing:2px;text-transform:uppercase;margin-bottom:3px;}
        .sc-pill-field{font-size:9px;color:#6B7280;}
        .sc-pill-count{font-family:'Cormorant Garamond',serif;font-size:22px;
          font-weight:300;color:#111827;}
        .sc-pill-unit{font-size:8px;color:#9CA3AF;}
        </style>
        <div class="sc-hdr">Season <em>Comparison</em></div>
        <div class="sc-sub">Compare stock across all systems · all models included</div>
        """,unsafe_allow_html=True)

        # Controls row
        _sc_c1,_sc_c2=st.columns([1,2])
        with _sc_c1:
            if st.button(t("⟳ Reload Season Fields","⟳ إعادة تحميل"),
                         type="secondary",key="sc_reload"):
                try: _sdiscover.clear(); _slocs.clear(); st.cache_data.clear()
                except: pass
                for _k in ["sc_info","sc_long","sc_sname","sc_types","sc_exact","sc_noseas"]:
                    st.session_state.pop(_k,None)
                st.rerun()
        with _sc_c2:
            _sc_inc=st.checkbox(t("Include archived products","تضمين المؤرشف"),
                                value=False,key="sc_inc")

        # ── Discovery (parallel, cached) ────────────────────────────────────
        if "sc_info" not in st.session_state:
            _sc_ph=st.empty()
            _sc_ph.info(t("Detecting season fields...","اكتشاف حقول الموسم..."))
            _sc_imap={}
            with _TPE(max_workers=len(SYSTEM_KEYS)) as _ex:
                _fts={_ex.submit(_sdiscover,_sk):_sk for _sk in SYSTEM_KEYS}
                for _f in _asc(_fts):
                    _sk=_fts[_f]
                    try:
                        _r=_f.result()
                        if _r: _sc_imap[_sk]=_r
                    except: pass
            _sc_ph.empty()
            st.session_state["sc_info"]=_sc_imap
            st.session_state["sc_noseas"]=[s for s in SYSTEM_KEYS if s not in _sc_imap]
            _at=set(); _ae=set()
            for _inf in _sc_imap.values():
                for _,_lb in _inf.get("seasons",[]):
                    _tt=_stype(_lb)
                    if _tt: _at.add(_tt)
                    if _lb.strip(): _ae.add(_lb.strip())
            st.session_state["sc_types"]=sorted(
                _at,key=lambda z:["SUMMER","WINTER","SPRING","FALL"].index(z)
                if z in ["SUMMER","WINTER","SPRING","FALL"] else 9)
            st.session_state["sc_exact"]=sorted(_ae)

        _sc_info  = st.session_state.get("sc_info",{})
        _sc_noseas= st.session_state.get("sc_noseas",[])
        _sc_types = st.session_state.get("sc_types",[])
        _sc_exact = st.session_state.get("sc_exact",[])

        # ── System status pills ─────────────────────────────────────────────
        if _sc_info or _sc_noseas:
            _sc_scols=st.columns(len(SYSTEM_KEYS))
            for _si,_sk in enumerate(SYSTEM_KEYS):
                with _sc_scols[_si]:
                    _sn=get_system_name(_sk)
                    if _sk in _sc_info:
                        _nf=len(_sc_info[_sk].get("seasons",[])); _ff=_sc_info[_sk].get("field","?")
                        st.markdown(f"<div class='sc-pill-ok'>"
                            f"<div class='sc-pill-label' style='color:#1A7A82;'>{_sn}</div>"
                            f"<div class='sc-pill-field'>{_ff}</div>"
                            f"<div class='sc-pill-count'>{_nf}</div>"
                            f"<div class='sc-pill-unit'>{t('seasons','مواسم')}</div></div>",
                            unsafe_allow_html=True)
                    else:
                        st.markdown(f"<div class='sc-pill-no'>"
                            f"<div class='sc-pill-label' style='color:#9CA3AF;'>{_sn}</div>"
                            f"<div class='sc-pill-field' style='color:rgba(212,168,75,0.6);'>"
                            f"{t('All products','كل المنتجات')}</div></div>",
                            unsafe_allow_html=True)

        if not _sc_info:
            st.warning(t("No season fields detected automatically.",
                         "لم يتم اكتشاف حقول موسم تلقائياً."))
            # Manual field setup
            with st.expander(t("Manual Setup","إعداد يدوي"),expanded=True):
                _mf_c1,_mf_c2,_mf_c3=st.columns([2,1,1])
                _mf=_mf_c1.text_input(t("Field name","اسم الحقل"),
                    placeholder="season_id",key="sc_mf").strip()
                _mft=_mf_c2.selectbox(t("Type","النوع"),
                    ["many2one","char","selection"],key="sc_mft")
                _ms=_mf_c3.selectbox(t("System","النظام"),
                    SYSTEM_KEYS,format_func=get_system_name,key="sc_ms")
                if st.button(t("Apply →","تطبيق →"),type="primary",key="sc_mapply") and _mf:
                    _cfg_m=get_system_config(_ms)
                    if _cfg_m:
                        _ar_m=_auth(_cfg_m["url"],_cfg_m["db"],_cfg_m["user"],_cfg_m["api_key"])
                        if _ar_m["ok"]:
                            _xm=_proxy(_cfg_m["url"],"object").execute_kw
                            _sl=_get_seasons(_xm,_cfg_m["db"],_ar_m["uid"],_cfg_m["api_key"],_mf,_mft)
                            if _sl:
                                _imap={_ms:{"field":_mf,"ftype":_mft,"relation":"","seasons":_sl}}
                                for _os in SYSTEM_KEYS:
                                    if _os==_ms: continue
                                    _cfg_o=get_system_config(_os)
                                    if not _cfg_o: continue
                                    _ar_o=_auth(_cfg_o["url"],_cfg_o["db"],_cfg_o["user"],_cfg_o["api_key"])
                                    if not _ar_o["ok"]: continue
                                    _xo=_proxy(_cfg_o["url"],"object").execute_kw
                                    _sl_o=_get_seasons(_xo,_cfg_o["db"],_ar_o["uid"],
                                                       _cfg_o["api_key"],_mf,_mft)
                                    if _sl_o:
                                        _imap[_os]={"field":_mf,"ftype":_mft,
                                                    "relation":"","seasons":_sl_o}
                                st.session_state["sc_info"]=_imap
                                st.session_state["sc_noseas"]=[s for s in SYSTEM_KEYS if s not in _imap]
                                st.success(f"{len(_sl)} seasons found!"); st.rerun()
                            else: st.error("Field found but no values.")
        else:
            st.markdown("<br>",unsafe_allow_html=True)
            # ── Season selector ─────────────────────────────────────────────
            _sc_mode=st.radio(t("Mode","الوضع"),
                [t("By Type (all years)","حسب النوع (كل السنوات)"),
                 t("Exact Season","موسم محدد")],
                horizontal=True,key="sc_mode")
            _sc_q=""; _sc_rm="type"
            if t("By Type","حسب النوع") in _sc_mode:
                _sc_rm="type"
                if _sc_types:
                    _sc_tp=st.selectbox(t("Season Type","نوع الموسم"),
                        [""]+_sc_types,
                        format_func=lambda z: t("— Choose —","— اختر —") if z=="" else _STL.get(z,z),
                        key="sc_tp")
                    if _sc_tp: _sc_q=_sc_tp
                else: st.warning(t("No season types found.","لا أنواع مواسم."))
            else:
                _sc_rm="exact"
                if _sc_exact:
                    _sc_ep=st.selectbox(t("Season","الموسم"),[""]+_sc_exact,
                        format_func=lambda z: t("— Choose —","— اختر —") if z=="" else z,
                        key="sc_ep")
                    if _sc_ep: _sc_q=_sc_ep
                else: st.warning(t("No seasons found. Reload.","لا مواسم. أعد التحميل."))

            if _sc_q:
                _sc_dp=_STL.get(_sc_q,_sc_q) if _sc_rm=="type" else _sc_q
                st.info(f"**{t('Fetching:','جلب:')}** {_sc_dp} — {t('all models from all systems','جميع الموديلات من جميع الأنظمة')}")
                # Preview matches per system
                _pvc=st.columns(len(_sc_info))
                for _pvi,(_psk,_pinf) in enumerate(_sc_info.items()):
                    _pvv,_pvl=_sresolve(_sc_q,_pinf,_sc_rm)
                    with _pvc[_pvi]:
                        _pvtxt="<br>".join(_pvl[:6])+("..." if len(_pvl)>6 else "") if _pvl else t("No match","لا تطابق")
                        st.markdown(f"<div style='background:#EEF9FA;"
                            f"border:1.5px solid #C5E3E5;border-radius:6px;"
                            f"padding:8px 12px;'>"
                            f"<div style='font-size:8px;letter-spacing:2px;text-transform:uppercase;"
                            f"color:#1A7A82;margin-bottom:4px;'>{get_system_name(_psk)}</div>"
                            f"<div style='font-size:11px;color:#111827;'>{_pvtxt}</div></div>",
                            unsafe_allow_html=True)

            _sc_bc,_=st.columns([1,4])
            with _sc_bc:
                _sc_go=st.button(t("Compare →","قارن →"),type="primary",
                    use_container_width=True,disabled=not bool(_sc_q),key="sc_go")

            if _sc_go and _sc_q:
                _pr=st.progress(0.0)
                _st=st.empty()
                _parts={}

                # ── STEP 1: Find master system ────────────────────────────
                # Try SWAG first. If no match, try other season-aware systems.
                # This handles cases where a season exists in LaRouche but not SWAG.
                _master={}; _master_sys_used=None; _master_df=pd.DataFrame()

                # Try systems in priority: SWAG first, then others
                _sys_priority=["SWAG"]+[s for s in SYSTEM_KEYS if s!="SWAG"]
                for _msys in _sys_priority:
                    _minfo=_sc_info.get(_msys)
                    if not _minfo: continue
                    _vals,_lbls=_sresolve(_sc_q,_minfo,_sc_rm)
                    if not _vals: continue
                    # This system has a match — use it as master
                    _st.info(
                        f"1/2 — {t('Fetching master from','جلب الرئيسي من')} "
                        f"{get_system_name(_msys)} ({len(_vals)} {t('seasons','مواسم')})...")
                    _master, _master_df = _fetch_swag_season(_minfo, _sc_q, _sc_rm, _sc_inc)
                    if _master:
                        _master_sys_used=_msys
                        _parts[_msys]=_master_df
                        break

                _pr.progress(0.35)

                if not _master:
                    st.error(
                        t("No products found for this season in any system.",
                          "لا منتجات لهذا الموسم في أي نظام."))
                else:
                    st.caption(
                        f"📦 {len(_master):,} {t('models from','موديل من')} "
                        f"{get_system_name(_master_sys_used)} "
                        f"{t('(master)','(رئيسي)')}")

                    # ── STEP 2: All other systems by model code ───────────────
                    _other=[s for s in SYSTEM_KEYS if s!=_master_sys_used]
                    _st.info(
                        f"2/2 — {len(_master):,} {t('models. Fetching from other systems...','موديل. جلب من الأنظمة الأخرى...')}")
                    _err_log={}
                    with _TPE(max_workers=max(len(_other),1)) as _ex2:
                        _f2={_ex2.submit(_sfetch,_sk,
                                         _sc_info.get(_sk,{}),
                                         _sc_q,_sc_rm,_sc_inc,_master):_sk
                             for _sk in _other}
                        for _ff in _asc(_f2):
                            _sk2=_f2[_ff]
                            try:
                                _d=_ff.result()
                                if _d is not None and not _d.empty:
                                    _parts[_sk2]=_d
                            except Exception as _fe:
                                _err_log[_sk2]=str(_fe)
                    if _err_log:
                        for _esk,_emsg in _err_log.items():
                            st.caption(f"⚠️ {get_system_name(_esk)}: {_emsg}")

                    _pr.progress(1.0); _pr.empty(); _st.empty()

                    if not _parts:
                        st.error(t("No data found.","لا بيانات."))
                    else:
                        _long_all=pd.concat(_parts.values(),ignore_index=True)
                        _sc_dp2=_STL.get(_sc_q,_sc_q) if _sc_rm=="type" else _sc_q
                        _noseas2=[s for s in SYSTEM_KEYS
                                  if s!=_master_sys_used and s not in _sc_info]
                        st.session_state["sc_long"]=_long_all
                        st.session_state["sc_sname"]=_sc_dp2
                        st.session_state["sc_noseas"]=_noseas2
                        st.session_state["sc_qt"]=_stype(_sc_q) if _sc_rm=="type" else None
                        st.session_state["sc_qraw"]=_sc_q
                        st.session_state["sc_mcount"]=len(_master)
                        st.session_state["sc_master_sys"]=_master_sys_used
                        st.rerun()

            # ── RESULTS ────────────────────────────────────────────────────
            if "sc_long" in st.session_state:
                _long_all = st.session_state["sc_long"]
                _sname    = st.session_state.get("sc_sname","—")
                _noseas2  = st.session_state.get("sc_noseas",[])
                _qt3      = st.session_state.get("sc_qt",None)
                _qraw     = st.session_state.get("sc_qraw","")

                # ── Filter: only show season-matched rows ─────────────────
                # For season-aware systems: keep rows where season matches OR season is blank
                # For no-season systems: keep all (they have no season field)
                # Data is already accurate — SWAG season filter applied at fetch
                # No post-filter needed
                _long=_long_all.copy()
                _mc_count=st.session_state.get("sc_mcount",0)
                if _mc_count:
                    st.caption(f"📦 {_mc_count:,} {t('models from SWAG master for this season','موديل من SWAG الرئيسي لهذا الموسم')}")

                _active=[s for s in SYSTEM_KEYS if s in _long["System"].unique()]

                # Info banner for no-season systems
                if _noseas2:
                    _cf_nms=", ".join(get_system_name(s) for s in _noseas2 if s in _active)
                    if _cf_nms:
                        st.info(f"ℹ️ {t('No season field:','لا حقل موسم:')} **{_cf_nms}** — "
                                f"{t('all products fetched. NOT AVAILABLE = model missing in that system.','كل المنتجات محضرة.')}")

                # ── Metrics ───────────────────────────────────────────────
                _m1,_m2,_m3,_m4=st.columns(4)
                _m1.metric(t("Total Models","إجمالي الموديلات"),
                           f"{_long['Model Code'].nunique():,}")
                _m2.metric(t("Total Units","إجمالي الوحدات"),
                           f"{int(_long[_long['_na']!='NOT AVAILABLE']['Qty'].sum()):,}")
                _m3.metric(t("Systems","الأنظمة"),len(_active))
                _br_with_stock = _long[
                    (_long["Branch"]!="—") &
                    (_long["_na"]!="NOT AVAILABLE") &
                    (pd.to_numeric(_long["Qty"],errors="coerce").fillna(0)>0)
                ]["Branch"].nunique()
                _m4.metric(t("Branches","الفروع"), _br_with_stock)

                # Stock value
                _svr=[]
                for _svs in _active:
                    _svdf=_long[(_long["System"]==_svs)&(_long["_na"]!="NOT AVAILABLE")]
                    _svv=(_svdf.groupby("Model Code").agg({"Qty":"sum","Price":"max"})
                           .assign(val=lambda d:d["Qty"]*d["Price"])["val"].sum())
                    _svr.append((get_system_name(_svs),_svv))
                if _svr:
                    with st.expander(t("Stock Value per System","قيمة المخزون حسب النظام"),False):
                        _svc=st.columns(len(_svr))
                        for _svi,(_svn,_svv) in enumerate(_svr):
                            _svc[_svi].metric(_svn,f"{_svv:,.0f} SAR")

                # ── Build company pivot (Model Code only as index) ────────
                _av_map={}  # (model_code, system) → "YES" | "NOT AVAILABLE"
                for _,_lr in _long.iterrows():
                    _k=(_lr["Model Code"],_lr["System"])
                    if _av_map.get(_k)!="YES":
                        _av_map[_k]=_lr.get("_na","") or "YES"

                # Numeric qty pivot
                _long_n=_long.copy()
                _long_n["Qty"]=pd.to_numeric(_long_n["Qty"],errors="coerce").fillna(0)
                _long_n["Price"]=pd.to_numeric(_long_n["Price"],errors="coerce").fillna(0)
                _lna=_long_n[_long_n["_na"]!="NOT AVAILABLE"]

                _qp=(_lna.pivot_table(index="Model Code",columns="System",
                                      values="Qty",aggfunc="sum",fill_value=0)
                     .reset_index())
                _qp.columns.name=None
                _pp=(_lna.pivot_table(index="Model Code",columns="System",
                                      values="Price",aggfunc="max",fill_value=0)
                     .reset_index())
                _pp.columns.name=None

                # Aggregate product name + season per model
                _pagg=_long.groupby("Model Code")["Product"].agg(
                    lambda s: next((x for x in s if str(x).strip()),"")).to_dict()
                _sagg=_long.groupby("Model Code")["Season"].agg(
                    lambda s: next((x for x in s if str(x).strip()),"")).to_dict()

                _comp=_qp.copy()
                _comp["Product"]=_comp["Model Code"].map(_pagg).fillna("")
                _comp["Season"] =_comp["Model Code"].map(_sagg).fillna("")
                _comp["Year"]   =_comp["Season"].apply(_syear)

                # Ensure all systems have columns
                _all_sys=[s for s in SYSTEM_KEYS if s in _active or s in _noseas2 or s in _sc_info]
                for _sk5 in _all_sys:
                    if _sk5 not in _comp.columns: _comp[_sk5]=0
                    else: _comp[_sk5]=_comp[_sk5].fillna(0).astype(int)
                    _pcol=f"_p_{_sk5}"
                    _comp[_pcol]=(_pp[_sk5].reindex(_comp.index,fill_value=0).fillna(0).round(2)
                                  if _sk5 in _pp.columns else 0.0)

                _comp["Total"]=_comp[
                    [s for s in _all_sys if s in _comp.columns]].sum(axis=1).astype(int)

                # Apply NOT AVAILABLE display
                for _sk5 in _all_sys:
                    if _sk5 not in _comp.columns: continue
                    def _apna(row,sk=_sk5):
                        st2=_av_map.get((row["Model Code"],sk),"YES")
                        return "NOT AVAILABLE" if st2=="NOT AVAILABLE" else int(row[sk])
                    _comp[_sk5]=_comp.apply(_apna,axis=1)

                # Rename + order
                _ord=["Model Code","Product","Season","Year"]
                _rmap={}
                for _sk5 in _all_sys:
                    _dn=get_system_name(_sk5)
                    _rmap[_sk5]=f"{_dn} Qty"; _rmap[f"_p_{_sk5}"]=f"{_dn} Price"
                    _ord.append(_sk5); _ord.append(f"_p_{_sk5}")
                _ord.append("Total")
                _comp=_comp[[c for c in _ord if c in _comp.columns]].rename(columns=_rmap)
                _comp["Total"]=pd.to_numeric(_comp["Total"],errors="coerce").fillna(0).astype(int)
                _comp=_comp.sort_values(["Total","Model Code"],ascending=[False,True]).reset_index(drop=True)

                # Numeric version for health stats
                _qdn=[f"{get_system_name(s)} Qty" for s in _all_sys]
                _qdn_ok=[c for c in _qdn if c in _comp.columns]
                _cn=pd.DataFrame(index=_comp.index)
                for _cnc in _qdn_ok:
                    _cn[_cnc]=_comp[_cnc].apply(
                        lambda v: -1 if str(v)=="NOT AVAILABLE"
                        else (int(float(v)) if str(v).replace(".","").lstrip("-").isdigit()
                              else -1))

                # Health stats
                if len(_qdn_ok)>=2:
                    _hin=(_cn>0).sum(axis=1)
                    _hc1,_hc2,_hc3=st.columns(3)
                    _hc1.metric(t("Zero-Stock","بلا مخزون"),int((_comp["Total"]==0).sum()))
                    _hc2.metric(t("Single-System Only","نظام واحد فقط"),int((_hin==1).sum()))
                    _hc3.metric(t("In All Systems","في كل الأنظمة"),int((_hin==len(_qdn_ok)).sum()))

                # Price mismatch
                if not _pp.empty:
                    _par=[]
                    for _,_pr2 in _pp.iterrows():
                        _prs={get_system_name(s):float(_pr2.get(s,0) or 0)
                              for s in _all_sys if s in _pr2.index and float(_pr2.get(s,0) or 0)>0}
                        if len(_prs)<2: continue
                        _mn2,_mx2=min(_prs.values()),max(_prs.values())
                        if _mn2>0 and ((_mx2-_mn2)/_mn2)*100>=10:
                            _par.append({"Model Code":_pr2["Model Code"],
                                "Min Price":round(_mn2,2),"Max Price":round(_mx2,2),
                                "Diff %":round(((_mx2-_mn2)/_mn2)*100,1),
                                "Cheapest":min(_prs,key=_prs.get),
                                "Highest":max(_prs,key=_prs.get)})
                    if _par:
                        _padf2=pd.DataFrame(_par).sort_values("Diff %",ascending=False).reset_index(drop=True)
                        with st.expander(f"⚠️ {t('Price Mismatch','فرق أسعار')} — {len(_padf2)} {t('items','أصناف')}",False):
                            st.dataframe(_padf2,use_container_width=True,height=260)

                # ── View selector ─────────────────────────────────────────
                _sc_view=st.radio(t("View","عرض"),
                    [t("Company","بحسب الشركة"),
                     t("Branch","بحسب الفرع"),
                     t("Size","بحسب الحجم")],
                    horizontal=True,key="sc_vw")
                _sc_s=st.text_input(t("Search model / product","بحث"),
                    placeholder="e.g. RVT196",key="sc_srch").strip()

                if t("Company","بحسب الشركة") in _sc_view:
                    _show=_comp.copy()
                    _od=st.checkbox(t("Only differences","الاختلافات فقط"),key="sc_diff")
                    if _od and len(_qdn_ok)>=2:
                        _dm=_cn.max(axis=1)!=_cn.min(axis=1)
                        _show=_show[_dm].reset_index(drop=True)
                    if _sc_s:
                        _q2=_sc_s.lower()
                        _mm=(_show["Model Code"].astype(str).str.lower().str.contains(_q2,regex=False)
                             |_show["Product"].astype(str).str.lower().str.contains(_q2,regex=False))
                        _show=_show[_mm]
                    st.dataframe(_show.head(500),use_container_width=True,height=540)
                    st.caption(f"{min(len(_show),500):,} / {len(_show):,} {t('models','موديل')}")
                    _dc1,_dc2=st.columns(2)
                    _dc1.download_button("Excel ↓",to_excel(_show),
                        dl_name(f"sc_company_{_sname}","xlsx"),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="sc_xl",use_container_width=True)
                    _dc2.download_button("CSV ↓",_show.to_csv(index=False).encode("utf-8-sig"),
                        dl_name(f"sc_company_{_sname}","csv"),"text/csv",
                        key="sc_csv",use_container_width=True)

                elif t("Branch","بحسب الفرع") in _sc_view:
                    _lb2=_long[(_long["Branch"]!="—")&(_long["_na"]!="NOT AVAILABLE")].copy()
                    if _lb2.empty: st.info(t("No branch stock data.","لا بيانات فروع."))
                    else:
                        _bsc=sorted(_lb2["System"].unique())
                        _bsel=st.multiselect(t("Systems","الأنظمة"),_bsc,default=_bsc,
                            format_func=get_system_name,key="sc_brsys")
                        _bnf=st.text_input(t("Branch filter","فلتر الفرع"),
                            placeholder="Jeddah / جدة",key="sc_brnf").strip()
                        if _bsel: _lb2=_lb2[_lb2["System"].isin(_bsel)]
                        if _bnf:
                            _lb2=_lb2[_lb2["Branch"].astype(str).str.lower()
                                      .str.contains(_bnf.lower(),regex=False,na=False)]
                        _bp=(_lb2.pivot_table(index="Model Code",
                                              columns=["System","Branch"],
                                              values="Qty",aggfunc="sum",fill_value=0))
                        _bp.columns=[f"{get_system_name(a)} | {b}" for a,b in _bp.columns]
                        _bp=_bp.reset_index()
                        _bc=[c for c in _bp.columns if " | " in c]
                        for _bcc in _bc: _bp[_bcc]=_bp[_bcc].astype(int)
                        _bp=_bp.copy()
                        _bp["Product"]=_bp["Model Code"].map(_pagg).fillna("")
                        _bp["Total"]=_bp[_bc].sum(axis=1).astype(int)
                        _bp=_bp[_bp["Total"]>0].sort_values(
                            ["Total","Model Code"],ascending=[False,True]).reset_index(drop=True)
                        if _sc_s:
                            _q2=_sc_s.lower()
                            _mm=(_bp["Model Code"].astype(str).str.lower().str.contains(_q2,regex=False)
                                 |_bp["Product"].astype(str).str.lower().str.contains(_q2,regex=False))
                            _bp=_bp[_mm]
                        if len(_bc)>50:
                            st.warning(f"{len(_bc)} {t('branch columns — use system filter above.','عمود فرع — استخدم فلتر الأنظمة.')}")
                        st.dataframe(_bp.head(300),use_container_width=True,height=540)
                        st.caption(f"{min(len(_bp),300):,} / {len(_bp):,} {t('models','موديل')} · {len(_bc)} {t('branches','فرع')}")
                        _db1,_db2=st.columns(2)
                        _db1.download_button("Excel ↓",to_excel(_bp),
                            dl_name(f"sc_branch_{_sname}","xlsx"),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="sc_brxl",use_container_width=True)
                        _db2.download_button("CSV ↓",_bp.to_csv(index=False).encode("utf-8-sig"),
                            dl_name(f"sc_branch_{_sname}","csv"),"text/csv",
                            key="sc_brcsv",use_container_width=True)

                else:  # Size view
                    _lw2=_long[_long["_na"]!="NOT AVAILABLE"].copy()
                    _lw2["_qty"]=pd.to_numeric(_lw2["Qty"],errors="coerce").fillna(0)
                    _lw2[["_base","_size"]]=_lw2["Model Code"].apply(
                        lambda c: pd.Series(_extract_size(str(c))))
                    _lws=_lw2[_lw2["_size"]!=""]
                    if _lws.empty:
                        st.info(t("No size suffixes found (e.g. RVT196-M, needs -S/-M/-L/-XL).",
                                  "لا لاحقات أحجام في رموز الموديل."))
                    else:
                        _sp=(_lws.pivot_table(index="_base",columns="_size",
                                              values="_qty",aggfunc="sum",fill_value=0)
                             .reset_index())
                        _sp.columns.name=None
                        _scf=[s for s in _SIZE_ORDER if s in _sp.columns]
                        _sp["Total"]=_sp[_scf].sum(axis=1).astype(int)
                        _pm3=_lws.groupby("_base")["Product"].first().to_dict()
                        _sp.insert(1,"Product",_sp["_base"].map(_pm3).fillna(""))
                        _sp=_sp.rename(columns={"_base":"Base Model"})
                        if _sc_s:
                            _q2=_sc_s.lower()
                            _sp=_sp[_sp["Base Model"].astype(str).str.lower().str.contains(_q2,regex=False)]
                        _sp=_sp.sort_values(["Total","Base Model"],ascending=[False,True]).reset_index(drop=True)
                        st.dataframe(_sp.head(300),use_container_width=True,height=540)
                        st.caption(f"{min(len(_sp),300):,} / {len(_sp):,} base models")
                        _ds1,_ds2=st.columns(2)
                        _ds1.download_button("Excel ↓",to_excel(_sp),
                            dl_name(f"sc_sizes_{_sname}","xlsx"),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="sc_szxl",use_container_width=True)
                        _ds2.download_button("CSV ↓",_sp.to_csv(index=False).encode("utf-8-sig"),
                            dl_name(f"sc_sizes_{_sname}","csv"),"text/csv",
                            key="sc_szcsv",use_container_width=True)

                # WhatsApp
                _wa=[f"🌾 *SWAG Season — {_sname}*",
                     datetime.now().strftime("%Y-%m-%d %H:%M"),"",
                     f"📦 {t('Models','الموديلات')}: {_long['Model Code'].nunique():,}",
                     f"📊 {t('Units','الوحدات')}: {int(_long[_long['_na']!='NOT AVAILABLE']['Qty'].sum()):,}","",
                     f"{t('By system:','حسب النظام:')}"]
                for _wsk in _active:
                    _wu=int(_long[(_long["System"]==_wsk)&(_long["_na"]!="NOT AVAILABLE")]["Qty"].sum())
                    _wa.append(f"  • {get_system_name(_wsk)}: {_wu:,}")
                _wa_msg="\n".join(_wa+["","_SWAG Dashboard_"])
                with st.expander(t("WhatsApp Summary","ملخص واتساب"),False):
                    st.text_area("",value=_wa_msg,height=180,key="sc_wa")
                    import urllib.parse as _ulp
                    st.markdown(
                        f'<a href="https://wa.me/?text={_ulp.quote(_wa_msg)}" target="_blank"'
                        f' style="display:inline-block;padding:9px 22px;background:#25D366;'
                        f'border-radius:100px;font-size:10px;font-weight:600;'
                        f'letter-spacing:2px;color:#111827;text-decoration:none;">'
                        f'WhatsApp →</a>',unsafe_allow_html=True)

                if st.button(t("Clear","مسح"),type="secondary",key="sc_clear"):
                    for _k in ["sc_long","sc_sname","sc_qt","sc_qraw"]:
                        st.session_state.pop(_k,None)
                    st.rerun()

    # ── TODAY SNAPSHOT (product comparison)
    # Season page already rendered above if _on_season
    if not _on_season:  # product comparison content below
        _snap      = st.session_state.last_run
        _stats     = st.session_state.sys_stats
        _tdf_cache = st.session_state.total_df

        _email     = st.session_state.user_email or ""
        _firstname = _email.split("@")[0].split(".")[0].capitalize()

        _hour = datetime.now().hour
        if _hour < 12:   _greet = t("Good morning","صباح الخير")
        elif _hour < 17: _greet = t("Good afternoon","مساء الخير")
        else:            _greet = t("Good evening","مساء الخير")

        # System pills
        _sys_pills = ""
        for _k in SYSTEM_KEYS:
            _cfg_ok = bool(get_system_config(_k))
            _stat   = _stats.get(_k,"UNKNOWN") if _snap else ("CONFIGURED" if _cfg_ok else "NO_CONFIG")
            _dname  = get_system_name(_k)
            if not _cfg_ok:           _cls="offline"; _lbl=t("No Config","لا إعداد")
            elif _stat=="OK":         _cls="online";  _lbl=t("Online","متصل")
            elif _stat=="NOT_FOUND":  _cls="nodata";  _lbl=t("No Data","لا بيانات")
            elif _stat=="ERROR":      _cls="error";   _lbl=t("Error","خطأ")
            else:                     _cls="offline"; _lbl="—"
            _sys_pills += (
                f"<div class='sp sp-{_cls}'>"
                f"<div class='sd sd-{_cls}'></div>"
                f"<span class='sn sn-{_cls}'>{_dname}</span>"
                f"<span class='sb sb-{_cls}'>{_lbl}</span>"
                f"</div>")

        # Portfolio value + low stock count
        _port_val  = ""; _low_count = 0
        if _tdf_cache is not None and not _tdf_cache.empty:
            _qcx = t("On Hand","متوفر"); _pcx = t("Sale Price","سعر البيع")
            _okx = _tdf_cache[_tdf_cache["_status"]=="OK"] if "_status" in _tdf_cache.columns else _tdf_cache
            if _qcx in _okx.columns and _pcx in _okx.columns:
                _qv = pd.to_numeric(_okx[_qcx],errors="coerce").fillna(0)
                _pv = pd.to_numeric(_okx[_pcx],errors="coerce").fillna(0)
                _tv = (_qv*_pv).sum()
                _port_val  = f"{_tv/1000:,.0f}K SAR" if _tv>=1000 else f"{_tv:,.0f} SAR"
                _thr_x = st.session_state.low_stock_thresh
                if _thr_x>0: _low_count=int(((_qv>0)&(_qv<=_thr_x)).sum())

        # Last run info
        _last_html = ""
        _online_count = sum(1 for _k in SYSTEM_KEYS if _stats.get(_k)=="OK") if _snap else 0

        if _snap:
            _ago_s  = (datetime.now()-datetime.strptime(_snap["time"],"%Y-%m-%d %H:%M:%S")).total_seconds()
            _ago_str = (f"{int(_ago_s//3600)}h ago" if _ago_s>=3600
                        else f"{int(_ago_s//60)}m ago" if _ago_s>=60 else "just now")
            _last_html = (
                "<div class='snap-last'>"
                f"<div class='sl-label'>{t('Last Run','آخر تشغيل')}</div>"
                f"<div class='sl-val'>{_snap.get('models','—')} {t('model(s)','موديل')}</div>"
                f"<div class='sl-meta'>{_snap['time']}</div>"
                f"<div class='sl-ago'>{_ago_str}</div>"
                + (f"<div class='sl-rows'>→ {_snap.get('rows','?')} rows</div>" if _snap.get('rows') else "")
                + "</div>")

        _online_cls = "teal" if _online_count==len(SYSTEM_KEYS) else ("gold" if _online_count>0 else "red-v")

        # ── Part 1: CSS + greeting + KPI cards ──────────────────────────────────
        _port_card  = (
            "<div class='snap-card' style='animation-delay:.08s;"
            "border-color:rgba(212,168,75,.12)'>"
            f"<div class='sc-label'>{t('Portfolio','المحفظة')}</div>"
            f"<div class='sc-val gold'>{_port_val}</div>"
            f"<div class='sc-sub'>{t('last search','آخر بحث')}</div></div>"
        ) if _port_val else ""

        _low_card = (
            "<div class='snap-card' style='animation-delay:.12s;"
            "border-color:rgba(255,100,80,.1)'>"
            f"<div class='sc-label'>{t('Low Stock','مخزون منخفض')}</div>"
            f"<div class='sc-val red-v'>{_low_count}</div>"
            f"<div class='sc-sub'>{t('items','صنف')}</div></div>"
        ) if _low_count > 0 else ""

        _run_card = (
            "<div class='snap-card' style='animation-delay:.16s'>"
            f"<div class='sc-label'>{t('Last Run','آخر تشغيل')}</div>"
            f"<div class='sc-val' style='font-size:20px'>{_snap.get('rows','—')}</div>"
            f"<div class='sc-sub'>{t('rows','صفوف')}</div></div>"
        ) if _snap else ""

        _sub_online = ("All connected" if _online_count==len(SYSTEM_KEYS)
                       else f"{len(SYSTEM_KEYS)-_online_count} offline")
        _warn_html  = (
            f"<div class='snap-warn'>&#9888; {_low_count} "
            f"{t('items below low stock threshold','صنف تحت حد المخزون المنخفض')}</div>"
        ) if _low_count > 0 else ""

        st.markdown(f"""
        <style>
        @keyframes snapIn{{from{{opacity:0;transform:translateY(-14px)}}to{{opacity:1;transform:translateY(0)}}}}
        @keyframes cardIn{{from{{opacity:0;transform:translateX(-6px)}}to{{opacity:1;transform:translateX(0)}}}}
        @keyframes dotBlink{{0%,100%{{opacity:1}}50%{{opacity:0.25}}}}
        .snap-wrap{{padding:32px 0 20px;animation:snapIn .5s cubic-bezier(.22,.68,0,1.2) both}}
        .snap-greeting{{font-family:'Cormorant Garamond',serif;font-size:40px;font-weight:300;
          color:#111827;margin-bottom:4px;line-height:1.15}}
        .snap-greeting em{{font-style:normal;color:#1A7A82}}
        .snap-date{{font-family:'Outfit',sans-serif;font-size:9px;letter-spacing:4px;
          text-transform:uppercase;color:#9CA3AF;margin-bottom:24px}}
        .snap-cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));
          gap:10px;margin-bottom:22px}}
        .snap-card{{background:#F9FAFB;border:1px solid #E2E8F0;
          border-radius:12px;padding:16px 14px;animation:cardIn .5s ease both;
          transition:border-color .2s,background .2s}}
        .snap-card:hover{{border-color:rgba(74,172,180,0.2);background:#EEF9FA}}
        .sc-label{{font-family:'Outfit',sans-serif;font-size:8px;letter-spacing:3px;
          text-transform:uppercase;color:#9CA3AF;margin-bottom:10px}}
        .sc-val{{font-family:'Cormorant Garamond',serif;font-size:32px;font-weight:300;
          color:#111827;line-height:1;margin-bottom:3px}}
        .sc-val.teal{{color:#1A7A82}}.sc-val.gold{{color:#D4A84B}}
        .sc-val.red-v{{color:rgba(255,100,80,.85)}}
        .sc-sub{{font-family:'Outfit',sans-serif;font-size:9px;
          color:#9CA3AF;letter-spacing:.5px}}
        .snap-sys-label{{font-family:'Outfit',sans-serif;font-size:8px;letter-spacing:3px;
          text-transform:uppercase;color:#D1D5DB;margin-bottom:10px}}
        .snap-sys-row{{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:10px}}
        .sp{{display:flex;align-items:center;gap:7px;border-radius:100px;padding:6px 14px}}
        .sp-online{{background:#EEF9FA;border:1px solid rgba(74,172,180,0.18)}}
        .sp-offline{{background:#F9FAFB;border:1px solid #E2E8F0}}
        .sp-error{{background:rgba(255,100,80,0.05);border:1px solid rgba(255,100,80,0.14)}}
        .sp-nodata{{background:rgba(212,168,75,0.05);border:1px solid rgba(212,168,75,0.14)}}
        .sd{{width:7px;height:7px;border-radius:50%;flex-shrink:0}}
        .sd-online{{background:#4AACB4;animation:dotBlink 2.5s ease-in-out infinite}}
        .sd-offline{{background:rgba(255,255,255,0.14)}}
        .sd-error{{background:rgba(255,100,80,.75)}}
        .sd-nodata{{background:#D4A84B}}
        .sn{{font-family:'Outfit',sans-serif;font-size:11px;font-weight:500;letter-spacing:.5px}}
        .sn-online{{color:#1A7A82}}.sn-offline{{color:#9CA3AF}}
        .sn-error{{color:rgba(255,100,80,.7)}}.sn-nodata{{color:#B45309}}
        .sb{{font-family:'Outfit',sans-serif;font-size:7px;letter-spacing:1.5px;
          text-transform:uppercase;padding:2px 6px;border-radius:100px}}
        .sb-online{{background:rgba(74,172,180,.1);color:rgba(74,172,180,.55)}}
        .sb-offline{{background:rgba(255,255,255,.03);color:rgba(255,255,255,.14)}}
        .sb-error{{background:rgba(255,100,80,.09);color:rgba(255,100,80,.55)}}
        .sb-nodata{{background:rgba(212,168,75,.09);color:rgba(212,168,75,.55)}}
        .snap-last{{display:flex;align-items:center;gap:14px;flex-wrap:wrap;
          background:#EEF9FA;border:1px solid rgba(74,172,180,0.09);
          border-left:3px solid rgba(74,172,180,0.35);
          border-radius:0 10px 10px 0;padding:12px 18px;margin-top:12px}}
        .sl-label{{font-family:'Outfit',sans-serif;font-size:8px;letter-spacing:3px;
          text-transform:uppercase;color:#1A7A82;flex-shrink:0}}
        .sl-val{{font-family:'Cormorant Garamond',serif;font-size:18px;
          font-weight:300;color:#111827;letter-spacing:1px}}
        .sl-meta{{font-family:'Outfit',sans-serif;font-size:10px;color:#9CA3AF}}
        .sl-ago{{font-family:'Outfit',sans-serif;font-size:10px;color:#9CA3AF}}
        .sl-rows{{font-family:'Outfit',sans-serif;font-size:11px;color:#1A7A82;font-weight:500}}
        .snap-warn{{display:inline-flex;align-items:center;gap:8px;
          background:rgba(212,168,75,.06);border:1px solid rgba(212,168,75,.18);
          border-radius:8px;padding:8px 14px;margin-top:8px;
          font-family:'Outfit',sans-serif;font-size:10px;
          letter-spacing:1px;color:rgba(212,168,75,.8)}}
        .snap-divider{{height:1px;margin:28px 0 20px;
          background:linear-gradient(90deg,rgba(74,172,180,.25),rgba(74,172,180,.06),transparent)}}
        </style>

        <div class='snap-wrap'>
          <div class='snap-greeting'>{_greet}, <em>{_firstname}</em></div>
          <div class='snap-date'>{datetime.now().strftime("%A, %d %B %Y")} &nbsp;·&nbsp; SWAG Product Intelligence</div>
          <div class='snap-cards'>
            <div class='snap-card' style='animation-delay:.04s'>
              <div class='sc-label'>{t("Systems Online","الأنظمة المتصلة")}</div>
              <div class='sc-val {_online_cls}'>{_online_count}/{len(SYSTEM_KEYS)}</div>
              <div class='sc-sub'>{_sub_online}</div>
            </div>
            {_port_card}{_low_card}{_run_card}
          </div>
          <div class='snap-sys-label'>{t("Connected Systems","الأنظمة المتصلة")}</div>
        """, unsafe_allow_html=True)

        # ── Part 2: sys_pills rendered separately (avoid f-string escape issue) ──
        st.markdown(
            "<div class='snap-sys-row'>" + _sys_pills + "</div>",
            unsafe_allow_html=True)

        # ── Part 3: warn + last run + divider ────────────────────────────────────
        st.markdown(
            _warn_html + _last_html +
            "</div><div class='snap-divider'></div>",
            unsafe_allow_html=True)

        # ── HERO ─────────────────────────────────────────────────────────────────
        st.markdown("""
        <div class="hero-section">

          <div class="hero-inner" style="padding:0 2rem;">
            <div class="eyebrow">Real-time · 4 Odoo Systems · Live Data</div>
            <div class="hero-title">مقارنة <em>المنتجات</em> والمخزون</div>
            <div class="hero-subtitle">Product Comparison Dashboard</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='padding:0 2rem;'>", unsafe_allow_html=True)

        # ── PDF UPLOAD ────────────────────────────────────────────────────────────
        st.markdown(f"<div class='section-tag' style='margin-top:32px;'>{t('Upload Invoice PDF','رفع فاتورة PDF')}</div>",
                    unsafe_allow_html=True)
        p1,p2 = st.columns([2.5,1.5])
        with p1:
            updf = st.file_uploader(t("Upload PDF","رفع PDF"), type=["pdf"],
                                    label_visibility="collapsed")
        with p2:
            emode = None
            if updf:
                emode = st.radio(t("Extract mode","وضع الاستخراج"),
                                 [t("Main models","موديلات رئيسية"),
                                  t("With sizes","مع المقاسات")], horizontal=True)
        if updf:
            fbytes = updf.read()
            fhash  = hashlib.md5(fbytes).hexdigest()
            ck     = f"pdf_{fhash}"
            if ck not in st.session_state:
                with st.spinner(t("Parsing PDF...","جاري قراءة الفاتورة...")):
                    st.session_state[ck] = parse_invoice_pdf_cached(fbytes)
            raw = st.session_state[ck]
            if raw:
                is_main = emode is None or "Main" in emode or "رئيسية" in emode
                unique  = get_unique_base_models(raw) if is_main else list(
                    dict.fromkeys([item["code"] for item in raw]))
                if isinstance(unique[0], dict):
                    unique_sorted = sorted(unique, key=lambda x: x["sequence"])
                    unique_codes  = [item["code"] for item in unique_sorted]
                else:
                    unique_codes  = unique
                    unique_sorted = [{"sequence":i+1,"code":c} for i,c in enumerate(unique_codes)]
                c1,c2 = st.columns(2)
                c1.metric(t("Raw codes","رموز مستخرجة"), len(raw))
                c2.metric(t("Unique models","موديلات فريدة"), len(unique_codes))
                with st.expander(t(f"{len(unique_codes)} codes found","الرموز المستخرجة"), expanded=False):
                    st.code("\n".join(f"{item['sequence']:>3}. {item['code']}"
                                      for item in unique_sorted))
                ca,cb = st.columns(2)
                with ca:
                    if st.button(t("Total Stock","مخزون إجمالي"),
                                 type="primary", use_container_width=True, key="pt"):
                        st.session_state.pdf_codes = unique_codes
                        st.session_state.pdf_mode  = "total"; st.rerun()
                with cb:
                    if st.button(t("Branch-wise","حسب الفرع"),
                                 type="secondary", use_container_width=True, key="pb"):
                        st.session_state.pdf_codes = unique_codes
                        st.session_state.pdf_mode  = "branch"; st.rerun()
            else:
                st.warning(t("No codes found in PDF.","لم يتم العثور على رموز."))

        st.divider()

        # ── MANUAL SEARCH ─────────────────────────────────────────────────────────
        st.markdown(f"<div class='section-tag'>{t('Manual Search','بحث يدوي')}</div>",
                    unsafe_allow_html=True)
        L,R = st.columns([1.5,1])
        with L:
            if not st.session_state.search_exact:
                st.markdown(f"<div class='info-banner'>{t('Variant mode — XP6013 finds all sizes','وضع المتغيرات — XP6013 يجد جميع المقاسات')}</div>",
                            unsafe_allow_html=True)
            else:
                st.markdown(f"<div class='warn-banner'>{t('Exact match — identical codes only','تطابق تام — رموز مطابقة فقط')}</div>",
                            unsafe_allow_html=True)
            ms   = t("Single Model","موديل واحد")
            mm   = t("Multiple Models","موديلات متعددة")
            mode = st.radio(t("Mode","الوضع"),[ms,mm], horizontal=True,
                            label_visibility="collapsed")
            if mode==mm:
                rt    = st.text_area(t("Codes","الرموز"), height=120, placeholder="ABC123\nDEF456")
                codes = [c.strip() for c in rt.replace(",","\n").splitlines() if c.strip()]
            else:
                sg    = st.text_input(t("Model Code","رمز الموديل"), placeholder="e.g. XP6013")
                codes = [sg.strip()] if sg.strip() else []

            t1,t2,t3,t4,t5 = st.columns(5)
            sz  = t1.toggle(t("Zero","الصفري"),     value=False)
            sb  = t2.toggle(t("Branch","فروع"),      value=False)
            ss  = t3.toggle(t("Sort","ترتيب"),       value=False)
            st_ = t4.toggle(t("Transfers","نقليات"), value=False)
            sr  = t5.toggle(t("Reorder","طلب"),      value=False)

            if sr:
                with st.expander(t("Reorder Settings","إعدادات إعادة الطلب"), expanded=True):
                    rx,ry = st.columns(2)
                    with rx:
                        st.session_state.reorder_target_days = st.slider(
                            t("Target days cover","أيام التغطية"), 7, 180,
                            st.session_state.reorder_target_days)
                    with ry:
                        st.session_state.reorder_point = st.number_input(
                            t("Reorder point","نقطة الطلب"), min_value=0, max_value=9999,
                            value=st.session_state.reorder_point, step=1)

            cbtn = st.button(t("Compare →","مقارنة →"), use_container_width=True, type="primary")

        with R:
            st.markdown(f"<div class='section-tag'>{t('Last Run','آخر تشغيل')}</div>",
                        unsafe_allow_html=True)
            snap  = st.session_state.last_run
            stats = st.session_state.sys_stats
            if not snap:
                st.markdown(f"<div class='info-banner'>{t('Run a comparison first.','قم بتشغيل مقارنة أولاً.')}</div>",
                            unsafe_allow_html=True)
            else:
                on = sum(1 for v in stats.values() if v=="OK")
                st.markdown(
                    f"<div class='snap-card'>"
                    f"<b>{t('Time','الوقت')}</b> &nbsp;{snap.get('time','—')}<br>"
                    f"<b>{t('Models','الموديلات')}</b> &nbsp;{snap.get('models','—')}<br>"
                    f"<b>{t('Online','متصل')}</b> &nbsp;{on}/{len(SYSTEM_KEYS)}<br>"
                    f"<b>{t('Rows','الصفوف')}</b> &nbsp;{snap.get('rows','—')}"
                    f"</div>", unsafe_allow_html=True)
                st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
                for key in SYSTEM_KEYS:
                    s  = stats.get(key,"—")
                    bc = "badge-ok" if s=="OK" else "badge-off" if s=="NOT_FOUND" else "badge-err"
                    bt = "Online" if s=="OK" else "Offline" if s=="NOT_FOUND" else "Error"
                    dn = get_system_name(key)
                    st.markdown(
                        f"<div class='sys-row'>"
                        f"<span style='font-family:Outfit,sans-serif;font-size:11px;"
                        f"letter-spacing:1px;color:#6B7280;'>{dn}</span>"
                        f"<span class='{bc}'>{bt}</span></div>",
                        unsafe_allow_html=True)

        # ── TRIGGER RUN ───────────────────────────────────────────────────────────
        run_codes    = None
        force_branch = False
        if st.session_state.get("pdf_codes"):
            run_codes    = st.session_state.pdf_codes
            force_branch = st.session_state.get("pdf_mode","total") == "branch"
            sb = True
            st.session_state.pdf_codes = None
            st.session_state.pdf_mode  = "total"
        elif cbtn:
            run_codes = codes

        if run_codes is not None:
            if not run_codes:
                st.warning(t("Enter at least one model code.","أدخل رمزاً واحداً.")); st.stop()
            run_codes = list(dict.fromkeys([c.strip() for c in run_codes if c.strip()]))
            ct = tuple(run_codes)
            with st.spinner(t("Fetching from 4 systems...","جلب البيانات من 4 أنظمة...")):
                data = fetch_all_data(
                    ct, exact=st.session_state.search_exact,
                    need_branch=sb or force_branch,
                    need_transfers=st_, need_reorder=sr,
                    target_days=st.session_state.reorder_target_days,
                    reorder_point=st.session_state.reorder_point)

            tdf  = prepare_df(data["total"])
            bdf  = prepare_df(data["branch"])
            trdf = prepare_df(data["transfers"])
            rdf  = prepare_df(data["reorder"])

            raw_tdf = data["total"]
            ns = {k:"NOT_FOUND" for k in SYSTEM_KEYS}
            # raw_tdf["System"] holds the RAW KEY (e.g. "SWAG"), not display name
            # because _one() uses sn = key before prepare_df() translates it.
            # We match on key directly.
            if "_status" in raw_tdf.columns and "System" in raw_tdf.columns:
                for key in SYSTEM_KEYS:
                    # Match both raw key AND display name (safety)
                    dn   = get_system_name(key)
                    mask = (raw_tdf["System"] == key) | (raw_tdf["System"] == dn)
                    if mask.any():
                        sv = raw_tdf.loc[mask,"_status"]
                        if   "OK"          in sv.values: ns[key]="OK"
                        elif "NOT_FOUND"   in sv.values: ns[key]="NOT_FOUND"
                        elif "ERROR"       in sv.values: ns[key]="ERROR"

            qc2     = t("On Hand","متوفر")
            sc2_loc = t("System","النظام")
            mc_loc  = t("Model Code","رمز الموديل")

            # Note: DO NOT mark 0 qty as "not_available"
            # 0 qty = product exists but no stock (should show as 0)
            # "not_available" = product doesn't exist in that system

            if ss and sc2_loc in tdf.columns:
                tdf = tdf.sort_values(sc2_loc).reset_index(drop=True)
            if not bdf.empty and ss and sc2_loc in bdf.columns:
                bdf = bdf.sort_values(sc2_loc).reset_index(drop=True)

            # Purchase Qty — fetch for SWAG + STOCK (both have purchase orders)
            # Other systems (LAROUCHE, DIFFC, FASHIONLIMITS) get 0
            _PUR_SYSTEMS = ["SWAG", "STOCK"]
            tdf["Purchase Qty"] = 0

            ed = datetime.now().date(); sd = ed - timedelta(days=365)
            for _pur_key in _PUR_SYSTEMS:
                _pur_name = get_system_name(_pur_key)
                _pur_mask = (tdf[sc2_loc] == _pur_name)
                if not _pur_mask.any():
                    continue
                _pur_models = tdf.loc[_pur_mask, mc_loc].dropna().unique().tolist()
                if not _pur_models:
                    continue
                with st.spinner(t(
                    f"Fetching purchase totals ({_pur_name})...",
                    f"جلب إجمالي المشتريات ({_pur_name})...")):
                    _pur_df = get_purchase_summary_by_model(
                        tuple(_pur_models),
                        sd.strftime("%Y-%m-%d"),
                        ed.strftime("%Y-%m-%d"),
                        system_key=_pur_key)
                if not _pur_df.empty:
                    _pur_df2 = _pur_df.rename(columns={"Model Code": mc_loc})
                    # Merge into tdf — use _pur_tmp to avoid column name clash
                    _tmp = tdf.merge(_pur_df2[[mc_loc,"Purchase Qty"]]
                                     .rename(columns={"Purchase Qty":"_pur_tmp"}),
                                     on=mc_loc, how="left")
                    # Only update rows for this system
                    _fill = _tmp["_pur_tmp"].fillna(0).astype(int)
                    tdf.loc[_pur_mask, "Purchase Qty"] = _fill[_pur_mask].values
                    # Drop temp column if it leaked into tdf
                    tdf = tdf.drop(columns=["_pur_tmp"], errors="ignore")

            pur_col = t("Purchase Qty","كمية المشتريات")
            tdf = tdf.rename(columns={"Purchase Qty":pur_col})
            desired = [sc2_loc,mc_loc,t("Product","المنتج"),
                       t("Sale Price","سعر البيع"),pur_col,qc2]
            existing = tdf.columns.tolist()
            final    = [c for c in desired if c in existing]
            for c in existing:
                # skip any internal temp columns starting with _
                if c not in final and not c.startswith("_"):
                    final.append(c)
            # Always include _status — needed for downstream filtering
            if "_status" in tdf.columns and "_status" not in final:
                final.append("_status")
            tdf = tdf[final]

            st.session_state.total_df       = tdf
            st.session_state.branch_df      = bdf
            st.session_state.transfers_df   = trdf
            st.session_state.reorder_df     = rdf
            st.session_state.show_transfers = st_
            st.session_state.show_reorder   = sr
            st.session_state.sys_stats      = ns
            st.session_state.last_run       = {
                "time"  : datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "models": len(run_codes),
                "rows"  : len(tdf),
            }
            st.rerun()

        # ── RESULTS ───────────────────────────────────────────────────────────────
        tdf  = st.session_state.total_df
        bdf  = st.session_state.branch_df
        trdf = st.session_state.transfers_df
        rdf  = st.session_state.reorder_df
        if tdf is None or tdf.empty:
            st.markdown("</div>", unsafe_allow_html=True)
            return

        st.divider()
        thr   = st.session_state.low_stock_thresh
        qc2   = t("On Hand","متوفر"); pc2 = t("Sale Price","سعر البيع")
        sc2   = t("System","النظام"); stats = st.session_state.sys_stats
        ok    = tdf[tdf["_status"]=="OK"] if "_status" in tdf.columns else tdf
        on    = sum(1 for v in stats.values() if v=="OK")

        # Low stock alert
        if thr>0 and qc2 in ok.columns:
            low = ok[(ok[qc2]>0)&(ok[qc2]<=thr)]
            if not low.empty:
                mc2 = t("Model Code","رمز الموديل")
                det = " · ".join(
                    f"{r.get(mc2,'?')} @ {r.get(sc2,'?')} ({r.get(qc2,0)})"
                    for _,r in low.head(6).iterrows())
                if len(low)>6: det += f" +{len(low)-6}"
                st.markdown(
                    f"<div class='alert-banner'>"
                    f"<div class='alert-dot'></div>"
                    f"<div class='alert-txt'>{t('Low Stock','مخزون منخفض')} — {len(low)} items ≤ {thr}</div>"
                    f"<div class='mono'>{det}</div></div>",
                    unsafe_allow_html=True)

        # Metrics — including stock value calculator
        _ok_qty   = pd.to_numeric(ok[qc2], errors="coerce").fillna(0) if qc2 in ok.columns else pd.Series(dtype=float)
        _ok_price = pd.to_numeric(ok[pc2], errors="coerce").fillna(0) if pc2 in ok.columns else pd.Series(dtype=float)

        # Stock value = qty * price per row, summed
        _stock_value = 0.0
        if qc2 in ok.columns and pc2 in ok.columns:
            _stock_value = (_ok_qty * _ok_price).sum()

        # Per-system stock value
        _sys_values = {}
        if qc2 in ok.columns and pc2 in ok.columns and sc2 in ok.columns:
            for _sys in ok[sc2].dropna().unique():
                _mask = ok[sc2] == _sys
                _sv   = (_ok_qty[_mask] * _ok_price[_mask]).sum()
                _sys_values[_sys] = _sv

        m1,m2,m3,m4,m5,m6 = st.columns(6)
        m1.metric(t("Total Rows","إجمالي الصفوف"), len(tdf))
        m2.metric(t("Systems Online","الأنظمة"), f"{on}/{len(SYSTEM_KEYS)}")
        if qc2 in ok.columns:
            m3.metric(t("Total Qty","إجمالي الكمية"),
                      f"{int(_ok_qty.sum()):,}")
        if pc2 in ok.columns:
            vp = _ok_price[_ok_price>0]
            m4.metric(t("Avg Price (SAR)","متوسط السعر ر.س"),
                      f"{vp.mean():,.0f}" if not vp.empty else "—")
        m5.metric(
            t("Stock Value (SAR)","قيمة المخزون ر.س"),
            f"{_stock_value/1000:,.1f}K" if _stock_value >= 1000 else (f"{_stock_value:,.0f}" if _stock_value > 0 else "—"))
        _zero_val = int((_ok_qty == 0).sum()) if not _ok_qty.empty else 0
        m6.metric(t("Zero Stock Items","أصناف بلا مخزون"), _zero_val)

        hb = bdf  is not None and not bdf.empty
        ht = st.session_state.show_transfers and trdf is not None and not trdf.empty
        hr = st.session_state.show_reorder   and rdf  is not None and not rdf.empty

        # Product tabs
        tlabels = [t("Total Stock","المخزون الإجمالي")]
        if hb: tlabels.append(t("Branch Stock","مخزون الفروع"))
        if ht: tlabels.append(t("Transfers","النقليات"))
        if hr: tlabels.append(t("Reorder","إعادة الطلب"))
        if not _on_season:
            tabs = st.tabs(tlabels); ti = 0
        else:
            tabs = None; ti = 0

        # TAB: TOTAL STOCK
        if not _on_season:
         with tabs[ti]:
            ti += 1

            # ── View toggle ───────────────────────────────────────────────────
            _vt_col, _sz_col = st.columns([3, 1])
            with _vt_col:
                st.markdown(
                    f"<div class='section-tag' style='margin-top:20px;'>"
                    f"{t('Total Stock','المخزون الإجمالي')}</div>",
                    unsafe_allow_html=True)
            with _sz_col:
                st.markdown("<div style='margin-top:18px;'></div>",
                            unsafe_allow_html=True)
                _size_view = st.toggle(
                    t("Size View","عرض الأحجام"),
                    value=False, key="sz_toggle",
                    help=t(
                        "Pivot table: one row per model, sizes as columns (S/M/L/XL/XXL).",
                        "جدول محوري: صف واحد لكل موديل، الأحجام كأعمدة."))

            if not _size_view:
                # ── Normal flat table ─────────────────────────────────────────
                _ft = display_df(tdf, thr, table_key="total")
            else:
                # ── Size Breakdown Pivot ──────────────────────────────────────
                st.markdown(
                    f"<div class='info-banner'>"
                    f"{t('Pivot view — one row per base model × system. Sizes as columns. '
                         'Red = 0 stock, Amber = low stock, Teal = OK.',
                         'عرض محوري — صف واحد لكل موديل × نظام. الأحجام كأعمدة. '
                         'أحمر = لا مخزون، عنبر = مخزون منخفض، تيل = كافٍ.')}"
                    f"</div>", unsafe_allow_html=True)

                _sz_qc = t("On Hand","متوفر")
                _sz_mc = t("Model Code","رمز الموديل")
                _sz_sc = t("System","النظام")
                _sz_pc = t("Sale Price","سعر البيع")

                # Use filtered df if available, else full tdf
                _sz_source = tdf.copy()
                # Apply same OK filter
                if "_status" in _sz_source.columns:
                    _sz_source = _sz_source[_sz_source["_status"]=="OK"].copy()
                # Numeric qty
                if _sz_qc in _sz_source.columns:
                    _sz_source[_sz_qc] = pd.to_numeric(
                        _sz_source[_sz_qc], errors="coerce").fillna(0)

                _pivot_df, _size_cols = build_size_pivot(
                    _sz_source, _sz_mc, _sz_qc, _sz_sc, _sz_pc, thr=thr)

                if _pivot_df is None or _pivot_df.empty:
                    st.markdown(
                        f"<div class='warn-banner'>"
                        f"{t('No size suffixes found in model codes (e.g. XP6013-M). '
                             'Size View works when model codes end with -S/-M/-L/-XL/-XXL etc.',
                             'لم يتم العثور على لاحقات أحجام في رموز الموديل (مثال: XP6013-M). '
                             'يعمل عرض الأحجام عندما تنتهي رموز الموديل بـ -S/-M/-L/-XL/-XXL إلخ.')}"
                        f"</div>", unsafe_allow_html=True)
                    # Fallback to normal view
                    _ft = display_df(tdf, thr, table_key="total")
                else:
                    # Summary metrics for size view
                    _sz_total = int(_pivot_df["Total"].sum()) if "Total" in _pivot_df.columns else 0
                    _base_col = t("Base Model","الموديل الأساسي")
                    _sz_models = _pivot_df[_base_col].nunique() if _base_col in _pivot_df.columns else 0
                    _sz_zero   = sum(
                        int((_pivot_df[s] == 0).sum())
                        for s in _size_cols if s in _pivot_df.columns)

                    _sm1,_sm2,_sm3,_sm4 = st.columns(4)
                    _sm1.metric(t("Base Models","الموديلات"), _sz_models)
                    _sm2.metric(t("Total Units","إجمالي الوحدات"), f"{_sz_total:,}")
                    _sm3.metric(t("Sizes Found","الأحجام"), len(_size_cols))
                    _sm4.metric(t("Zero-Size Slots","خانات فارغة"), f"{_sz_zero:,}")

                    # Optional system filter
                    if _sz_sc in _pivot_df.columns:
                        _sys_opts = sorted(_pivot_df[_sz_sc].dropna().unique().tolist())
                        if len(_sys_opts) > 1:
                            _sel_sys_sz = st.multiselect(
                                t("Filter by System","فلتر حسب النظام"),
                                options=_sys_opts, default=_sys_opts,
                                key="sz_sys_filter")
                            if _sel_sys_sz:
                                _pivot_df = _pivot_df[
                                    _pivot_df[_sz_sc].isin(_sel_sys_sz)]

                    # Model search
                    _sz_search = st.text_input(
                        t("Search base model","بحث موديل أساسي"),
                        placeholder="e.g. XP6013", key="sz_search").strip().upper()
                    if _sz_search:
                        _bc = t("Base Model","الموديل الأساسي")
                        if _bc in _pivot_df.columns:
                            _pivot_df = _pivot_df[
                                _pivot_df[_bc].str.upper().str.contains(
                                    _sz_search, regex=False, na=False)]

                    st.caption(
                        f"{len(_pivot_df)} {t('models','موديل')} · "
                        f"{len(_size_cols)} {t('sizes','حجم')} · "
                        f"{t('Red=0 · Amber=Low · Teal=OK','أحمر=صفر · عنبر=منخفض · تيل=كافٍ')}")

                    render_size_pivot(_pivot_df, _size_cols, thr=thr)

                    # Excel export for size view
                    st.markdown("<br>", unsafe_allow_html=True)
                    _sz_ex1, _sz_ex2 = st.columns([1, 3])
                    _sz_ex1.download_button(
                        t("Size View Excel ↓","Excel عرض الأحجام ↓"),
                        _excel_generic(
                            _pivot_df.fillna(0),
                            t("Size Breakdown","تفصيل الأحجام")),
                        dl_name("size_breakdown","xlsx"),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="sz_excel_dl")

                _ft = None   # no filtered df in size view

            # ── STOCK VALUE BREAKDOWN ─────────────────────────────────────────
            st.markdown(f"<div class='section-tag'>{t('Stock Value by System','قيمة المخزون حسب النظام')}</div>",
                        unsafe_allow_html=True)

            _qc = t("On Hand","متوفر"); _pc = t("Sale Price","سعر البيع"); _sc = t("System","النظام")
            _mc = t("Model Code","رمز الموديل"); _prc = t("Product","المنتج")

            if _qc in tdf.columns and _pc in tdf.columns:
                _ok2  = tdf[tdf["_status"]=="OK"].copy() if "_status" in tdf.columns else tdf.copy()
                _ok2["_qty"]   = pd.to_numeric(_ok2[_qc], errors="coerce").fillna(0)
                _ok2["_price"] = pd.to_numeric(_ok2[_pc], errors="coerce").fillna(0)
                _ok2["_value"] = _ok2["_qty"] * _ok2["_price"]

                # ── per-system value cards ────────────────────────────────────
                if _sc in _ok2.columns:
                    _sys_list = sorted(_ok2[_sc].dropna().unique().tolist())
                    _cols = st.columns(len(_sys_list)) if _sys_list else []
                    for _i, _sn in enumerate(_sys_list):
                        _smask  = _ok2[_sc] == _sn
                        _sval   = _ok2.loc[_smask, "_value"].sum()
                        _sqty   = int(_ok2.loc[_smask, "_qty"].sum())
                        _scount = _smask.sum()
                        _cols[_i].markdown(f"""
                        <div style='background:#EEF9FA;border:1.5px solid #C5E3E5;
                                    border-radius:10px;padding:16px;text-align:center;'>
                          <div style='font-family:Outfit,sans-serif;font-size:8px;letter-spacing:3px;
                                      text-transform:uppercase;color:#1A7A82;margin-bottom:8px;'>{_sn}</div>
                          <div style='font-family:"Cormorant Garamond",serif;font-size:28px;font-weight:300;
                                      color:#111827;line-height:1;margin-bottom:4px;'>
                            {_sval:,.0f}
                          </div>
                          <div style='font-family:Outfit,sans-serif;font-size:9px;letter-spacing:2px;
                                      color:#9CA3AF;margin-bottom:8px;'>SAR</div>
                          <div style='display:flex;justify-content:center;gap:12px;'>
                            <div style='font-family:Outfit,sans-serif;font-size:9px;color:#9CA3AF;'>
                              {_sqty:,} {t("units","وحدة")}
                            </div>
                            <div style='font-family:Outfit,sans-serif;font-size:9px;color:#9CA3AF;'>
                              {_scount} {t("SKUs","صنف")}
                            </div>
                          </div>
                        </div>""", unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)

                # ── top 10 models by value ────────────────────────────────────
                st.markdown(f"<div class='section-tag'>{t('Top 10 Models by Stock Value','أعلى 10 موديلات بقيمة المخزون')}</div>",
                            unsafe_allow_html=True)

                _top_cols = [c for c in [_mc, _prc, _sc, "_qty", "_price", "_value"]
                             if c in _ok2.columns]
                _top = (_ok2[_ok2["_qty"]>0][_top_cols]
                        .sort_values("_value", ascending=False)
                        .head(10)
                        .reset_index(drop=True))

                if not _top.empty:
                    _display_top = _top.copy()
                    _display_top["_qty"]   = _display_top["_qty"].astype(int).map(lambda v: f"{v:,}")
                    _display_top["_price"] = _display_top["_price"].map(lambda v: f"{v:.2f} SAR")
                    _display_top["_value"] = _display_top["_value"].map(lambda v: f"{v:,.0f} SAR")
                    _display_top = _display_top.rename(columns={
                        "_qty"  : t("Qty","الكمية"),
                        "_price": t("Unit Price","سعر الوحدة"),
                        "_value": t("Stock Value","قيمة المخزون"),
                    })
                    # remove internal cols
                    _display_top = _display_top[[c for c in _display_top.columns
                                                 if not c.startswith("_")]]

                    # render as html table
                    _cols_t = _display_top.columns.tolist()
                    _th     = "".join(f"<th>{c}</th>" for c in _cols_t)
                    def _tr(ir):
                        _, row = ir
                        cells = "".join(
                            f'<td class="cf">{v}</td>' if ci==0 else
                            (f'<td style="color:#D4A84B;font-family:Outfit,monospace;">{v}</td>'
                             if ci == len(row)-1 else f"<td>{v}</td>")
                            for ci,v in enumerate(row))
                        return f"<tr>{cells}</tr>"
                    _tbody = "".join(_tr(x) for x in _display_top.iterrows())
                    _TABLE_CSS2 = """<style>
    .swag-wrap{width:100%;overflow-x:auto;border:1px solid rgba(74,172,180,0.08);border-radius:4px;overflow:hidden;margin-bottom:4px;}
    .swag-tbl{width:100%;border-collapse:collapse;font-family:'Outfit','Tajawal',sans-serif;}
    .swag-tbl thead tr{background:#EEF9FA;border-bottom:1px solid rgba(74,172,180,0.1);}
    .swag-tbl thead th{color:#1A7A82;font-family:'Outfit',sans-serif;font-size:8px;letter-spacing:3px;text-transform:uppercase;font-weight:400;padding:13px 16px;text-align:center;white-space:nowrap;}
    .swag-tbl tbody tr{border-bottom:1px solid #F3F4F6;transition:background 0.15s;}
    .swag-tbl tbody tr:hover td{background:#EEF9FA;}
    .swag-tbl tbody td{padding:12px 16px;text-align:center;font-size:12px;color:#111827;}
    .swag-tbl tbody td.cf{font-family:'Outfit',monospace;font-size:11px;letter-spacing:0.5px;color:#111827;font-weight:500;border-right:1px solid rgba(74,172,180,0.08);}
    </style>"""
                    st.markdown(
                        f'{_TABLE_CSS2}<div class="swag-wrap">'
                        f'<table class="swag-tbl"><thead><tr>{_th}</tr></thead>'
                        f'<tbody>{_tbody}</tbody></table></div>',
                        unsafe_allow_html=True)

                # ── total value summary bar ───────────────────────────────────
                _total_val  = _ok2["_value"].sum()
                _zero_val2  = int((_ok2["_qty"]==0).sum())
                _avail_val  = _ok2.loc[_ok2["_qty"]>0,"_value"].sum()

                st.markdown(f"""
                <div style='background:rgba(212,168,75,0.06);border:1px solid rgba(212,168,75,0.2);
                            border-radius:10px;padding:20px 24px;margin-top:16px;
                            display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:16px;'>
                  <div>
                    <div style='font-family:Outfit,sans-serif;font-size:8px;letter-spacing:4px;
                                text-transform:uppercase;color:#D4A84B;margin-bottom:6px;'>
                      {t("Total Portfolio Value","إجمالي قيمة المحفظة")}
                    </div>
                    <div style='font-family:"Cormorant Garamond",serif;font-size:42px;
                                font-weight:300;color:#111827;line-height:1;'>
                      {_total_val:,.0f}
                      <span style='font-size:18px;color:#D4A84B;letter-spacing:2px;'> SAR</span>
                    </div>
                  </div>
                  <div style='display:flex;gap:28px;flex-wrap:wrap;'>
                    <div style='text-align:center;'>
                      <div style='font-family:"Cormorant Garamond",serif;font-size:24px;
                                  font-weight:300;color:#1A7A82;'>{_avail_val:,.0f}</div>
                      <div style='font-family:Outfit,sans-serif;font-size:8px;letter-spacing:2px;
                                  text-transform:uppercase;color:#9CA3AF;margin-top:2px;'>
                        {t("In-Stock Value","قيمة المتوفر")} SAR
                      </div>
                    </div>
                    <div style='text-align:center;'>
                      <div style='font-family:"Cormorant Garamond",serif;font-size:24px;
                                  font-weight:300;color:#D4A84B;'>{_zero_val2}</div>
                      <div style='font-family:Outfit,sans-serif;font-size:8px;letter-spacing:2px;
                                  text-transform:uppercase;color:#9CA3AF;margin-top:2px;'>
                        {t("Zero-Stock SKUs","أصناف بلا مخزون")}
                      </div>
                    </div>
                    <div style='text-align:center;'>
                      <div style='font-family:"Cormorant Garamond",serif;font-size:24px;
                                  font-weight:300;color:#374151;'>
                        {int(_ok2.loc[_ok2["_qty"]>0,"_qty"].sum()):,}
                      </div>
                      <div style='font-family:Outfit,sans-serif;font-size:8px;letter-spacing:2px;
                                  text-transform:uppercase;color:#9CA3AF;margin-top:2px;'>
                        {t("Total Units","إجمالي الوحدات")}
                      </div>
                    </div>
                  </div>
                </div>""", unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            d1,d2,d3,d4 = st.columns(4)
            d1.download_button("CSV ↓", to_csv(tdf), dl_name("total","csv"),
                               "text/csv", use_container_width=True)
            d2.download_button("Excel ↓", to_excel(tdf), dl_name("total","xlsx"),
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
            d3.download_button(t("All Systems ↓","كل الأنظمة ↓"), to_excel_bulk(tdf),
                               dl_name("bulk","xlsx"),
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
            if _ft is not None and not _ft.empty:
                d4.download_button(t("Filtered ↓","مفلتر ↓"), to_excel(_ft),
                                   dl_name("filtered","xlsx"),
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)

            # ── WhatsApp Share ────────────────────────────────────────────────
            st.divider()
            st.markdown(
                f"<div class='section-tag'>{t('WhatsApp Share','مشاركة واتساب')}</div>",
                unsafe_allow_html=True)

            # Build formatted message from tdf
            _wa_src = (_ft if _ft is not None and not _ft.empty else tdf).copy()
            _wa_src  = _wa_src[_wa_src.get("_status","OK") != "ERROR"] if "_status" in _wa_src.columns else _wa_src
            _wa_src  = _wa_src.drop(columns=["_status"], errors="ignore")

            _wa_qc  = t("On Hand","متوفر")
            _wa_mc  = t("Model Code","رمز الموديل")
            _wa_sc  = t("System","النظام")
            _wa_pc  = t("Sale Price","سعر البيع")
            _wa_pur = t("Purchase Qty","كمية المشتريات")

            def _build_wa_msg(df):
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                lines   = []
                lines.append(f"📦 *SWAG Stock Report*")
                lines.append(f"🕒 {now_str}")
                lines.append("")

                # Group by model code
                mc_col  = _wa_mc  if _wa_mc  in df.columns else (df.columns[1] if len(df.columns)>1 else None)
                sys_col = _wa_sc  if _wa_sc  in df.columns else None
                qty_col = _wa_qc  if _wa_qc  in df.columns else None
                prc_col = _wa_pc  if _wa_pc  in df.columns else None
                pur_col = _wa_pur if _wa_pur in df.columns else None

                if mc_col is None:
                    return t("No data to share.","لا توجد بيانات للمشاركة.")

                models = df[mc_col].dropna().unique().tolist()[:20]  # max 20 models
                for model in models:
                    mask  = df[mc_col] == model
                    rows  = df[mask]
                    lines.append(f"*{model}*")

                    for _, row in rows.iterrows():
                        sys_nm = str(row.get(sys_col,"")).strip() if sys_col else ""
                        qty    = row.get(qty_col, 0) if qty_col else 0
                        price  = row.get(prc_col, 0) if prc_col else 0

                        try: qty_v = int(float(qty))
                        except Exception: qty_v = 0
                        try: prc_v = float(price)
                        except Exception: prc_v = 0.0

                        qty_emoji = "🔴" if qty_v == 0 else ("🟡" if qty_v <= 5 else "🟢")
                        parts = [f"  {qty_emoji} {sys_nm}" if sys_nm else f"  {qty_emoji}"]
                        parts.append(f"Qty: {qty_v:,}")
                        if prc_v > 0:
                            parts.append(f"Price: {prc_v:.0f} SAR")
                        if pur_col and pur_col in row:
                            try:
                                pur_v = int(float(row[pur_col]))
                                if pur_v > 0:
                                    parts.append(f"Purchased: {pur_v:,}")
                            except Exception:
                                pass
                        lines.append(" | ".join(parts))

                    lines.append("")

                if len(models) == 20 and len(df[mc_col].dropna().unique()) > 20:
                    lines.append(f"_...and {len(df[mc_col].dropna().unique())-20} more models_")
                    lines.append("")

                # Summary footer
                if qty_col:
                    _tot_qty = int(pd.to_numeric(df[qty_col], errors="coerce").fillna(0).sum())
                    lines.append(f"📊 *Total Qty: {_tot_qty:,}*")
                if prc_col and qty_col:
                    _tot_val = (
                        pd.to_numeric(df[qty_col], errors="coerce").fillna(0) *
                        pd.to_numeric(df[prc_col], errors="coerce").fillna(0)
                    ).sum()
                    lines.append(f"💰 *Stock Value: {_tot_val:,.0f} SAR*")

                lines.append("")
                lines.append("_Powered by SWAG Dashboard_")
                return "\n".join(lines)

            _wa_msg = _build_wa_msg(_wa_src)

            # Show preview + copy button
            _wac1, _wac2 = st.columns([2.5, 1])
            with _wac1:
                st.text_area(
                    t("Message Preview (copy & paste to WhatsApp)",
                      "معاينة الرسالة (انسخ والصق في واتساب)"),
                    value=_wa_msg,
                    height=200,
                    key="wa_preview",
                    help=t(
                        "Select all text → copy → paste in WhatsApp",
                        "حدد كل النص ← انسخ ← الصق في واتساب"))
            with _wac2:
                st.markdown("<br><br>", unsafe_allow_html=True)

                # Direct WhatsApp link (mobile friendly)
                import urllib.parse as _urlparse
                _wa_encoded = _urlparse.quote(_wa_msg)
                _wa_url     = f"https://wa.me/?text={_wa_encoded}"

                st.markdown(f"""
                <a href="{_wa_url}" target="_blank" rel="noopener"
                   style='display:block;width:100%;padding:12px 0;
                          background:#25D366;border:none;border-radius:100px;
                          font-family:Outfit,sans-serif;font-size:10px;font-weight:600;
                          letter-spacing:2px;text-transform:uppercase;color:#111827;
                          text-align:center;text-decoration:none;
                          transition:background 0.2s;'>
                  WhatsApp →
                </a>""", unsafe_allow_html=True)

                st.markdown("<div style='margin-top:8px;'></div>",
                            unsafe_allow_html=True)

                # Plain text download as fallback
                st.download_button(
                    t("Download .txt ↓","تحميل .txt ↓"),
                    _wa_msg.encode("utf-8"),
                    dl_name("stock_report","txt"),
                    "text/plain",
                    use_container_width=True,
                    key="wa_txt_dl")



        # TAB: BRANCH STOCK
        if not _on_season and hb:
            with tabs[ti]:
                ti += 1
                st.markdown(f"<div class='section-tag' style='margin-top:20px;'>{t('Branch-wise Stock','مخزون حسب الفرع')}</div>",
                            unsafe_allow_html=True)
                _fb = display_df(bdf, thr, table_key="branch")
                bc2 = t("Branch","الفرع")
                okb = bdf[bdf["_status"]=="OK"] if "_status" in bdf.columns else bdf
                if not okb.empty and bc2 in okb.columns and qc2 in okb.columns:
                    chart = okb.groupby([sc2,bc2])[qc2].sum().reset_index()
                    if not chart.empty:
                        st.markdown(f"<div class='section-tag'>{t('Qty by Branch','الكميات حسب الفرع')}</div>",
                                    unsafe_allow_html=True)
                        st.bar_chart(chart.set_index(bc2)[qc2], use_container_width=True)
                b1,b2,b3,b4 = st.columns(4)
                b1.download_button("CSV ↓", to_csv(bdf), dl_name("branch","csv"),
                                   "text/csv", use_container_width=True)
                b2.download_button("Excel ↓", to_excel(bdf), dl_name("branch","xlsx"),
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)
                if _fb is not None and not _fb.empty:
                    b3.download_button(t("Filtered ↓","مفلتر ↓"), to_excel(_fb),
                                       dl_name("filtered_branch","xlsx"),
                                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       use_container_width=True)
                    b4.download_button(t("Matrix ↓","مصفوفة ↓"),
                                       to_excel_branch_matrix(_fb, get_lang()),
                                       dl_name("matrix","xlsx"),
                                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       use_container_width=True)

        # TAB: TRANSFERS
        if not _on_season and ht:
            with tabs[ti]:
                ti += 1
                st.markdown(f"<div class='section-tag' style='margin-top:20px;'>{t('Pending Transfers','النقليات المعلقة')}</div>",
                            unsafe_allow_html=True)
                okt = trdf[trdf["_status"]=="OK"] if "_status" in trdf.columns else trdf
                if not okt.empty:
                    k1,k2,k3 = st.columns(3)
                    k1.metric(t("Total","إجمالي"), len(okt))
                    qd = t("Qty","الكمية")
                    if qd in okt.columns: k2.metric(t("Total Qty","إجمالي الكمية"), int(okt[qd].sum()))
                    if sc2 in okt.columns: k3.metric(t("Systems","الأنظمة"), okt[sc2].nunique())
                display_df(trdf, thresh=0, table_key="transfers")
                x1,x2 = st.columns([1,1])
                x1.download_button("CSV ↓", to_csv(trdf), dl_name("transfers","csv"),
                                   "text/csv", use_container_width=True)
                x2.download_button("Excel ↓", to_excel(trdf), dl_name("transfers","xlsx"),
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)

        # TAB: REORDER
        if not _on_season and hr:
            with tabs[ti]:
                ti += 1
                CPRI  = t("Priority","الأولوية"); CSUGG = t("Suggest","المقترح")
                st.markdown(f"<div class='section-tag' style='margin-top:20px;'>{t('Reorder Suggestions','اقتراحات إعادة الطلب')}</div>",
                            unsafe_allow_html=True)
                okr = rdf[rdf["_status"]=="OK"] if "_status" in rdf.columns else rdf
                if not okr.empty:
                    crit = okr[okr[CPRI].str.contains("Critical",na=False)].shape[0] if CPRI in okr.columns else 0
                    lo   = okr[okr[CPRI].str.contains("Low",na=False)].shape[0]      if CPRI in okr.columns else 0
                    okn  = okr[okr[CPRI].str.contains("OK",na=False)].shape[0]       if CPRI in okr.columns else 0
                    sg   = int(okr[CSUGG].sum()) if CSUGG in okr.columns else 0
                    r1,r2,r3,r4 = st.columns(4)
                    r1.metric(t("Critical","حرج"), crit)
                    r2.metric(t("Low","منخفض"), lo)
                    r3.metric(t("OK","كافٍ"), okn)
                    r4.metric(t("To Order","للطلب"), sg)
                    if crit+lo>0:
                        st.markdown(
                            f"<div class='warn-banner'>{crit+lo} {t('products need reordering','منتجات تحتاج إعادة طلب')}</div>",
                            unsafe_allow_html=True)
                    sa = st.toggle(t("Show all","عرض الكل"), value=False)
                    dr = (okr if sa else
                          okr[okr[CPRI].str.contains("Critical|Low",na=False)] if CPRI in okr.columns else okr)
                    display_df(dr.reset_index(drop=True), table_key="reorder")
                o1,o2 = st.columns([1,1])
                o1.download_button("CSV ↓", to_csv(rdf), dl_name("reorder","csv"),
                                   "text/csv", use_container_width=True)
                o2.download_button("Excel ↓", to_excel(rdf), dl_name("reorder","xlsx"),
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)

    # TAB: SEASON COMPARISON
# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
restore_session()
if not st.session_state.authenticated:
    show_login()
else:
    show_dashboard()
